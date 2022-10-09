from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from fpdf import FPDF
import uuid
import imghdr
from email.message import EmailMessage
from django.contrib import messages
from django.contrib.auth.models import User
import warnings
from datetime import datetime
warnings.filterwarnings("ignore")
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, Frame, SimpleDocTemplate
from reportlab.lib.styles import getSampleStyleSheet
from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter
from Assessment.models import Master
from Members.models import Profile
from PyPDF2 import PdfMerger
from openpyxl import load_workbook

from django.conf import settings
from django.core.mail import send_mail

from PyPDF2 import PdfFileWriter, PdfFileReader
import io



# Create your views here.
@login_required
def home(request):
    if request.method == 'GET':
        session_dict = get_session_data()
        count = 1
        context = {}
        session_data = False
        for i in session_dict:
            if request.session['title'] == session_dict.get('session_dict_{}'.format(count)).get('title') \
                    and request.session['date'] == session_dict.get('session_dict_{}'.format(count)).get('date'):
                context['value_dict'] = session_dict.get('session_dict_{}'.format(count))
                session_data = True
            count += 1
        if not session_data:
            context['value_dict'] = 0
        context['session_dict'] = session_dict
        return render(request, 'index.html', context)

    if request.method == 'POST':
        request.session['title'] = request.POST.get('title')
        request.session['author'] = request.POST.get('author')
        request.session['role'] = request.POST.get('role')
        request.session['department'] = request.POST.get('department')
        request.session['manager'] = request.POST.get('manager')
        request.session['status'] = request.POST.get('status')
        request.session['date'] = request.POST.get('date')
        title = request.session.get('title')
        author = request.session.get('author')
        role = request.session.get('role')
        department = request.session.get('department')
        manager = request.session.get('manager')
        status = request.session.get('status')
        date = request.session.get('date')

        input_data = {
            'title': title,
            'author': author,
            'department': department,
            'role': role,
            'manager': manager,
            'status': status,
            'date': date,
        }
        username = request.session.get('username')
        session_dict = get_session_data()
        count = 1
        context = {}
        session_data = False
        for i in session_dict:
            if (title == session_dict.get('session_dict_{}'.format(count)).get('title')) and (
                    status == session_dict.get('session_dict_{}'.format(count)).get('status') and
                    date == session_dict.get('session_dict_{}'.format(count)).get('date')
            ):
                session_data = True
                context['value_dict'] = session_dict.get('session_dict_{}'.format(count))
                request.session['value_dict_final'] = session_dict.get('session_dict_{}'.format(count))
            count+=1
        if not session_data:
            context['value_dict'] = 0
        context['input_data'] = input_data
        context['session_dict'] = session_dict
        return render(request, 'screening.html', context)
    session_dict = get_session_data()
    context = {
        'authorized': True,
        'session_dict': session_dict
    }
    return render(request, 'index.html', context)


def no_session(request):
    context = {}
    if request.method == "GET":
        db_dict_num = request.GET.get('search')
        if db_dict_num == '0':
            context['value_dict'] = 0
        else:
            session_dict = get_session_data()
            for data in range(1, len(session_dict) + 1):
                if db_dict_num == session_dict.get('session_dict_{}'.format(data)).get('title'):
                    context['value_dict'] = session_dict.get('session_dict_{}'.format(data))
            context['session_dict'] = session_dict
            request.session['value_dict_final'] = session_dict
            return render(request, 'index.html', context)
    return render(request, 'index.html', context)


def get_session_data():
    session_data = Master.objects.all()
    dict_count = 1
    session_dict = {}
    for items in session_data:
        session_dict['session_dict_{}'.format(dict_count)] = items.__dict__
        session_dict.get('session_dict_{}'.format(dict_count))['_state'] = \
            str(session_dict.get('session_dict_{}'.format(dict_count))['_state'])
        session_dict.get('session_dict_{}'.format(dict_count))['date'] = \
            session_dict.get('session_dict_{}'.format(dict_count))['date'].strftime("%Y-%m-%d")
        dict_count+=1
    return session_dict


@login_required
def temp(request):
    return render(request, 'temp.html')


@login_required
def session_screen(request):
    return render(request, 'index.html')


@login_required
def risk_summary(request):
    if request.method == 'GET':
        session_dict = get_session_data()
        count = 1
        value = request.GET.get('search')
        context = {}
        session_data = False
        for i in session_dict:
            if request.session['title'] == session_dict.get('session_dict_{}'.format(count)).get('title') \
                    and request.session['date'] == session_dict.get('session_dict_{}'.format(count)).get('date'):
                context['value_dict'] = session_dict.get('session_dict_{}'.format(count))
                context['approve_value'] = value
                context['risk_score1'] = request.session.get('risk_score1')
                context['risk_score2'] = request.session.get('risk_score2')
                context['risk_score3'] = request.session.get('risk_score3')
                context['risk_score4'] = request.session.get('risk_score4')
                context['risk_score5'] = request.session.get('risk_score5')
                context['risk_score6'] = request.session.get('risk_score6')
                context['risk_score7'] = request.session.get('risk_score7')
                context['risk_score8'] = request.session.get('risk_score8')
                context['form1_percentage'] = request.session.get('form1_percentage')
                context['form2_percentage'] = request.session.get('form2_percentage')
                context['form3_percentage'] = request.session.get('form3_percentage')
                context['form4_percentage'] = request.session.get('form4_percentage')
                context['form5_percentage'] = request.session.get('form5_percentage')
                context['form6_percentage'] = request.session.get('form6_percentage')
                context['form7_percentage'] = request.session.get('form7_percentage')
                context['form8_percentage'] = request.session.get('form8_percentage')
                context['total_no_of_risk'] = request.session.get('total_no_of_risk')
                session_data = True
            count += 1
        if not session_data:
            context['value_dict'] = 0
        context['session_dict'] = session_dict
        return render(request, 'risk_summary.html', context)
    return render(request, 'risk_summary.html')


@login_required
def heat_map(request):
    return render(request, 'heat_map.html')


@login_required
def result(request):
    return render(request, 'result.html')


@login_required
def gdpr_report(request):
    if request.method == 'POST':
        print('Enter GDPR report')
        text_box_1 = request.POST.get('text_box_1')
        text_box_2 = request.POST.get('text_box_2')
        text_box_2_2 = request.POST.get('text_box_2_2')
        text_box_2_3 = request.POST.get('text_box_2_3')
        text_box_2_4 = request.POST.get('text_box_2_4')
        text_box_3 = request.POST.get('text_box_3')
        text_box_4 = request.POST.get('text_box_4')
        text_box_5 = request.POST.get('text_box_5')
        text_box_6 = request.POST.get('text_box_6')
        text_box_7 = request.POST.get('text_box_7')
        text_box_8 = request.POST.get('text_box_8')

        form2_f1_1 = request.POST.get('form2_f1_1')
        form2_f1_2 = request.POST.get('form2_f1_2')
        form2_f1_3 = request.POST.get('form2_f1_3')

        form2_f2_1 = request.POST.get('form2_f2_1')
        form2_f2_2 = request.POST.get('form2_f2_2')
        form2_f2_3 = request.POST.get('form2_f2_3')

        form2_f3_1 = request.POST.get('form2_f3_1')
        form2_f3_2 = request.POST.get('form2_f3_2')
        form2_f3_3 = request.POST.get('form2_f3_3')
        form2_f3_4 = request.POST.get('form2_f3_4')
        form2_f3_5 = request.POST.get('form2_f3_5')
        form2_f3_6 = request.POST.get('form2_f3_6')
        form2_f3_7 = request.POST.get('form2_f3_7')

        request.session['form2_f1_1'] = form2_f1_1
        request.session['form2_f1_2'] = form2_f1_2
        request.session['form2_f1_3'] = form2_f1_3

        request.session['form2_f2_1'] = form2_f2_1
        request.session['form2_f2_2'] = form2_f2_2
        request.session['form2_f2_3'] = form2_f2_3

        request.session['form2_f3_1'] = form2_f3_1
        request.session['form2_f3_2'] = form2_f3_2
        request.session['form2_f3_3'] = form2_f3_3
        request.session['form2_f3_4'] = form2_f3_4
        request.session['form2_f3_5'] = form2_f3_5
        request.session['form2_f3_6'] = form2_f3_6
        request.session['form2_f3_7'] = form2_f3_7

        box2_section2_text1 = request.session.get('form2_f1_1') + '.' + ' ' + request.session.get('form2_f1_2') + '.' + ' ' + request.session.get('form2_f1_3') + '.'
        request.session['box2_section2_text1'] = box2_section2_text1

        box2_section2_text2 = request.session.get('form2_f2_1') + '.' + ' ' + request.session.get(
            'form2_f2_2') + '.' + ' ' + request.session.get('form2_f2_3') + '.'
        request.session['box2_section2_text2'] = box2_section2_text2

        box2_section2_text3 = request.session.get('form2_f3_1') + '.' + ' ' + request.session.get(
            'form2_f3_2') + '.' + ' ' + request.session.get('form2_f3_3') + '.' + ' ' + request.session.get('form2_f3_4') + '.'+ ' ' + request.session.get('form2_f3_5') + '.'+ ' ' + request.session.get('form2_f3_6') + '.'+ ' ' + request.session.get('form2_f3_7') + '.'
        request.session['box2_section2_text3'] = box2_section2_text3

        request.session['text_box_2_2'] = text_box_2_2
        request.session['text_box_2_3'] = text_box_2_3
        request.session['text_box_2_4'] = text_box_2_4

        request.session['f1_final_text'] = text_box_1
        request.session['f2_final_text'] = text_box_2
        request.session['f3_final_text'] = text_box_3
        request.session['f4_final_text'] = text_box_4
        request.session['f5_final_text'] = text_box_5
        request.session['f6_final_text'] = text_box_6
        request.session['f7_final_text'] = text_box_7
        request.session['f8_final_text'] = text_box_8

        m_approved_by = request.POST.get('m_approved_by')
        r_approved_by = request.POST.get('r_approved_by')
        dpo_advice = request.POST.get('dpo_advice')
        summary_dpo_advice = request.POST.get('summary_dpo_advice')
        dpo_advice_or_overruled = request.POST.get('dpo_advice_or_overruled')
        response_reviewed_by = request.POST.get('response_reviewed_by')
        under_review_by = request.POST.get('under_review_by')
        comments1 = request.POST.get('comments1')
        comments2 = request.POST.get('comments2')
        dpo_comments = request.POST.get('dpo_comments')
        do_comments = request.POST.get('do_comments')
        request.session['dpo_comments'] = dpo_comments
        request.session['do_comments'] = do_comments

        # form5_f1 = request.POST.get('form5_f1')
        # request.session['form5_f1'] = form5_f1
        # form5_f1_1 = request.POST.get('form5_f1_1')
        # form5_f1_2 = request.POST.get('form5_f1_2')
        # form5_f1_3 = request.POST.get('form5_f1_3')

        form5_f2 = request.POST.get('form5_f2')
        form5_f2_1 = request.POST.get('form5_f2_1')
        form5_f2_2 = request.POST.get('form5_f2_2')
        form5_f2_3 = request.POST.get('form5_f2_3')

        form5_f3 = request.POST.get('form5_f3')
        form5_f3_1 = request.POST.get('form5_f3_1')
        form5_f3_2 = request.POST.get('form5_f3_2')
        form5_f3_3 = request.POST.get('form5_f3_3')

        form5_f4 = request.POST.get('form5_f4')
        form5_f4_1 = request.POST.get('form5_f4_1')
        form5_f4_2 = request.POST.get('form5_f4_2')
        form5_f4_3 = request.POST.get('form5_f4_3')

        form5_f5 = request.POST.get('form5_f5')
        form5_f5_1 = request.POST.get('form5_f5_1')
        form5_f5_2 = request.POST.get('form5_f5_2')
        form5_f5_3 = request.POST.get('form5_f5_3')

        form6_f1 = request.POST.get('form6_f1')
        form6_f1_1 = request.POST.get('form6_f1_1')
        form6_f1_2 = request.POST.get('form6_f1_2')
        form6_f1_3 = request.POST.get('form6_f1_3')
        form6_f1_4 = request.POST.get('form6_f1_4')

        form6_f2 = request.POST.get('form6_f2')
        form6_f2_1 = request.POST.get('form6_f2_1')
        form6_f2_2 = request.POST.get('form6_f2_2')
        form6_f2_3 = request.POST.get('form6_f2_3')
        form6_f2_4 = request.POST.get('form6_f2_4')

        form6_f3 = request.POST.get('form6_f3')
        form6_f3_1 = request.POST.get('form6_f3_1')
        form6_f3_2 = request.POST.get('form6_f3_2')
        form6_f3_3 = request.POST.get('form6_f3_3')
        form6_f3_4 = request.POST.get('form6_f3_4')

        form6_f4 = request.POST.get('form6_f4')
        form6_f4_1 = request.POST.get('form6_f4_1')
        form6_f4_2 = request.POST.get('form6_f4_2')
        form6_f4_3 = request.POST.get('form6_f4_3')
        form6_f4_4 = request.POST.get('form6_f4_4')

        form6_f5 = request.POST.get('form6_f5')
        form6_f5_1 = request.POST.get('form6_f5_1')
        form6_f5_2 = request.POST.get('form6_f5_2')
        form6_f5_3 = request.POST.get('form6_f5_3')
        form6_f5_4 = request.POST.get('form6_f5_4')

        request.session['m_approved_by'] = m_approved_by
        request.session['r_approved_by'] = r_approved_by
        request.session['dpo_advice'] = dpo_advice
        request.session['summary_dpo_advice'] = summary_dpo_advice
        request.session['dpo_advice_or_overruled'] = dpo_advice_or_overruled
        request.session['response_reviewed_by'] = response_reviewed_by
        request.session['under_review_by'] = under_review_by
        request.session['comments1'] = comments1
        request.session['comments2'] = comments2


        # request.session['form5_f1_1'] = form5_f1_1
        # request.session['form5_f1_2'] = form5_f1_2
        # request.session['form5_f1_3'] = form5_f1_3

        request.session['form5_f2'] = form5_f2
        request.session['form5_f2_1'] = form5_f2_1
        request.session['form5_f2_2'] = form5_f2_2
        request.session['form5_f2_3'] = form5_f2_3

        request.session['form5_f3'] = form5_f3
        request.session['form5_f3_1'] = form5_f3_1
        request.session['form5_f3_2'] = form5_f3_2
        request.session['form5_f3_3'] = form5_f3_3

        request.session['form5_f4'] = form5_f4
        request.session['form5_f4_1'] = form5_f4_1
        request.session['form5_f4_2'] = form5_f4_2
        request.session['form5_f4_3'] = form5_f4_3

        request.session['form5_f5'] = form5_f5
        request.session['form5_f5_1'] = form5_f5_1
        request.session['form5_f5_2'] = form5_f5_2
        request.session['form5_f5_3'] = form5_f5_3

        request.session['form6_f1'] = form6_f1
        request.session['form6_f1_1'] = form6_f1_1
        request.session['form6_f1_2'] = form6_f1_2
        request.session['form6_f1_3'] = form6_f1_3
        request.session['form6_f1_4'] = form6_f1_4

        request.session['form6_f2'] = form6_f2
        request.session['form6_f2_1'] = form6_f2_1
        request.session['form6_f2_2'] = form6_f2_2
        request.session['form6_f2_3'] = form6_f2_3
        request.session['form6_f2_4'] = form6_f2_4

        request.session['form6_f3'] = form6_f3
        request.session['form6_f3_1'] = form6_f3_1
        request.session['form6_f3_2'] = form6_f3_2
        request.session['form6_f3_3'] = form6_f3_3
        request.session['form6_f3_4'] = form6_f3_4

        request.session['form6_f4'] = form6_f4
        request.session['form6_f4_1'] = form6_f4_1
        request.session['form6_f4_2'] = form6_f4_2
        request.session['form6_f4_3'] = form6_f4_3
        request.session['form6_f4_4'] = form6_f4_4

        request.session['form6_f5'] = form6_f5
        request.session['form6_f5_1'] = form6_f5_1
        request.session['form6_f5_2'] = form6_f5_2
        request.session['form6_f5_3'] = form6_f5_3
        request.session['form6_f5_4'] = form6_f5_4

        pdf_gen = get_pdf(request)
        print('PDF Created: ', pdf_gen)

        context = {
            'f1_final_text': text_box_1,
            'f2_final_text': text_box_2,
            'f3_final_text': text_box_3,
            'f4_final_text': text_box_4,
            'f5_final_text': text_box_5,
            'f6_final_text': text_box_6,
            'f7_final_text': text_box_7,
            'f8_final_text': text_box_8,

            'm_approved_by': m_approved_by,
            'r_approved_by': r_approved_by,
            'dpo_advice': dpo_advice,
            'dpo_advice_or_overruled': dpo_advice_or_overruled,
            'response_reviewed_by': response_reviewed_by,
            'under_review_by': under_review_by,
            'comments1': comments1,
            'comments2': comments2,

        }
        return render(request, 'gdpr_report.html', context)

    final_pdf_dict = request.session.get('final_excel_data')

    # Get Overall risk logic
    if final_pdf_dict is not None:
        step5_box1_likelihood = final_pdf_dict.get('step5_box1_likelihood')
        step5_box1_severity = final_pdf_dict.get('step5_box1_severity')
        step5_box1_overall_risk = overall_risk_logic(step5_box1_likelihood, step5_box1_severity)
        request.session['step5_box1_overall_risk'] = step5_box1_overall_risk
        final_pdf_dict['step5_box1_overall_risk'] = request.session.get('step5_box1_overall_risk')

        final_pdf_dict['s5_b1_option_reduce_risk'] = request.session.get('s5_b1_option_reduce_risk')
        final_pdf_dict['s5_b1_residual_risk'] = request.session.get('s5_b1_residual_risk')
        final_pdf_dict['s5_b1_measure_approved'] = request.session.get('s5_b1_measure_approved')
        final_pdf_dict['s5_b1_effect_on_risk'] = request.session.get('s5_b1_effect_on_risk')



        step5_box2_likelihood = final_pdf_dict.get('step5_box2_likelihood')
        step5_box2_severity = final_pdf_dict.get('step5_box2_severity')
        step5_box2_overall_risk = overall_risk_logic(step5_box2_likelihood, step5_box2_severity)
        request.session['step5_box2_overall_risk'] = step5_box2_overall_risk
        final_pdf_dict['step5_box2_overall_risk'] = request.session.get('step5_box2_overall_risk')

        final_pdf_dict['s5_b2_option_reduce_risk'] = request.session.get('s5_b2_option_reduce_risk')
        final_pdf_dict['s5_b2_residual_risk'] = request.session.get('s5_b2_residual_risk')
        final_pdf_dict['s5_b2_measure_approved'] = request.session.get('s5_b2_measure_approved')
        final_pdf_dict['s5_b2_effect_on_risk'] = request.session.get('s5_b2_effect_on_risk')



        step5_box3_likelihood = final_pdf_dict.get('step5_box3_likelihood')
        step5_box3_severity = final_pdf_dict.get('step5_box3_severity')
        step5_box3_overall_risk = overall_risk_logic(step5_box3_likelihood, step5_box3_severity)
        request.session['step5_box3_overall_risk'] = step5_box3_overall_risk
        final_pdf_dict['step5_box3_overall_risk'] = request.session.get('step5_box3_overall_risk')

        final_pdf_dict['s5_b3_option_reduce_risk'] = request.session.get('s5_b3_option_reduce_risk')
        final_pdf_dict['s5_b3_residual_risk'] = request.session.get('s5_b3_residual_risk')
        final_pdf_dict['s5_b3_measure_approved'] = request.session.get('s5_b3_measure_approved')
        final_pdf_dict['s5_b3_effect_on_risk'] = request.session.get('s5_b3_effect_on_risk')



        step5_box4_likelihood = final_pdf_dict.get('step5_box4_likelihood')
        step5_box4_severity = final_pdf_dict.get('step5_box4_severity')
        step5_box4_overall_risk = overall_risk_logic(step5_box4_likelihood, step5_box4_severity)
        request.session['step5_box4_overall_risk'] = step5_box4_overall_risk
        final_pdf_dict['step5_box4_overall_risk'] = request.session.get('step5_box4_overall_risk')

        final_pdf_dict['s5_b4_option_reduce_risk'] = request.session.get('s5_b4_option_reduce_risk')
        final_pdf_dict['s5_b4_residual_risk'] = request.session.get('s5_b4_residual_risk')
        final_pdf_dict['s5_b4_measure_approved'] = request.session.get('s5_b4_measure_approved')
        final_pdf_dict['s5_b4_effect_on_risk'] = request.session.get('s5_b4_effect_on_risk')

    final_dict = request.session.get('value_dict_final')
    f1_1 = request.session.get('f1_1')
    if f1_1 == '1':
        f1_1_text = 'Our organization does not define the purpose of processing data.'
    elif f1_1 == '2':
        f1_1_text = 'For this scope of service , the project lead defines the purpose of processing the data.'
    elif f1_1 == '3':
        f1_1_text = 'For this scope of service , the project lead collects information from the stakeholders before defining the purpose of processing the data.'
    else:
        f1_1_text = 'For this scope of service , the project lead collects information from the stakeholders and probes them on possible requirements before defining the purpose of processing the data.'

    f1_2 = request.session.get('f1_2')
    if f1_2 == '1':
        f1_2_text = ' We do not define the aim of the project.'
    elif f1_2 == '2':
        f1_2_text = ' The project lead defines the aim of the project by using purpose of processing as a reference.'
    elif f1_2 == '3':
        f1_2_text = ' The project team analyzes the purpose of processing and then collectively decide the aim of the project.'
    else:
        f1_2_text = ' The project team analyzes the purpose of processing and then collectively decide the aim of the project which is further approved by the project lead and other stakeholders.'

    f1_3 = request.session.get('f1_3')
    if f1_3 == '1':
        f1_3_text = ' At present , method of processing the data is not defined for this scope of service.'
    elif f1_3 == '2':
        f1_3_text = ' At present , basic research is done before finalizing the method of processing data.'
    elif f1_3 == '3':
        f1_3_text = ' At present , the project team researches for possible methods and then the final method is decided after discussion with the project lead.'
    else:
        f1_3_text = ' At present ,an extensive research is done by the project team  for possible methods and then the final method is decided after discussion with the project lead and related stakeholders.'


    f1_4 = request.session.get('f1_4')
    if f1_4 == '1':
        f1_4_text = ' Expected benefits from processing data is also not defined for this project.'
    elif f1_4 == '2':
        f1_4_text = ' Expected benefits from processing data are assumed by the project team on the basis of the aim of the project and not documented in the report.'
    elif f1_4 == '3':
        f1_4_text = ' Expected benefits from processing data is well defined by the project lead and also documented in the report.'
    else:
        f1_4_text = ' Expected benefits from processing data is well defined by the project lead and is reviewed by the senior management.'


    f1_5 = request.session.get('f1_5')
    if f1_5 == '1':
        f1_5_text = 'We do not prepare a list of departments who will be affected by this data processing.'
    elif f1_5 == '2':
        f1_5_text = 'The project lead prepares the list of  benefits for each department and documents it.'
    elif f1_5 == '3':
        f1_5_text = 'The project lead sends a message to all the concerned departments about the benefits and documents the list of departments in the report.'
    else:
        f1_5_text = 'The project lead reaches out to all the concerned departments and present them the benefits they will receive from this data processing and documents both the benefits and recommendations/suggestions from each department.'

    f1_6 = request.session.get('f1_6')
    if f1_6 == '1':
        f1_6_text = ' There is no such process to capture the list of team members involved in this data processing.'
    elif f1_6 == '2':
        f1_6_text = ' The list of team members who will be involved in the data processing are decided by the project lead and not documented at any stage.'
    elif f1_6 == '3':
        f1_6_text = ' The list of team members who will be involved in this data processing is documented during the commencement of the project'
    else:
        f1_6_text = ' The list of team members who will be involved in this data processing is well prepared before the start of the project and are given a unique project code to monitor any changes in the team.'

    f4_1 = request.session.get('f4_1')
    if f4_1 == '1':
        f4_1_final_text = 'Consent'
    elif f4_1 == '2':
        f4_1_final_text = 'Performance of contract'
    elif f4_1 == '3':
        f4_1_final_text = 'Legitimate interest'
    elif f4_1 == '4':
        f4_1_final_text = 'Vital interest'
    elif f4_1 == '5':
        f4_1_final_text = 'Legal requirement'
    else:
        f4_1_final_text = 'Public interest'

    f4_1_text = 'The legal basis for processing are - {} Select all from the list; separate by comma if more than 1'.format(f4_1_final_text) + '\n'

    f4_2 = request.session.get('f4_2')
    if f4_2 == '1':
        f4_2_text = ' For this scope of service , we have not defined the legality of data processing.'
    elif f4_2 == '2':
        f4_2_text = ' For this scope of service , our organization recommends project lead to define the legality of data processing.'
    elif f4_2 == '3':
        f4_2_text = ' For this scope of service , it is important for the project lead to define the legality of data processing.'
    else:
        f4_2_text = ' For this scope of service , our organization recommends project lead to define the legality of data processing.'


    f4_3 = request.session.get('f4_3')
    if f4_3 == '1':
        f4_3_text = ' Currently we do not have a process to evaluate the outcomes and check if the assessment reached the desired goals.'
    elif f4_3 == '2':
        f4_3_text = ' To evaluate the outcomes fo the data processing , the project lead checks if they have reached the desired aim which was defined before the start of the assessment.'
    elif f4_3 == '3':
        f4_3_text = ' To evaluate the outcomes fo the data processing , the project lead checks if all the requirements are met and the goals are met within the defined timelines.'
    else:
        f4_3_text = ' To evaluate the outcomes fo the data processing , the project lead checks if all the requirements are met in each phase and the goals are met within the defined timelines.'

    f4_4 = request.session.get('f4_4')
    if f4_4 == '0':
        f4_4_text = ' Yes, data processing will achieve the desired goals.'
    else:
        f4_4_text = ' No, The data processing will not achieve the desired goals.'

    f4_5 = request.session.get('f4_5')
    if f4_5 == '0':
        f4_5_text = ' Yes, there are alternate approaches to achieve the same outcome.'
    else:
        f4_5_text = ' No, we do not have any alternate approach to achieve the same outcome.'


    f4_6 = request.session.get('f4_6')
    if f4_6 == '1':
        f4_6_text = ' We have not have any process to explore the  alternate approaches which can be used to achieve the same results.'
    elif f4_6 == '2':
        f4_6_text = ' We explore a couple of more approaches to achieve the same results. All these approaches are documented in the final report.'
    elif f4_6 == '3':
        f4_6_text = ' We explore all the possible approaches which can achieve the same results and measure their intrusive, post which project team shares it with lead to finalize the final approach to process data.'
    else:
        f4_6_text = ' We explore all the possible approaches which can achieve the same results and measure their intrusive, post which project team check the feasibility before finalizing the final approach to process data.'


    f4_7 = request.session.get('f4_7')
    if f4_7 == '1':
        f4_7_text = r" Our organization does not currently have any KPI's to monitor data quality and integrity for this specific scope of service."
    elif f4_7 == '2':
        f4_7_text = r" We have a list of KPI's in our organization which is used to monitor data quality and integrity for all data related services."
    elif f4_7 == '3':
        f4_7_text = r" Project lead decided the list of  KPI's which will be used to monitor data quality and integrity for this specific scope of service."
    else:
        f4_7_text = r" Our organization has tailored a list of KPI's which will be used to monitor data quality and integrity for this specific scope of service. This list is prepared by the team and reviewed by the lead."

    f4_8 = request.session.get('f4_8')
    if f4_8 == '1':
        f4_8_text = r" Data Subjects are not provided any information after collection of the data."
    elif f4_8 == '2':
        f4_8_text = r" Data Subjects are informed about the intent of processing data. "
    elif f4_8 == '3':
        f4_8_text = ' Data Subjects are informed about the intent of processing data and also notified if there is any change in the scope of service. '
    else:
        f4_8_text = ' Data Subjects are informed about the intent of processing data and also notified if there is any change in the scope of service. Also , they will be notified once the processing is done and data can be archived.'


    f4_9 = request.session.get('f4_9')
    if f4_9 == '1':
        f4_9_text = ' Currently , we do not have process to uphold data subject rights while processing data.'
    elif f4_9 == '2':
        f4_9_text = ' Our processes related  to data subject rights recommends the data processing team  to uphold these rights and aim for complete compliance.'
    elif f4_9 == '3':
        f4_9_text = ' We make it essential for the team to upload all the Data subjects rights and make sure that the team comply to it throughout the processing.'
    else:
        f4_9_text = ' Data subjects rights are given utmost importance while designing data processing plan. We make sure that Data Subject rights are upheld throughout the processing.'

    f4_10 = request.session.get('f4_10')
    if f4_10 == '1':
        f4_10_text = ' At present , we do not have a process to monitor the compliance of the designated entities while data processing.'
    elif f4_10 == '2':
        f4_10_text = " We have a generic process to monitor the compliance of the designated processing entities. This monitoring process considers a few of the KPI's to measure the compliance."
    elif f4_10 == '3':
        f4_10_text = ' We have a robust process to monitor the compliance of the designated processing entities which aims to achieve full compliance.'
    else:
        f4_10_text = ' Compliance of the designated processing entities is a must and aim is always to achieve 100% compliance. We have a robust process to monitor the compliance of the designated processing entities.'

    f4_11 = request.session.get('f4_11')
    if f4_11 == '1':
        f4_11_text = ' We do not have a process to ensure any out of scope data processing activities.'
    elif f4_11 == '2':
        f4_11_text = ' We have a process which ensures that we do not deviate from the decided scope of service while processing data.'
    elif f4_11 == '3':
        f4_11_text = ' We make sure that we stick to the decided scope of service and have no deviations from it while processing. We do not accept any out of scope requirements and prevent any function creep.'
    else:
        f4_11_text = ' We make sure that we stick to the decided scope of service and have no deviations from it while processing. We document all the ad-hoc requirements and plan it out for second phase rather than adding it to on-going processing with aim to prevent function creep.'

    f4_12 = request.session.get('f4_12')
    if f4_12 == '1':
        f4_12_text = ' For this scope of service, we do not have any measures to ensure that processors comply with the scope of project.'
    elif f4_12 == '2':
        f4_12_text = ' Data processors are informed about the boundaries of the scope and asked to esnure compliance while processing data.'
    elif f4_12 == '3':
        f4_12_text = ' We have well defined controls in place to ensure that data processors are aware about the scope of service and they comply by them'
    else:
        f4_12_text = ' We have well defined controls in place to ensure that data processors are aware about the scope of service and they comply by them. Processors are also asked to provide full disclosure about their processing activities and ensure that they document all the data transfers.'

    f4_13 = request.session.get('f4_13')
    if f4_13 == '1':
        f4_13_text = 'There is no measures defined to safeguard any international transfer of the data.'
    elif f4_13 == '2':
        f4_13_text = 'We have a few controls to keep the data safe in the designated geography and prevent them from any international transfer.'
    elif f4_13 == '3':
        f4_13_text = 'We keep a strict check on any international transfer of the data and have several safeguard aiming to do same. We have efficient controls to keep the data safe in the designated geography.'
    else:
        f4_13_text = 'We keep a strict check on any international transfer of the data and have several safeguard aiming to do same. We also make sure that all the data controllers and processors who have access to our data also have efficient controls to keep the data safe in the designated geography.'

    f8_1 = request.session.get('f8_1')
    if f8_1 == '0':
        f8_1_text = 'After completion of DPIA , We assign an associate who will be responsible for collecting and analysing the DPIA outcomes.'
    else:
        f8_1_text = 'Post DPIA process , we do not assign any associate who will be responsible for post DPIA implementation activites'

    f8_2 = request.session.get('f8_2')
    if f8_2 == '0':
        f8_2_text = 'An Associate will be also be responsible for integrates DPIA outcomes in a project plan.'
    else:
        f8_2_text = 'Currently we do not have any associate who will  integrate DPIA outcomes in a project plan.'


    f8_3 = request.session.get('f8_3')
    if f8_3 == '1':
        f8_3_text = 'For this specific scope of service, our organization do not have a process to integrate DPIA Outocomes in our current Processing activities.'
    elif f8_3 == '2':
        f8_3_text = 'For this specific scope of service , Our organization recommends the project lead to integrate DPIA outcomes to prevent any data leak or loss.'
    elif f8_3 == '3':
        f8_3_text = ' In our organization, It is important for the project lead to integrate DPIA outcomes to prevent any data leak or loss.'
    else:
        f8_3_text = ' In our organization, It is mandatory for every project lead to integrate DPIA outcomes before going live with the services.'

    f8_4 = request.session.get('f8_4')
    if f8_4 == '1':
        f8_4_text = 'We do not define any timelines to implement the DPIA outcomes.'
    elif f8_4 == '2':
        f8_4_text = 'Our Post DPIA policy has a basic structure to estimate the timeline for  implementing DPIA outcomes.'
    elif f8_4 == '3':
        f8_4_text = 'Our Post DPIA policy takes good number of factors in consideration which will be used to estimate the timeline for  implementing DPIA outcomes.'
    else:
        f8_4_text = 'Our Post DPIA policy has a well defined process with multiple steps of approval before finalization of the timeline for  implementing DPIA outcomes.'



    f1_final_text = 'Scope of Project:{}'.format(request.session.get('scope_of_service_project')) + '\n' + f1_1_text + f1_2_text + f1_3_text + f1_4_text + f1_5_text + f1_6_text
    f4_final_text =  f4_1_text + f4_2_text + f4_3_text + f4_4_text + f4_5_text + f4_6_text + f4_7_text + f4_8_text + f4_9_text + f4_10_text + f4_11_text + f4_12_text + f4_13_text
    f8_final_text = f8_1_text + f8_2_text + f8_3_text + f8_4_text


    f3_1 = request.session.get('f3_1')
    f3_1_text = f3_1 + '\n'
    f3_2 = request.session.get('f3_2')
    if f3_2 == '0':
        f3_2_text = ' We have designed a consultation process which will seek views of Data subjects on the intended processing.'
    else:
        f3_2_text = ' We do not have any specific consultation process to seeks views of data subjects or their representatives.'
    f3_3 = request.session.get('f3_3')
    if f3_3 == '1':
        f3_3_text = ' At present , we do not follow any process to choose the data subjects to be consulted.'
    elif f3_3 == '2':
        f3_3_text = ' We recommend that the project lead should design a process which run consultation process efficiently.'
    elif f3_3 == '3':
        f3_3_text = ' It is important  for the project team to design a process which run consultation process efficiently.'
    else:
        f3_3_text = ' It is necessary for the project team to design a process which run consultation process efficiently.'

    f3_4 = request.session.get('f3_4')
    if f3_4 == '1':
        f3_4_text = ' Currently , we do not have any process to evaluate consultants.'
    elif f3_4 == '2':
        f3_4_text = ' It is recommended that we should have a process to evaluate the possible options before finalizing the consultants.'
    elif f3_4 == '3':
        f3_4_text = ' It is always important to follow the process in order to find the most suitable consultants.'
    else:
        f3_4_text = ' It is essential part  of our process to evaluate the options based on various factors and then decide who will be designated as the consultants.'

    f3_5 = request.session.get('f3_5')
    if f3_5 == '1':
        f3_5_text = ' There is no process which defines how data subjects can reach out to any consultant.'
    elif f3_5 == '2':
        f3_5_text = ' It is recommended to define a process  which will have all the steps on how data subjects , project teams and others can reach out to the consultants.'
    elif f3_5 == '3':
        f3_5_text = ' Importance is given to the process which has all the steps on how data subjects , project teams and others can reach out to the consultants. A timestamp is maintained for all the meetings.'
    else:
        f3_5_text = ' There is very well defined process which has all the steps on how data subjects , project teams and others can reach out to the consultants. Also , a time is set in which the consultants will have to revert back and also a timestamp is maintained for all the meetings.'

    f3_6 = request.session.get('f3_6')
    if f3_6 == '1':
        f3_6_text = ' We do not have any process to mapping the roles of the consultants.'
    elif f3_6 == '2':
        f3_6_text = ' We map the roles to the consultants on the basis of their most recent work experience.'
    elif f3_6 == '3':
        f3_6_text = ' We ensure that It is ensured that experience, expertise and knowledge is considered before we map roles to the consultants.'
    else:
        f3_6_text = ' It is ensured that experience, expertise, knowledge and choice of the consultants are considered before we map roles to the consultants.'

    f3_7 = request.session.get('f3_7')
    if f3_7 == '0':
        f3_7_text = ' A list is maintained which contains the names of all the involved consultants.'
    else:
        f3_7_text = ' No list is maintained which contains the names of the involved consultants.'

    f3_8 = request.session.get('f3_8')
    if f3_8 == '0':
        f3_8_text = ' Yes , we have a process to monitor the consultation process.'
    else:
        f3_8_text = ' No, we do not have any process to monitor the consultation process.'

    f3_9 = request.session.get('f3_9')
    if f3_9 == '1':
        f3_9_text = ' At present , we do not have any process to deliberate the issues rasied by the consultants.'
    elif f3_9 == '2':
        f3_9_text = ' We have a basic plan which deliberate the issues raised by the consultants.'
    elif f3_9 == '3':
        f3_9_text = ' Our process to deliberate the issues raised by the consultants is well defined and ensures all compliances.'
    else:
        f3_9_text = ' It is our utmost priority to ensure that we follow all the steps of the designed process which enseure that proper deliberation is done on the issues raised by the consultants.'

    f3_10 = request.session.get('f3_10')
    if f3_10 == '0':
        f3_10_text = ' It is our utmost priority to ensure that we follow all the steps of the designed process which ensure that proper deliberation is done on the issues raised by the consultants.'
    else:
        f3_10_text = ' Currently , we do not have any forum where data subjects can rasie their concerns.'

    f3_11 = request.session.get('f3_11')
    if f3_11 == '0':
        f3_11_text = ' We have assigned an associate who will be responsible for keeping a check that all the issues raised by the consultants have been resolved. '
    else:
        f3_11_text = ' We do not have assigned an associate who will be responsible for keeping a check that all the issues raised by the consultants have been resolved. '
    f3_12 = request.session.get('f3_12')
    f3_12_text = f3_12

    f3_final_text = f3_1_text + f3_2_text + f3_3_text + f3_4_text + f3_5_text + f3_6_text + f3_7_text + f3_8_text + f3_9_text + f3_10_text + f3_11_text + f3_12_text

    request.session['f1_final_text'] = f1_final_text
    request.session['f4_final_text'] = f4_final_text
    request.session['f3_final_text'] = f3_final_text
    request.session['f8_final_text'] = f8_final_text
    scope = request.session.get('scope_of_service_project')

    excel_question_dict = get_excel_risk_questions(request)
    liklihood_dict_value = request.session.get('liklihood_dict')

    form5_dict = request.session.get('pdf_print_dict_section5')
    final_excel_data_value = request.session.get('final_excel_data')

    context = {
        'f1_final_text': f1_final_text,
        'f3_final_text': f3_final_text,
        'excel_question_dict': excel_question_dict,
        'liklihood_dict_value': liklihood_dict_value,
        'form5_tot_risk': request.session.get('risk_score5')[1],
        'data_owner': request.session.get('manager'),
        'name_of_dpo': request.session.get('name_of_DPO'),
        'dpo_comments': request.session.get('dpo_comments'),
        'do_comments': request.session.get('do_comments'),
        'consultant_name': request.session.get('f3_1'),
        'f4_final_text': f4_final_text,
        'form5_dict': form5_dict,
        'f8_final_text': f8_final_text,
        'final_dict': final_dict,
        'final_pdf_dict': final_excel_data_value,
        'final_5_dict': request.session.get('pdf_print_dict_section5'),
        'scope': scope,
        'total_no_of_risk': request.session.get('total_no_of_risk')
    }
    return render(request, 'gdpr_report.html', context)


def overall_risk_logic(likelihood, severity):
    logic = ''
    if likelihood == 'Unlikely' and severity == 'Minor':
        logic = 'Low'
    elif likelihood == 'Unlikely' and severity == 'Moderate':
        logic = 'Low'
    elif likelihood == 'Unlikely' and severity == 'Significant':
        logic = 'Medium'
    elif likelihood == 'Possibly' and severity == 'Minor':
        logic = 'Low'
    elif likelihood == 'Possibly' and severity == 'Moderate':
        logic = 'Medium'
    elif likelihood == 'Possibly' and severity == 'Significant':
        logic = 'High'
    elif likelihood == 'Highly Likely' and severity == 'Minor':
        logic = 'Medium'
    elif likelihood == 'Highly Likely' and severity == 'Moderate':
        logic = 'High'
    else:
        logic = 'High'

    return logic

@login_required
def status(request):
    value = request.GET.get('search')
    request.session['approve_or_rejected'] = value
    # generate_pdf = get_pdf(request)
    # print('PDF Status: ', generate_pdf)
    if value == '1':
        app_status = 'Approved'
    else:
        app_status = 'Rejected'
    db_objects = Master.objects.all()
    db_dict_list = []
    for instance in db_objects:
        db_dict_list.append(instance.__dict__)

    session_title = request.session.get('title')
    session_status = request.session.get('status')
    session_date = request.session.get('date')
    for item in db_dict_list:
        if item['title'] == session_title and item['status'] == session_status \
                and item['date'] == session_date or item['title'] == session_title \
                and item['status'] == session_status or item['title'] == session_title \
                and item['date'] == session_date:
            Master.objects.filter(title=session_title).update(status=app_status)

    context = {
        'status': value,
    }
    return render(request, 'status.html', context)


def forget_password(request):
    if request.method == 'POST':
        try:
            f_username = request.POST.get('f_username')
            if not User.objects.filter(username=f_username).first():
                messages.success(request, 'Not {} user found with this username.'.format(f_username))
                text = 'Username not found. Please check the username and try again.'
                context = {
                    'Text': text
                }
                return render(request, 'forget_password.html', context)

            user_obj = User.objects.filter(username=f_username).first()
            profile_obj_check = Profile.objects.filter(user=user_obj)
            if not profile_obj_check:
                profile_obj = Profile.objects.create(user=user_obj)
                profile_obj.save()

            user_obj = User.objects.get(username=f_username)
            user_email = user_obj.email

            token = str(uuid.uuid4())
            print('Token: ', token)
            profile_obj = Profile.objects.get(user=user_obj)
            profile_obj.forget_password_token = token
            profile_obj.save()

            user_obj = User.objects.get(username=f_username)
            send_forget_password_mail(user_obj.email, token)
            messages.success(request, 'An email has been sent to your registered email address.')
            text = 'Password reset mail has been sent to the provided mail address.'
            context = {
                'Text': text
            }
            return render(request, 'forget_password.html', context)
        except Exception as e:
            print(e)
    else:
        return render(request, 'forget_password.html')
    context = {
    }
    return render(request, 'forget_password.html', context)


def change_password(request, token):
    context = {}
    try:
        profile_obj = Profile.objects.filter(forget_password_token=token).first()
        context = {'user_id': profile_obj.user.id}

        if request.method == 'POST':
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('reconfirm_password')
            user_id = request.POST.get('user_id')

            if user_id is None:
                text = 'No {} user id found.'.format(user_id)
                messages.success(request, 'No {} user id found.'.format(user_id))
                context['Text'] = text
                return render(request, f'/change_password/{token}/', context)

            if new_password != confirm_password:
                text2 = 'Passwords do not match. Please check and try again.'
                context['Text'] = text2
                return render(request, f'/change_password/{token}/', context)

            user_obj = User.objects.get(id=user_id)
            user_obj.set_password(new_password)
            user_obj.save()
            return redirect('login')

    except Exception as e:
        print(e)
    return render(request, 'change_password.html', context)


def send_forget_password_mail(email, token):
    subject = ' DPIA Application forget password link'
    # message = f'Hi , click on the link to reset your password https://privacyone.azurewebsites.net/change_password/{token}/'

    # for local env test
    message = f'Hi , click on the link to reset your password http://127.0.0.1:8900/change_password/{token}/'
    email_from = settings.EMAIL_HOST_USER
    recipient_list = [email]
    send_mail(subject, message, email_from, recipient_list)
    return True


@login_required
def pdf_button(request):

    value = request.GET.get('search')
    risk_score1 = request.session.get('risk_score1')
    risk_score2 = request.session.get('risk_score2')
    risk_score3 = request.session.get('risk_score3')
    risk_score4 = request.session.get('risk_score4')
    risk_score5 = request.session.get('risk_score5')
    risk_score6 = request.session.get('risk_score6')
    risk_score7 = request.session.get('risk_score7')
    risk_score8 = request.session.get('risk_score8')
    form1_percentage = request.session.get('form1_percentage')
    form2_percentage = request.session.get('form2_percentage')
    form3_percentage = request.session.get('form3_percentage')
    form4_percentage = request.session.get('form4_percentage')
    form5_percentage = request.session.get('form5_percentage')
    form6_percentage = request.session.get('form6_percentage')
    form7_percentage = request.session.get('form7_percentage')
    form8_percentage = request.session.get('form8_percentage')
    context = {
        'value': value,
        'risk_score1': risk_score1,
        'risk_score2': risk_score2,
        'risk_score3': risk_score3,
        'risk_score4': risk_score4,
        'risk_score5': risk_score5,
        'risk_score6': risk_score6,
        'risk_score7': risk_score7,
        'risk_score8': risk_score8,
        'form1_percentage': form1_percentage,
        'form2_percentage': form2_percentage,
        'form3_percentage': form3_percentage,
        'form4_percentage': form4_percentage,
        'form5_percentage': form5_percentage,
        'form6_percentage': form6_percentage,
        'form7_percentage': form7_percentage,
        'form8_percentage': form8_percentage,
        'name_of_controller': request.session.get('manager'),
        'name_of_dpo': request.session.get('name_of_DPO'),
        'title_of_dpo': request.session.get('title_of_DPO')
    }

    return render(request, 'risk_summary.html', context)


def get_excel_risk_questions(request):
    value_excel_dict = {}
    response_list = []
    f1_1 = request.session.get('f1_1')
    if f1_1 == '1' or f1_1 == '2':
        f1_1_question = 'To what extent is the purpose of the processing defined in your organization?'
        response_list.append(replace_value_1_2(
            f1_1))
        request.session['f1_1_question'] = f1_1_question
        value_excel_dict['f1_1'] = f1_1_question

    f1_2 = request.session.get('f1_2')
    if f1_2 == '1' or f1_2 == '2':
        f1_2_question = 'To what extent is the aim of the project defined in your organization?'
        response_list.append(replace_value_1_2(f1_2))
        request.session['f1_2_question'] = f1_2_question
        value_excel_dict['f1_2'] = f1_2_question

    f1_3 = request.session.get('f1_3')
    if f1_3 == '1' or f1_3 == '2':
        f1_3_question = 'To what extent is method of processing defined in your organization?'
        response_list.append(replace_value_1_2(f1_3))
        request.session['f1_3_question'] = f1_3_question
        value_excel_dict['f1_3'] = f1_3_question

    f1_4 = request.session.get('f1_4')
    if f1_4 == '1' or f1_4 == '2':
        f1_4_question = 'To what extent are the expected benefit(s) from the processing defined?'
        response_list.append(replace_value_1_2(f1_4))
        request.session['f1_4_question'] = f1_4_question
        value_excel_dict['f1_4'] = f1_4_question

    f1_5 = request.session.get('f1_5')
    if f1_5 == '1' or f1_5 == '2':
        f1_5_question = 'To what extent is the list of departments who will benefit from this processing defined in your organization?'
        response_list.append(replace_value_1_2(f1_5))
        request.session['f1_5_question'] = f1_5_question
        value_excel_dict['f1_5'] = f1_5_question

    f1_6 = request.session.get('f1_6')
    if f1_6 == '1' or f1_6 == '2':
        f1_6_question = 'To what extent is the list of members / teams who will be involved in the processing defined in your organization?'
        response_list.append(replace_value_1_2(f1_6))
        request.session['f1_6_question'] = f1_6_question
        value_excel_dict['f1_6'] = f1_6_question

    number_of_risk = request.session.get('risk_score1')[1]

    request.session[
        'f5_1_question'] = 'Have you assigned a person/team who will be responsible for Identifying all the risks?'

    # Section 2
    f2_1 = request.session.get('f2_1')
    if f2_1 == '1' or f2_1 == '2':
        f2_1_question = 'How matured is the process of collecting, using, storing and deleting data in your organization?'
        response_list.append(replace_value_init_define(f2_1))
        value_excel_dict['f2_1'] = f2_1_question

    f2_2 = request.session.get('f2_2')
    if f2_2 == '1' or f2_2 == '2':
        f2_2_question = 'To what extent are the data source(s) used to initiate processing defined in your organization?'
        response_list.append(replace_value_init_define(f2_2))
        value_excel_dict['f2_2'] = f2_2_question

    f2_3 = request.session.get('f2_3')
    if f2_3 == '1' or f2_3 == '2':
        f2_3_question = 'Do you have a list of people who will have access to this data?'
        value_excel_dict['f2_3'] = f2_3_question
        response_list.append(replace_value_yes_no(f2_3))

    f2_4 = request.session.get('f2_4')
    if f2_4 == '1' or f2_4 == '2':
        f2_4_question = 'To what extent is the types of High-risk processing defined for this data processing?'
        response_list.append(replace_value_init_define(f2_4))
        value_excel_dict['f2_4'] = f2_4_question

    f2_5 = request.session.get('f2_5')
    if f2_5 == '1' or f2_5 == '2':
        f2_5_question = 'Do you comply by legal requirements to collect the data (Do you have consent of the data subjects)?'
        response_list.append(replace_value_yes_no(f2_5))
        value_excel_dict['f2_5'] = f2_5_question

    f2_6 = request.session.get('f2_6')
    if f2_6 == '1' or f2_6 == '2':
        f2_6_question = 'How many data subjects are likely to be affected by the project?'
        value_excel_dict['f2_6'] = f2_6_question
        response_list.append(replace_value_1_2(f2_6))

    f2_7 = request.session.get('f2_7')
    if f2_7 == '1' or f2_7 == '2':
        f2_7_question = 'Where is data stored ?'
        value_excel_dict['f2_7'] = f2_7_question
        response_list.append(replace_value_1_2(f2_7))

    f2_8 = request.session.get('f2_8')
    if f2_8 == '1' or f2_8 == '2':
        f2_8_question = 'Do you have appropriate measures to destroy data after use?'
        value_excel_dict['f2_8'] = f2_8_question
        response_list.append(replace_value_yes_no(f2_8))

    f2_9 = request.session.get('f2_9')
    if f2_9 == '1' or f2_9 == '2':
        f2_9_question = 'To what extent is the Data retention policy defined in your organization?'
        response_list.append(replace_value_init_define(f2_9))
        value_excel_dict['f2_9'] = f2_9_question

    # Section 3
    f3_1 = request.session.get('f3_1')
    if f3_1 == '1' or f3_1 == '2':
        f3_1_question = 'Who all will be involved in the consultation process ?'
        value_excel_dict['f3_1'] = f3_1_question

    f3_2 = request.session.get('f3_2')
    if f3_2 == '1' or f3_2 == '2':
        f3_2_question = 'Have you designed a Consultation Process . This Consultation process will involve seeking views of Data subjects or their representatives on the intended processing ?'
        value_excel_dict['f3_2'] = f3_2_question
        response_list.append(replace_value_yes_no(f3_2))

    f3_3 = request.session.get('f3_3')
    if f3_3 == '1' or f3_3 == '2':
        f3_3_question = 'Is there a process in your organization to choose the people for consultation on the proposed data processing ?'
        value_excel_dict['f3_3'] = f3_3_question
        response_list.append(replace_value_process_no_process(f3_3))

    f3_4 = request.session.get('f3_4')
    if f3_4 == '1' or f3_4 == '2':
        f3_4_question = 'How matured is the process in your organization of evaluating data processors, information security experts or other staff as consultants ?;'
        value_excel_dict['f3_4'] = f3_4_question
        response_list.append(replace_value_process_no_process(f3_4))

    f3_5 = request.session.get('f3_5')
    if f3_5 == '1' or f3_5 == '2':
        f3_5_question = 'How matured is the process in your organization of reaching out to the consultants for consultation by data subjects, project teams etc ?;'
        value_excel_dict['f3_5'] = f3_5_question
        response_list.append(replace_value_process_no_process(f3_5))

    f3_6 = request.session.get('f3_6')
    if f3_6 == '1' or f3_6 == '2':
        f3_6_question = 'To what extent is your process defined for mapping roles of the selected consultant(s) ?'
        value_excel_dict['f3_6'] = f3_6_question
        response_list.append(replace_value_process_no_process(f3_6))

    f3_7 = request.session.get('f3_7')
    if f3_7 == '1' or f3_7 == '2':
        f3_7_question = 'Do you have a list of members, other than designated consultant, who will also involve from the organization in the consultation process?'
        value_excel_dict['f3_7'] = f3_7_question
        response_list.append(replace_value_yes_no(f3_7))

    f3_8 = request.session.get('f3_8')
    if f3_8 == '1' or f3_8 == '2':
        f3_8_question = 'Do you have a process to monitor the consultation process?'
        value_excel_dict['f3_8'] = f3_8_question
        response_list.append(replace_value_yes_no(f3_8))

    f3_9 = request.session.get('f3_9')
    if f3_9 == '1' or f3_9 == '2':
        f3_9_question = 'How matured is the process in your organization to deliberate on the issues raised by the consultants or data subjects ?'
        value_excel_dict['f3_9'] = f3_9_question
        response_list.append(replace_value_1_2(f3_9))

    f3_10 = request.session.get('f3_10')
    if f3_10 == '1' or f3_10 == '2':
        f3_10_question = 'Do you have a proper forum where these consultants can provide assurances and raise their concerns ?'
        value_excel_dict['f3_10'] = f3_10_question
        response_list.append(replace_value_yes_no(f3_10))

    f3_11 = request.session.get('f3_11')
    if f3_11 == '1' or f3_11 == '2':
        f3_11_question = 'Have you assigned a person who will be responsible for keeping a check that all the issues raised by these consultants have been resolved ?'
        value_excel_dict['f3_11'] = f3_11_question
        response_list.append(replace_value_yes_no(f3_11))

    f3_12 = request.session.get('f3_12')
    if f3_12 == '1' or f3_12 == '2':
        f3_12_question = 'If Response to the above question is Yes, Please provide the name of the assigned personnel.'
        value_excel_dict['f3_12'] = f3_12_question

    # Section 4
    f4_2 = request.session.get('f4_2')
    if f4_2 == '1' or f4_2 == '2':
        f4_2_question = 'How matured is the process of defining the legality of data processing with respect to the data subjects in your organization?'
        value_excel_dict['f4_2'] = f4_2_question
        response_list.append(replace_value_process_no_process(f4_2))

    f4_3 = request.session.get('f4_3')
    if f4_3 == '1' or f4_3 == '2':
        f4_3_question = 'Is there a process to evaluate the outcome of the processing with respect to the desired goals?'
        value_excel_dict['f4_3'] = f4_3_question
        response_list.append(replace_value_1_2(f4_3))

    f4_4 = request.session.get('f4_4')
    if f4_4 == '1' or f4_4 == '2':
        f4_4_question = 'Will the data processing achieve the desired goal ?'
        value_excel_dict['f4_1'] = f4_4_question
        response_list.append(replace_value_1_2(f4_4))

    f4_5 = request.session.get('f4_5')
    if f4_5 == '1' or f4_5 == '2':
        f4_5_question = 'Is there an alternate approach/s to achieve the same outcome ?'
        value_excel_dict['f4_5'] = f4_5_question
        response_list.append(replace_value_1_2(f4_5))

    f4_6 = request.session.get('f4_6')
    if f4_6 == '1' or f4_6 == '2':
        f4_6_question = 'How matured is the process to explore the alternate approaches (less intrusive measures) in order to achieve the same results ?'
        value_excel_dict['f4_6'] = f4_6_question
        response_list.append(replace_value_1_2(f4_6))

    f4_7 = request.session.get('f4_7')
    if f4_7 == '1' or f4_7 == '2':
        f4_7_question = 'What is the maturity of defined data quality KPIs and metrics in your organization to monitor data quality and integrity ?'
        value_excel_dict['f4_7'] = f4_7_question
        response_list.append(replace_value_1_2(f4_7))

    f4_8 = request.session.get('f4_8')
    if f4_8 == '1' or f4_8 == '2':
        f4_8_question = 'To what extent is the information that will be given to the data subjects defined in your organization ?'
        value_excel_dict['f4_8'] = f4_8_question
        response_list.append(replace_value_1_2(f4_8))

    f4_9 = request.session.get('f4_9')
    if f4_9 == '1' or f4_9 == '2':
        f4_9_question = 'How matured is the process in your organization of upholding the Data Subject Rights while processing their data ?'
        value_excel_dict['f4_9'] = f4_9_question
        response_list.append(replace_value_1_2(f4_9))

    f4_10 = request.session.get('f4_10')
    if f4_10 == '1' or f4_10 == '2':
        f4_10_question = 'How matured is the process to monitor how compliant designated processing entities are while preforming the proposed processing ?'
        value_excel_dict['f4_10'] = f4_10_question
        response_list.append(replace_value_1_2(f4_10))

    f4_11 = request.session.get('f4_11')
    if f4_11 == '1' or f4_11 == '2':
        f4_11_question = 'How matured is your process is to ensure that data processing is not used for out of scope requirement in order to prevent function creep ?'
        value_excel_dict['f4_11'] = f4_11_question
        response_list.append(replace_value_1_2(f4_11))

    f4_12 = request.session.get('f4_12')
    if f4_12 == '1' or f4_12 == '2':
        f4_12_question = 'To what extreme the measures defined to ensure processors comply to the scope of the project during data processing ?'
        value_excel_dict['f4_12'] = f4_12_question
        response_list.append(replace_value_1_2(f4_12))

    f4_13 = request.session.get('f4_13')
    if f4_13 == '1' or f4_13 == '2':
        f4_13_question = 'To what extend the measures defined to safeguard any international transfers ?'
        value_excel_dict['f4_13'] = f4_13_question
        response_list.append(replace_value_1_2(f4_13))

    # Section 5
    f5_1 = request.session.get('f5_1')
    if f5_1 == '1' or f5_1 == '2':
        f5_1_question = 'Have you assigned a person/team who will be responsible for Identifying all the risks ?'
        request.session['f5_1_question'] = f5_1_question
        value_excel_dict['f5_1'] = f5_1_question
        response_list.append(replace_value_yes_no(f5_1))

    f5_2 = request.session.get('f5_2')
    if f5_2 == '1' or f5_2 == '2':
        f5_2_question = 'How matured is the process of documenting type and details of the risks from the proposed processing in your organization?'
        request.session['f5_2_question'] = f5_2_question
        value_excel_dict['f5_2'] = f5_2_question
        response_list.append(replace_value_process_no_process(f5_2))

    f5_3 = request.session.get('f5_3')
    if f5_3 == '1' or f5_3 == '2':
        f5_3_question = 'How matured is the process of calculating likehlihood and severity of the risk from the proposed processing in your organization?'
        request.session['f5_3_question'] = f5_3_question
        value_excel_dict['f5_3'] = f5_3_question
        response_list.append(replace_value_process_no_process(f5_3))

    f5_4 = request.session.get('f5_4')
    if f5_4 == '1' or f5_4 == '2':
        f5_4_question = 'How matured is the process of finding residual risks < define what is residual risks > defined in your organization?'
        request.session['f5_4_question'] = f5_4_question
        value_excel_dict['f5_4'] = f5_4_question
        response_list.append(replace_value_process_no_process(f5_4))

    f5_5 = request.session.get('f5_5')
    if f5_5 == '1' or f5_5 == '2':
        f5_5_question = 'To what extent is the plan to quantify the impact of the risks defined in your organization?'
        request.session['f5_5_question'] = f5_5_question
        value_excel_dict['f5_5'] = f5_5_question
        response_list.append(replace_value_process_no_process(f5_5))

    # Section 6
    f6_1 = request.session.get('f6_1')
    if f6_1 == '1' or f6_1 == '2':
        f6_1_question = 'Have you assigned a person/ team who will be responsible for describing the risk mitigation measures ?'
        value_excel_dict['f6_1'] = f6_1_question
        response_list.append(replace_value_yes_no(f6_1))

    f6_2 = request.session.get('f6_2')
    if f6_2 == '1' or f6_2 == '2':
        f6_2_question = 'Have you assigned a person/ team who will be responsible who will be responsible for defining the impact of risk ?'
        value_excel_dict['f6_2'] = f6_2_question
        response_list.append(replace_value_yes_no(f6_2))

    f6_3 = request.session.get('f6_3')
    if f6_3 == '1' or f6_3 == '2':
        f6_3_question = 'To what extent is the plan to reduce or eliminate risk defined in your organization ?'
        value_excel_dict['f6_3'] = f6_3_question
        response_list.append(replace_value_1_2(f6_3))

    f6_4 = request.session.get('f6_4')
    if f6_4 == '1' or f6_4 == '2':
        f6_4_question = 'Have you assigned a person who will be responsible for planning and leading the risk mitigation action plan ?'
        value_excel_dict['f6_4'] = f6_4_question
        response_list.append(replace_value_yes_no(f6_4))

    f6_5 = request.session.get('f6_5')
    if f6_5 == '1' or f6_1 == '2':
        f6_5_question = 'Have you assigned a person who will be responsible for auditing the plan ?'
        value_excel_dict['f6_5'] = f6_5_question
        response_list.append(replace_value_yes_no(f6_5))

    # Section 7
    f7_1 = request.session.get('f7_1')
    if f7_1 == '1' or f7_1 == '2':
        f7_1_question = 'Have you assigned a person who will be responsible for examining the residual risks ?'
        value_excel_dict['f7_1'] = f7_1_question
        response_list.append(replace_value_yes_no(f7_1))

    f7_2 = request.session.get('f7_2')
    if f7_2 == '1' or f7_2 == '2':
        f7_2_question = 'Comments provided by Consultation team'
        value_excel_dict['f7_2'] = f7_2_question
        response_list.append(replace_value_1_2(f7_2))

    # Section 8    import ipdb
    f8_1 = request.session.get('f8_1')
    if f8_1 == '1':
        f8_1_question = 'Have you assigned a person who will be responsible for collecting and analysing the DPIA outcomes ?'
        value_excel_dict['f8_1'] = f8_1_question
        response_list.append(replace_value_yes_no(f8_1))

    f8_2 = request.session.get('f8_2')
    if f8_2 == '0':
        f8_2_question = 'Have you assigned a person who will be responsible for planning intergration of DPIA outcomes into the project plan?'
        value_excel_dict['f8_2'] = f8_2_question
        response_list.append(replace_value_yes_no(f8_2))

    f8_3 = request.session.get('f8_3')
    if f8_3 == '1' or f8_3 == '2':
        f8_3_question = 'To what extent is your policy for assigning time period to Implement the DPIA outcomes defined in your organization ?'
        value_excel_dict['f8_3'] = f8_3_question
        response_list.append(replace_value_1_2(f8_3))

    f8_4 = request.session.get('f8_4')
    if f8_4 == '1' or f8_4 == '2':
        f8_4_question = 'To what extent are the KPIs which will depict the success of Integration defined in your organization ?'
        value_excel_dict['f8_4'] = f8_4_question
        response_list.append(replace_value_1_2(f8_4))
    request.session['response_list'] = response_list

    return value_excel_dict

liklihood_dict = {}

def risk_summary_box_1(request):
    if request.method == 'POST':
        risk_score1 = request.session.get('risk_score1')
        risk_score2 = request.session.get('risk_score2')
        risk_score3 = request.session.get('risk_score3')
        risk_score4 = request.session.get('risk_score4')
        risk_score5 = request.session.get('risk_score5')
        risk_score6 = request.session.get('risk_score6')
        risk_score7 = request.session.get('risk_score7')
        risk_score8 = request.session.get('risk_score8')
        form1_percentage = request.session.get('form1_percentage')
        form2_percentage = request.session.get('form2_percentage')
        form3_percentage = request.session.get('form3_percentage')
        form4_percentage = request.session.get('form4_percentage')
        form5_percentage = request.session.get('form5_percentage')
        form6_percentage = request.session.get('form6_percentage')
        form7_percentage = request.session.get('form7_percentage')
        form8_percentage = request.session.get('form8_percentage')


        s1_b1_likelihood = request.POST.get('1_1_likelihood')  # Step1 box1
        s1_b1_severity = request.POST.get('1_1_severity')  # Step1 box1
        s1_b1_option_reduce_risk = request.POST.get('1_1_option_reduce_risk')
        s1_b1_effect_on_risk = request.POST.get('1_1_effect_on_risk')
        s1_b1_residual_risk = request.POST.get('1_1_residual_risk')
        s1_b1_measure_approved = request.POST.get('1_1_measure_approved')

        if s1_b1_likelihood and s1_b1_severity is not None:
            liklihood_dict['s1_b1'] = {
                's1_b1_likelihood': replace_values_likelihood(s1_b1_likelihood),
                's1_b1_severity': replace_values_severity(s1_b1_severity),
                's1_b1_option_reduce_risk': s1_b1_option_reduce_risk,
                's1_b1_overall_risk': overall_risk_logic(replace_values_likelihood(s1_b1_likelihood), replace_values_severity(s1_b1_severity)),
                's1_b1_effect_on_risk': effect_on_risk_logic(s1_b1_effect_on_risk),
                's1_b1_residual_risk': residual_risk_logic(s1_b1_residual_risk),
                's1_b1_measure_approved': measure_approved_logic(s1_b1_measure_approved)

            }

        s1_b2_likelihood = request.POST.get('1_2_likelihood')  # Step1 box2
        s1_b2_severity = request.POST.get('1_2_severity')  # Step1 box2
        s1_b2_option_reduce_risk = request.POST.get('1_2_option_reduce_risk')
        s1_b2_effect_on_risk = request.POST.get('1_2_effect_on_risk')
        s1_b2_residual_risk = request.POST.get('1_2_residual_risk')
        s1_b2_measure_approved = request.POST.get('1_2_measure_approved')
        if s1_b2_likelihood and s1_b2_severity is not None:
            liklihood_dict['s1_b2'] = {
                's1_b2_likelihood': replace_values_likelihood(s1_b2_likelihood),
                's1_b2_severity': replace_values_severity(s1_b2_severity),
                's1_b2_option_reduce_risk': s1_b2_option_reduce_risk,
                's1_b2_overall_risk': overall_risk_logic(replace_values_likelihood(s1_b2_likelihood),
                                                         replace_values_severity(s1_b2_severity)),
                's1_b2_effect_on_risk': effect_on_risk_logic(s1_b2_effect_on_risk),
                's1_b2_residual_risk': residual_risk_logic(s1_b2_residual_risk),
                's1_b2_measure_approved': measure_approved_logic(s1_b2_measure_approved)

            }

        s1_b3_likelihood = request.POST.get('1_3_likelihood')  # Step1 box3
        s1_b3_severity = request.POST.get('1_3_severity')  # Step1 box3
        s1_b3_option_reduce_risk = request.POST.get('1_3_option_reduce_risk')
        s1_b3_effect_on_risk = request.POST.get('1_3_effect_on_risk')
        s1_b3_residual_risk = request.POST.get('1_3_residual_risk')
        s1_b3_measure_approved = request.POST.get('1_3_measure_approved')
        if s1_b3_likelihood and s1_b3_severity is not None:
            liklihood_dict['s1_b3'] = {
                's1_b3_likelihood': replace_values_likelihood(s1_b3_likelihood),
                's1_b3_severity': replace_values_severity(s1_b3_severity),
                's1_b3_option_reduce_risk': s1_b3_option_reduce_risk,
                's1_b3_overall_risk': overall_risk_logic(replace_values_likelihood(s1_b3_likelihood),
                                                         replace_values_severity(s1_b3_severity)),
                's1_b3_effect_on_risk': effect_on_risk_logic(s1_b3_effect_on_risk),
                's1_b3_residual_risk': residual_risk_logic(s1_b3_residual_risk),
                's1_b3_measure_approved': measure_approved_logic(s1_b3_measure_approved)

            }

        s1_b4_likelihood = request.POST.get('1_4_likelihood')  # Step1 box4
        s1_b4_severity = request.POST.get('1_4_severity')  # Step1 box4
        s1_b4_option_reduce_risk = request.POST.get('1_4_option_reduce_risk')
        s1_b4_effect_on_risk = request.POST.get('1_4_effect_on_risk')
        s1_b4_residual_risk = request.POST.get('1_4_residual_risk')
        s1_b4_measure_approved = request.POST.get('1_4_measure_approved')
        if s1_b4_likelihood and s1_b4_severity is not None:
            liklihood_dict['s1_b4'] = {
                's1_b4_likelihood': replace_values_likelihood(s1_b4_likelihood),
                's1_b4_severity': replace_values_severity(s1_b4_severity),
                's1_b4_option_reduce_risk': s1_b4_option_reduce_risk,
                's1_b4_overall_risk': overall_risk_logic(replace_values_likelihood(s1_b4_likelihood),
                                                         replace_values_severity(s1_b4_severity)),
                's1_b4_effect_on_risk': effect_on_risk_logic(s1_b4_effect_on_risk),
                's1_b4_residual_risk': residual_risk_logic(s1_b4_residual_risk),
                's1_b4_measure_approved': measure_approved_logic(s1_b4_measure_approved)

            }

        s1_b5_likelihood = request.POST.get('1_5_likelihood')  # Step1 box5
        s1_b5_severity = request.POST.get('1_5_severity')  # Step1 box5
        s1_b5_option_reduce_risk = request.POST.get('1_5_option_reduce_risk')
        s1_b5_effect_on_risk = request.POST.get('1_5_effect_on_risk')
        s1_b5_residual_risk = request.POST.get('1_5_residual_risk')
        s1_b5_measure_approved = request.POST.get('1_5_measure_approved')
        if s1_b5_likelihood and s1_b5_severity is not None:
            liklihood_dict['s1_b5'] = {
                's1_b5_likelihood': replace_values_likelihood(s1_b5_likelihood),
                's1_b5_severity': replace_values_severity(s1_b5_severity),
                's1_b5_option_reduce_risk': s1_b5_option_reduce_risk,
                's1_b5_effect_on_risk': effect_on_risk_logic(s1_b5_effect_on_risk),
                's1_b5_overall_risk': overall_risk_logic(replace_values_likelihood(s1_b5_likelihood),
                                                         replace_values_severity(s1_b5_severity)),
                's1_b5_residual_risk': residual_risk_logic(s1_b5_residual_risk),
                's1_b5_measure_approved': measure_approved_logic(s1_b5_measure_approved)

            }

        s1_b6_likelihood = request.POST.get('1_6_likelihood')  # Step1 box6
        s1_b6_severity = request.POST.get('1_6_severity')  # Step1 box6
        s1_b6_option_reduce_risk = request.POST.get('1_6_option_reduce_risk')
        s1_b6_effect_on_risk = request.POST.get('1_6_effect_on_risk')
        s1_b6_residual_risk = request.POST.get('1_6_residual_risk')
        s1_b6_measure_approved = request.POST.get('1_6_measure_approved')
        if s1_b6_likelihood and s1_b6_severity is not None:
            liklihood_dict['s1_b6'] = {
                's1_b6_likelihood': replace_values_likelihood(s1_b6_likelihood),
                's1_b6_severity': replace_values_severity(s1_b6_severity),
                's1_b6_option_reduce_risk': s1_b6_option_reduce_risk,
                's1_b6_overall_risk': overall_risk_logic(replace_values_likelihood(s1_b6_likelihood),
                                                         replace_values_severity(s1_b6_severity)),
                's1_b6_effect_on_risk': effect_on_risk_logic(s1_b6_effect_on_risk),
                's1_b6_residual_risk': residual_risk_logic(s1_b6_residual_risk),
                's1_b6_measure_approved': measure_approved_logic(s1_b6_measure_approved)

            }
        excel_question_dict = get_excel_risk_questions(request)
        text_to_excel(request, excel_question_dict, liklihood_dict)
        request.session['liklihood_dict'] = liklihood_dict
        context = {
            'risk_score1': risk_score1,
            'risk_score2': risk_score2,
            'risk_score3': risk_score3,
            'risk_score4': risk_score4,
            'risk_score5': risk_score5,
            'risk_score6': risk_score6,
            'risk_score7': risk_score7,
            'risk_score8': risk_score8,
            'form1_percentage': form1_percentage,
            'form2_percentage': form2_percentage,
            'form3_percentage': form3_percentage,
            'form4_percentage': form4_percentage,
            'form5_percentage': form5_percentage,
            'form6_percentage': form6_percentage,
            'form7_percentage': form7_percentage,
            'form8_percentage': form8_percentage,
            'total_no_of_risk': request.session.get('total_no_of_risk'),
            'name_of_controller': request.session.get('manager'),
            'name_of_dpo': request.session.get('name_of_DPO'),
            'title_of_dpo': request.session.get('title_of_DPO')
        }
        return render(request, 'risk_summary.html', context)


def risk_summary_box_2(request):
    if request.method == 'POST':
        risk_score1 = request.session.get('risk_score1')
        risk_score2 = request.session.get('risk_score2')
        risk_score3 = request.session.get('risk_score3')
        risk_score4 = request.session.get('risk_score4')
        risk_score5 = request.session.get('risk_score5')
        risk_score6 = request.session.get('risk_score6')
        risk_score7 = request.session.get('risk_score7')
        risk_score8 = request.session.get('risk_score8')
        form1_percentage = request.session.get('form1_percentage')
        form2_percentage = request.session.get('form2_percentage')
        form3_percentage = request.session.get('form3_percentage')
        form4_percentage = request.session.get('form4_percentage')
        form5_percentage = request.session.get('form5_percentage')
        form6_percentage = request.session.get('form6_percentage')
        form7_percentage = request.session.get('form7_percentage')
        form8_percentage = request.session.get('form8_percentage')

        s2_b1_likelihood = request.POST.get('2_1_likelihood')  # Step2 box1
        s2_b1_severity = request.POST.get('2_1_severity')  # Step2 box1
        s2_b1_option_reduce_risk = request.POST.get('2_1_option_reduce_risk')
        s2_b1_effect_on_risk = request.POST.get('2_1_effect_on_risk')
        s2_b1_residual_risk = request.POST.get('2_1_residual_risk')
        s2_b1_measure_approved = request.POST.get('2_1_measure_approved')

        if s2_b1_likelihood and s2_b1_severity is not None:
            liklihood_dict['s2_b1'] = {
                's2_b1_likelihood': replace_values_likelihood(s2_b1_likelihood),
                's2_b1_severity': replace_values_severity(s2_b1_severity),
                's2_b1_option_reduce_risk': s2_b1_option_reduce_risk,
                's2_b1_overall_risk': overall_risk_logic(replace_values_likelihood(s2_b1_likelihood),
                                                         replace_values_severity(s2_b1_severity)),
                's2_b1_effect_on_risk': effect_on_risk_logic(s2_b1_effect_on_risk),
                's2_b1_residual_risk': residual_risk_logic(s2_b1_residual_risk),
                's2_b1_measure_approved': measure_approved_logic(s2_b1_measure_approved)

            }

        s2_b2_likelihood = request.POST.get('2_2_likelihood')  # Step2 box2
        s2_b2_severity = request.POST.get('2_2_severity')  # Step2 box2
        s2_b2_option_reduce_risk = request.POST.get('2_2_option_reduce_risk')
        s2_b2_effect_on_risk = request.POST.get('2_2_effect_on_risk')
        s2_b2_residual_risk = request.POST.get('2_2_residual_risk')
        s2_b2_measure_approved = request.POST.get('2_2_measure_approved')

        if s2_b2_likelihood and s2_b2_severity is not None:
            liklihood_dict['s2_b2'] = {
                's2_b2_likelihood': replace_values_likelihood(s2_b2_likelihood),
                's2_b2_severity': replace_values_severity(s2_b2_severity),
                's2_b2_option_reduce_risk': s2_b2_option_reduce_risk,
                's2_b2_overall_risk': overall_risk_logic(replace_values_likelihood(s2_b2_likelihood),
                                                         replace_values_severity(s2_b2_severity)),
                's2_b2_effect_on_risk': effect_on_risk_logic(s2_b2_effect_on_risk),
                's2_b2_residual_risk': residual_risk_logic(s2_b2_residual_risk),
                's2_b2_measure_approved': measure_approved_logic(s2_b2_measure_approved)

            }

        s2_b3_likelihood = request.POST.get('2_3_likelihood')  # Step2 box3
        s2_b3_severity = request.POST.get('2_3_severity')  # Step2 box3
        s2_b3_option_reduce_risk = request.POST.get('2_3_option_reduce_risk')
        s2_b3_effect_on_risk = request.POST.get('2_3_effect_on_risk')
        s2_b3_residual_risk = request.POST.get('2_3_residual_risk')
        s2_b3_measure_approved = request.POST.get('2_3_measure_approved')

        if s2_b3_likelihood and s2_b3_severity is not None:
            liklihood_dict['s2_b3'] = {
                's2_b3_likelihood': replace_values_likelihood(s2_b3_likelihood),
                's2_b3_severity': replace_values_severity(s2_b3_severity),
                's2_b3_option_reduce_risk': s2_b3_option_reduce_risk,
                's2_b3_overall_risk': overall_risk_logic(replace_values_likelihood(s2_b3_likelihood),
                                                         replace_values_severity(s2_b3_severity)),
                's2_b3_effect_on_risk': effect_on_risk_logic(s2_b3_effect_on_risk),
                's2_b3_residual_risk': residual_risk_logic(s2_b3_residual_risk),
                's2_b3_measure_approved': measure_approved_logic(s2_b3_measure_approved)

            }

        s2_b4_likelihood = request.POST.get('2_4_likelihood')  # Step2 box4
        s2_b4_severity = request.POST.get('2_4_severity')  # Step2 box4
        s2_b4_option_reduce_risk = request.POST.get('2_4_option_reduce_risk')
        s2_b4_effect_on_risk = request.POST.get('2_4_effect_on_risk')
        s2_b4_residual_risk = request.POST.get('2_4_residual_risk')
        s2_b4_measure_approved = request.POST.get('2_4_measure_approved')

        if s2_b4_likelihood and s2_b4_severity is not None:
            liklihood_dict['s2_b4'] = {
                's2_b4_likelihood': replace_values_likelihood(s2_b4_likelihood),
                's2_b4_severity': replace_values_severity(s2_b4_severity),
                's2_b4_option_reduce_risk': s2_b4_option_reduce_risk,
                's2_b4_overall_risk': overall_risk_logic(replace_values_likelihood(s2_b4_likelihood),
                                                         replace_values_severity(s2_b4_severity)),
                's2_b4_effect_on_risk': effect_on_risk_logic(s2_b4_effect_on_risk),
                's2_b4_residual_risk': residual_risk_logic(s2_b4_residual_risk),
                's2_b4_measure_approved': measure_approved_logic(s2_b4_measure_approved)

            }

        s2_b5_likelihood = request.POST.get('2_5_likelihood')  # Step2 box5
        s2_b5_severity = request.POST.get('2_5_severity')  # Step2 box5
        s2_b5_option_reduce_risk = request.POST.get('2_5_option_reduce_risk')
        s2_b5_effect_on_risk = request.POST.get('2_5_effect_on_risk')
        s2_b5_residual_risk = request.POST.get('2_5_residual_risk')
        s2_b5_measure_approved = request.POST.get('2_5_measure_approved')

        if s2_b5_likelihood and s2_b5_severity is not None:
            liklihood_dict['s2_b5'] = {
                's2_b5_likelihood': replace_values_likelihood(s2_b5_likelihood),
                's2_b5_severity': replace_values_severity(s2_b5_severity),
                's2_b5_option_reduce_risk': s2_b5_option_reduce_risk,
                's2_b5_overall_risk': overall_risk_logic(replace_values_likelihood(s2_b5_likelihood),
                                                         replace_values_severity(s2_b5_severity)),
                's2_b5_effect_on_risk': effect_on_risk_logic(s2_b5_effect_on_risk),
                's2_b5_residual_risk': residual_risk_logic(s2_b5_residual_risk),
                's2_b5_measure_approved': measure_approved_logic(s2_b5_measure_approved)

            }

        s2_b6_likelihood = request.POST.get('2_6_likelihood')  # Step2 box6
        s2_b6_severity = request.POST.get('2_6_severity')  # Step2 box6
        s2_b6_option_reduce_risk = request.POST.get('2_6_option_reduce_risk')
        s2_b6_effect_on_risk = request.POST.get('2_6_effect_on_risk')
        s2_b6_residual_risk = request.POST.get('2_6_residual_risk')
        s2_b6_measure_approved = request.POST.get('2_6_measure_approved')

        if s2_b6_likelihood and s2_b6_severity is not None:
            liklihood_dict['s2_b6'] = {
                's2_b6_likelihood': replace_values_likelihood(s2_b6_likelihood),
                's2_b6_severity': replace_values_severity(s2_b6_severity),
                's2_b6_option_reduce_risk': s2_b6_option_reduce_risk,
                's2_b6_overall_risk': overall_risk_logic(replace_values_likelihood(s2_b6_likelihood),
                                                         replace_values_severity(s2_b6_severity)),
                's2_b6_effect_on_risk': effect_on_risk_logic(s2_b6_effect_on_risk),
                's2_b6_residual_risk': residual_risk_logic(s2_b6_residual_risk),
                's2_b6_measure_approved': measure_approved_logic(s2_b6_measure_approved)

            }

        s2_b7_likelihood = request.POST.get('2_7_likelihood')  # Step2 box7
        s2_b7_severity = request.POST.get('2_7_severity')  # Step2 box7
        s2_b7_option_reduce_risk = request.POST.get('2_7_option_reduce_risk')
        s2_b7_effect_on_risk = request.POST.get('2_7_effect_on_risk')
        s2_b7_residual_risk = request.POST.get('2_7_residual_risk')
        s2_b7_measure_approved = request.POST.get('2_7_measure_approved')

        if s2_b7_likelihood and s2_b7_severity is not None:
            liklihood_dict['s2_b7'] = {
                's2_b7_likelihood': replace_values_likelihood(s2_b7_likelihood),
                's2_b7_severity': replace_values_severity(s2_b7_severity),
                's2_b7_option_reduce_risk': s2_b7_option_reduce_risk,
                's2_b7_overall_risk': overall_risk_logic(replace_values_likelihood(s2_b7_likelihood),
                                                         replace_values_severity(s2_b7_severity)),
                's2_b7_effect_on_risk': effect_on_risk_logic(s2_b7_effect_on_risk),
                's2_b7_residual_risk': residual_risk_logic(s2_b7_residual_risk),
                's2_b7_measure_approved': measure_approved_logic(s2_b7_measure_approved)

            }

        s2_b8_likelihood = request.POST.get('2_8_likelihood')  # Step2 box8
        s2_b8_severity = request.POST.get('2_8_severity')  # Step2 box8
        s2_b8_option_reduce_risk = request.POST.get('2_8_option_reduce_risk')
        s2_b8_effect_on_risk = request.POST.get('2_8_effect_on_risk')
        s2_b8_residual_risk = request.POST.get('2_8_residual_risk')
        s2_b8_measure_approved = request.POST.get('2_8_measure_approved')

        if s2_b8_likelihood and s2_b8_severity is not None:
            liklihood_dict['s2_b8'] = {
                's2_b8_likelihood': replace_values_likelihood(s2_b8_likelihood),
                's2_b8_severity': replace_values_severity(s2_b8_severity),
                's2_b8_option_reduce_risk': s2_b8_option_reduce_risk,
                's2_b8_overall_risk': overall_risk_logic(replace_values_likelihood(s2_b8_likelihood),
                                                         replace_values_severity(s2_b8_severity)),
                's2_b8_effect_on_risk': effect_on_risk_logic(s2_b8_effect_on_risk),
                's2_b8_residual_risk': residual_risk_logic(s2_b8_residual_risk),
                's2_b8_measure_approved': measure_approved_logic(s2_b8_measure_approved)

            }

        s2_b9_likelihood = request.POST.get('2_9_likelihood')  # Step2 box9
        s2_b9_severity = request.POST.get('2_9_severity')  # Step2 box9
        s2_b9_option_reduce_risk = request.POST.get('2_9_option_reduce_risk')
        s2_b9_effect_on_risk = request.POST.get('2_9_effect_on_risk')
        s2_b9_residual_risk = request.POST.get('2_9_residual_risk')
        s2_b9_measure_approved = request.POST.get('2_9_measure_approved')

        if s2_b9_likelihood and s2_b9_severity is not None:
            liklihood_dict['s2_b9'] = {
                's2_b9_likelihood': replace_values_likelihood(s2_b9_likelihood),
                's2_b9_severity': replace_values_severity(s2_b9_severity),
                's2_b9_option_reduce_risk': s2_b9_option_reduce_risk,
                's2_b9_overall_risk': overall_risk_logic(replace_values_likelihood(s2_b9_likelihood),
                                                         replace_values_severity(s2_b9_severity)),
                's2_b9_effect_on_risk': effect_on_risk_logic(s2_b9_effect_on_risk),
                's2_b9_residual_risk': residual_risk_logic(s2_b9_residual_risk),
                's2_b9_measure_approved': measure_approved_logic(s2_b9_measure_approved)

            }
        excel_question_dict = get_excel_risk_questions(request)
        text_to_excel(request, excel_question_dict, liklihood_dict)
        request.session['liklihood_dict'] = liklihood_dict
        context = {
            'risk_score1': risk_score1,
            'risk_score2': risk_score2,
            'risk_score3': risk_score3,
            'risk_score4': risk_score4,
            'risk_score5': risk_score5,
            'risk_score6': risk_score6,
            'risk_score7': risk_score7,
            'risk_score8': risk_score8,
            'form1_percentage': form1_percentage,
            'form2_percentage': form2_percentage,
            'form3_percentage': form3_percentage,
            'form4_percentage': form4_percentage,
            'form5_percentage': form5_percentage,
            'form6_percentage': form6_percentage,
            'form7_percentage': form7_percentage,
            'form8_percentage': form8_percentage,
            'total_no_of_risk': request.session.get('total_no_of_risk'),
            'name_of_controller': request.session.get('manager'),
            'name_of_dpo': request.session.get('name_of_DPO'),
            'title_of_dpo': request.session.get('title_of_DPO')
        }
        return render(request, 'risk_summary.html', context)


def risk_summary_box_3(request):
    if request.method == 'POST':
        risk_score1 = request.session.get('risk_score1')
        risk_score2 = request.session.get('risk_score2')
        risk_score3 = request.session.get('risk_score3')
        risk_score4 = request.session.get('risk_score4')
        risk_score5 = request.session.get('risk_score5')
        risk_score6 = request.session.get('risk_score6')
        risk_score7 = request.session.get('risk_score7')
        risk_score8 = request.session.get('risk_score8')
        form1_percentage = request.session.get('form1_percentage')
        form2_percentage = request.session.get('form2_percentage')
        form3_percentage = request.session.get('form3_percentage')
        form4_percentage = request.session.get('form4_percentage')
        form5_percentage = request.session.get('form5_percentage')
        form6_percentage = request.session.get('form6_percentage')
        form7_percentage = request.session.get('form7_percentage')
        form8_percentage = request.session.get('form8_percentage')

        s3_b1_likelihood = request.POST.get('3_1_likelihood')  # Step3 box1
        s3_b1_severity = request.POST.get('3_1_severity')  # Step3 box1
        s3_b1_option_reduce_risk = request.POST.get('3_1_option_reduce_risk')
        s3_b1_effect_on_risk = request.POST.get('3_1_effect_on_risk')
        s3_b1_residual_risk = request.POST.get('3_1_residual_risk')
        s3_b1_measure_approved = request.POST.get('3_1_measure_approved')

        if s3_b1_likelihood and s3_b1_severity is not None:
            liklihood_dict['s3_b1'] = {
                's3_b1_likelihood': replace_values_likelihood(s3_b1_likelihood),
                's3_b1_severity': replace_values_severity(s3_b1_severity),
                's3_b1_option_reduce_risk': s3_b1_option_reduce_risk,
                's3_b1_overall_risk': overall_risk_logic(replace_values_likelihood(s3_b1_likelihood),
                                                         replace_values_severity(s3_b1_severity)),
                's3_b1_effect_on_risk': effect_on_risk_logic(s3_b1_effect_on_risk),
                's3_b1_residual_risk': residual_risk_logic(s3_b1_residual_risk),
                's3_b1_measure_approved': measure_approved_logic(s3_b1_measure_approved)

            }

        s3_b2_likelihood = request.POST.get('3_2_likelihood')  # Step3 box2
        s3_b2_severity = request.POST.get('3_2_severity')  # Step3 box2
        s3_b2_option_reduce_risk = request.POST.get('3_2_option_reduce_risk')
        s3_b2_effect_on_risk = request.POST.get('3_2_effect_on_risk')
        s3_b2_residual_risk = request.POST.get('3_2_residual_risk')
        s3_b2_measure_approved = request.POST.get('3_2_measure_approved')

        if s3_b2_likelihood and s3_b2_severity is not None:
            liklihood_dict['s3_b2'] = {
                's3_b2_likelihood': replace_values_likelihood(s3_b2_likelihood),
                's3_b2_severity': replace_values_severity(s3_b2_severity),
                's3_b2_option_reduce_risk': s3_b2_option_reduce_risk,
                's3_b2_overall_risk': overall_risk_logic(replace_values_likelihood(s3_b2_likelihood),
                                                         replace_values_severity(s3_b2_severity)),
                's3_b2_effect_on_risk': effect_on_risk_logic(s3_b2_effect_on_risk),
                's3_b2_residual_risk': residual_risk_logic(s3_b2_residual_risk),
                's3_b2_measure_approved': measure_approved_logic(s3_b2_measure_approved)

            }

        s3_b3_likelihood = request.POST.get('3_3_likelihood')  # Step3 box3
        s3_b3_severity = request.POST.get('3_3_severity')  # Step3 box3
        s3_b3_option_reduce_risk = request.POST.get('3_3_option_reduce_risk')
        s3_b3_effect_on_risk = request.POST.get('3_3_effect_on_risk')
        s3_b3_residual_risk = request.POST.get('3_3_residual_risk')
        s3_b3_measure_approved = request.POST.get('3_3_measure_approved')

        if s3_b3_likelihood and s3_b3_severity is not None:
            liklihood_dict['s3_b3'] = {
                's3_b3_likelihood': replace_values_likelihood(s3_b3_likelihood),
                's3_b3_severity': replace_values_severity(s3_b3_severity),
                's3_b3_option_reduce_risk': s3_b3_option_reduce_risk,
                's3_b3_overall_risk': overall_risk_logic(replace_values_likelihood(s3_b3_likelihood),
                                                         replace_values_severity(s3_b3_severity)),
                's3_b3_effect_on_risk': effect_on_risk_logic(s3_b3_effect_on_risk),
                's3_b3_residual_risk': residual_risk_logic(s3_b3_residual_risk),
                's3_b3_measure_approved': measure_approved_logic(s3_b3_measure_approved)

            }

        s3_b4_likelihood = request.POST.get('3_4_likelihood')  # Step3 box4
        s3_b4_severity = request.POST.get('3_4_severity')  # Step3 box4
        s3_b4_option_reduce_risk = request.POST.get('3_4_option_reduce_risk')
        s3_b4_effect_on_risk = request.POST.get('3_4_effect_on_risk')
        s3_b4_residual_risk = request.POST.get('3_4_residual_risk')
        s3_b4_measure_approved = request.POST.get('3_4_measure_approved')

        if s3_b4_likelihood and s3_b4_severity is not None:
            liklihood_dict['s3_b4'] = {
                's3_b4_likelihood': replace_values_likelihood(s3_b4_likelihood),
                's3_b4_severity': replace_values_severity(s3_b4_severity),
                's3_b4_option_reduce_risk': s3_b4_option_reduce_risk,
                's3_b4_overall_risk': overall_risk_logic(replace_values_likelihood(s3_b4_likelihood),
                                                         replace_values_severity(s3_b4_severity)),
                's3_b4_effect_on_risk': effect_on_risk_logic(s3_b4_effect_on_risk),
                's3_b4_residual_risk': residual_risk_logic(s3_b4_residual_risk),
                's3_b4_measure_approved': measure_approved_logic(s3_b4_measure_approved)

            }

        s3_b5_likelihood = request.POST.get('3_5_likelihood')  # Step3 box5
        s3_b5_severity = request.POST.get('3_5_severity')  # Step3 box5
        s3_b5_option_reduce_risk = request.POST.get('3_5_option_reduce_risk')
        s3_b5_effect_on_risk = request.POST.get('3_5_effect_on_risk')
        s3_b5_residual_risk = request.POST.get('3_5_residual_risk')
        s3_b5_measure_approved = request.POST.get('3_5_measure_approved')

        if s3_b5_likelihood and s3_b5_severity is not None:
            liklihood_dict['s3_b5'] = {
                's3_b5_likelihood': replace_values_likelihood(s3_b5_likelihood),
                's3_b5_severity': replace_values_severity(s3_b5_severity),
                's3_b5_option_reduce_risk': s3_b5_option_reduce_risk,
                's3_b5_overall_risk': overall_risk_logic(replace_values_likelihood(s3_b5_likelihood),
                                                         replace_values_severity(s3_b5_severity)),
                's3_b5_effect_on_risk': effect_on_risk_logic(s3_b5_effect_on_risk),
                's3_b5_residual_risk': residual_risk_logic(s3_b5_residual_risk),
                's3_b5_measure_approved': measure_approved_logic(s3_b5_measure_approved)

            }

        s3_b6_likelihood = request.POST.get('3_6_likelihood')  # Step3 box6
        s3_b6_severity = request.POST.get('3_6_severity')  # Step3 box6
        s3_b6_option_reduce_risk = request.POST.get('3_6_option_reduce_risk')
        s3_b6_effect_on_risk = request.POST.get('3_6_effect_on_risk')
        s3_b6_residual_risk = request.POST.get('3_6_residual_risk')
        s3_b6_measure_approved = request.POST.get('3_6_measure_approved')

        if s3_b6_likelihood and s3_b6_severity is not None:
            liklihood_dict['s3_b6'] = {
                's3_b6_likelihood': replace_values_likelihood(s3_b6_likelihood),
                's3_b6_severity': replace_values_severity(s3_b6_severity),
                's3_b6_option_reduce_risk': s3_b6_option_reduce_risk,
                's3_b6_overall_risk': overall_risk_logic(replace_values_likelihood(s3_b6_likelihood),
                                                         replace_values_severity(s3_b6_severity)),
                's3_b6_effect_on_risk': effect_on_risk_logic(s3_b6_effect_on_risk),
                's3_b6_residual_risk': residual_risk_logic(s3_b6_residual_risk),
                's3_b6_measure_approved': measure_approved_logic(s3_b6_measure_approved)

            }

        s3_b7_likelihood = request.POST.get('3_7_likelihood')  # Step3 box7
        s3_b7_severity = request.POST.get('3_7_severity')  # Step3 box7
        s3_b7_option_reduce_risk = request.POST.get('3_7_option_reduce_risk')
        s3_b7_effect_on_risk = request.POST.get('3_7_effect_on_risk')
        s3_b7_residual_risk = request.POST.get('3_7_residual_risk')
        s3_b7_measure_approved = request.POST.get('3_7_measure_approved')

        if s3_b7_likelihood and s3_b7_severity is not None:
            liklihood_dict['s3_b7'] = {
                's3_b7_likelihood': replace_values_likelihood(s3_b7_likelihood),
                's3_b7_severity': replace_values_severity(s3_b7_severity),
                's3_b7_option_reduce_risk': s3_b7_option_reduce_risk,
                's3_b7_overall_risk': overall_risk_logic(replace_values_likelihood(s3_b7_likelihood),
                                                         replace_values_severity(s3_b7_severity)),
                's3_b7_effect_on_risk': effect_on_risk_logic(s3_b7_effect_on_risk),
                's3_b7_residual_risk': residual_risk_logic(s3_b7_residual_risk),
                's3_b7_measure_approved': measure_approved_logic(s3_b7_measure_approved)

            }

        s3_b8_likelihood = request.POST.get('3_8_likelihood')  # Step3 box8
        s3_b8_severity = request.POST.get('3_8_severity')  # Step3 box8
        s3_b8_option_reduce_risk = request.POST.get('3_8_option_reduce_risk')
        s3_b8_effect_on_risk = request.POST.get('3_8_effect_on_risk')
        s3_b8_residual_risk = request.POST.get('3_8_residual_risk')
        s3_b8_measure_approved = request.POST.get('3_8_measure_approved')

        if s3_b8_likelihood and s3_b8_severity is not None:
            liklihood_dict['s3_b8'] = {
                's3_b8_likelihood': replace_values_likelihood(s3_b8_likelihood),
                's3_b8_severity': replace_values_severity(s3_b8_severity),
                's3_b8_option_reduce_risk': s3_b8_option_reduce_risk,
                's3_b8_overall_risk': overall_risk_logic(replace_values_likelihood(s3_b8_likelihood),
                                                         replace_values_severity(s3_b8_severity)),
                's3_b8_effect_on_risk': effect_on_risk_logic(s3_b8_effect_on_risk),
                's3_b8_residual_risk': residual_risk_logic(s3_b8_residual_risk),
                's3_b8_measure_approved': measure_approved_logic(s3_b8_measure_approved)

            }

        s3_b9_likelihood = request.POST.get('3_9_likelihood')  # Step3 box9
        s3_b9_severity = request.POST.get('3_9_severity')  # Step3 box9
        s3_b9_option_reduce_risk = request.POST.get('3_9_option_reduce_risk')
        s3_b9_effect_on_risk = request.POST.get('3_9_effect_on_risk')
        s3_b9_residual_risk = request.POST.get('3_9_residual_risk')
        s3_b9_measure_approved = request.POST.get('3_9_measure_approved')

        if s3_b9_likelihood and s3_b9_severity is not None:
            liklihood_dict['s3_b9'] = {
                's3_b9_likelihood': replace_values_likelihood(s3_b9_likelihood),
                's3_b9_severity': replace_values_severity(s3_b9_severity),
                's3_b9_option_reduce_risk': s3_b9_option_reduce_risk,
                's3_b9_overall_risk': overall_risk_logic(replace_values_likelihood(s3_b9_likelihood),
                                                         replace_values_severity(s3_b9_severity)),
                's3_b9_effect_on_risk': effect_on_risk_logic(s3_b9_effect_on_risk),
                's3_b9_residual_risk': residual_risk_logic(s3_b9_residual_risk),
                's3_b9_measure_approved': measure_approved_logic(s3_b9_measure_approved)

            }

        s3_b10_likelihood = request.POST.get('3_10_likelihood')  # Step3 box10
        s3_b10_severity = request.POST.get('3_10_severity')  # Step3 box10
        s3_b10_option_reduce_risk = request.POST.get('3_10_option_reduce_risk')
        s3_b10_effect_on_risk = request.POST.get('3_10_effect_on_risk')
        s3_b10_residual_risk = request.POST.get('3_10_residual_risk')
        s3_b10_measure_approved = request.POST.get('3_10_measure_approved')

        if s3_b10_likelihood and s3_b10_severity is not None:
            liklihood_dict['s3_b10'] = {
                's3_b10_likelihood': replace_values_likelihood(s3_b10_likelihood),
                's3_b10_severity': replace_values_severity(s3_b10_severity),
                's3_b10_option_reduce_risk': s3_b10_option_reduce_risk,
                's3_b10_overall_risk': overall_risk_logic(replace_values_likelihood(s3_b10_likelihood),
                                                         replace_values_severity(s3_b10_severity)),
                's3_b10_effect_on_risk': effect_on_risk_logic(s3_b10_effect_on_risk),
                's3_b10_residual_risk': residual_risk_logic(s3_b10_residual_risk),
                's3_b10_measure_approved': measure_approved_logic(s3_b10_measure_approved)

            }

        s3_b11_likelihood = request.POST.get('3_11_likelihood')  # Step3 box11
        s3_b11_severity = request.POST.get('3_11_severity')  # Step3 box11
        s3_b11_option_reduce_risk = request.POST.get('3_11_option_reduce_risk')
        s3_b11_effect_on_risk = request.POST.get('3_11_effect_on_risk')
        s3_b11_residual_risk = request.POST.get('3_11_residual_risk')
        s3_b11_measure_approved = request.POST.get('3_11_measure_approved')

        if s3_b11_likelihood and s3_b11_severity is not None:
            liklihood_dict['s3_b11'] = {
                's3_b11_likelihood': replace_values_likelihood(s3_b11_likelihood),
                's3_b11_severity': replace_values_severity(s3_b11_severity),
                's3_b11_option_reduce_risk': s3_b11_option_reduce_risk,
                's3_b11_overall_risk': overall_risk_logic(replace_values_likelihood(s3_b11_likelihood),
                                                          replace_values_severity(s3_b11_severity)),
                's3_b11_effect_on_risk': effect_on_risk_logic(s3_b11_effect_on_risk),
                's3_b11_residual_risk': residual_risk_logic(s3_b11_residual_risk),
                's3_b11_measure_approved': measure_approved_logic(s3_b11_measure_approved)

            }

        s3_b12_likelihood = request.POST.get('3_12_likelihood')  # Step3 box12
        s3_b12_severity = request.POST.get('3_12_severity')  # Step3 box12
        s3_b12_option_reduce_risk = request.POST.get('3_12_option_reduce_risk')
        s3_b12_effect_on_risk = request.POST.get('3_12_effect_on_risk')
        s3_b12_residual_risk = request.POST.get('3_12_residual_risk')
        s3_b12_measure_approved = request.POST.get('3_12_measure_approved')

        if s3_b12_likelihood and s3_b12_severity is not None:
            liklihood_dict['s3_b12'] = {
                's3_b12_likelihood': replace_values_likelihood(s3_b12_likelihood),
                's3_b12_severity': replace_values_severity(s3_b12_severity),
                's3_b12_option_reduce_risk': s3_b12_option_reduce_risk,
                's3_b12_overall_risk': overall_risk_logic(replace_values_likelihood(s3_b12_likelihood),
                                                          replace_values_severity(s3_b12_severity)),
                's3_b12_effect_on_risk': effect_on_risk_logic(s3_b12_effect_on_risk),
                's3_b12_residual_risk': residual_risk_logic(s3_b12_residual_risk),
                's3_b12_measure_approved': measure_approved_logic(s3_b12_measure_approved)

            }

        excel_question_dict = get_excel_risk_questions(request)
        text_to_excel(request, excel_question_dict, liklihood_dict)
        request.session['liklihood_dict'] = liklihood_dict
        context = {
            'risk_score1': risk_score1,
            'risk_score2': risk_score2,
            'risk_score3': risk_score3,
            'risk_score4': risk_score4,
            'risk_score5': risk_score5,
            'risk_score6': risk_score6,
            'risk_score7': risk_score7,
            'risk_score8': risk_score8,
            'form1_percentage': form1_percentage,
            'form2_percentage': form2_percentage,
            'form3_percentage': form3_percentage,
            'form4_percentage': form4_percentage,
            'form5_percentage': form5_percentage,
            'form6_percentage': form6_percentage,
            'form7_percentage': form7_percentage,
            'form8_percentage': form8_percentage,
            'total_no_of_risk': request.session.get('total_no_of_risk'),
            'name_of_controller': request.session.get('manager'),
            'name_of_dpo': request.session.get('name_of_DPO'),
            'title_of_dpo': request.session.get('title_of_DPO')
        }
        return render(request, 'risk_summary.html', context)



def risk_summary_box_4(request):
    if request.method == 'POST':
        risk_score1 = request.session.get('risk_score1')
        risk_score2 = request.session.get('risk_score2')
        risk_score3 = request.session.get('risk_score3')
        risk_score4 = request.session.get('risk_score4')
        risk_score5 = request.session.get('risk_score5')
        risk_score6 = request.session.get('risk_score6')
        risk_score7 = request.session.get('risk_score7')
        risk_score8 = request.session.get('risk_score8')
        form1_percentage = request.session.get('form1_percentage')
        form2_percentage = request.session.get('form2_percentage')
        form3_percentage = request.session.get('form3_percentage')
        form4_percentage = request.session.get('form4_percentage')
        form5_percentage = request.session.get('form5_percentage')
        form6_percentage = request.session.get('form6_percentage')
        form7_percentage = request.session.get('form7_percentage')
        form8_percentage = request.session.get('form8_percentage')

        s4_b1_likelihood = request.POST.get('4_1_likelihood')  # Step4 box1
        s4_b1_severity = request.POST.get('4_1_severity')  # Step4 box1
        s4_b1_option_reduce_risk = request.POST.get('4_1_option_reduce_risk')
        s4_b1_effect_on_risk = request.POST.get('4_1_effect_on_risk')
        s4_b1_residual_risk = request.POST.get('4_1_residual_risk')
        s4_b1_measure_approved = request.POST.get('4_1_measure_approved')

        if s4_b1_likelihood and s4_b1_severity is not None:
            liklihood_dict['s4_b1'] = {
                's4_b1_likelihood': replace_values_likelihood(s4_b1_likelihood),
                's4_b1_severity': replace_values_severity(s4_b1_severity),
                's4_b1_option_reduce_risk': s4_b1_option_reduce_risk,
                's4_b1_overall_risk': overall_risk_logic(replace_values_likelihood(s4_b1_likelihood),
                                                          replace_values_severity(s4_b1_severity)),
                's4_b1_effect_on_risk': effect_on_risk_logic(s4_b1_effect_on_risk),
                's4_b1_residual_risk': residual_risk_logic(s4_b1_residual_risk),
                's4_b1_measure_approved': measure_approved_logic(s4_b1_measure_approved)

            }

        s4_b2_likelihood = request.POST.get('4_2_likelihood')  # Step4 box2
        s4_b2_severity = request.POST.get('4_2_severity')  # Step4 box2
        s4_b2_option_reduce_risk = request.POST.get('4_2_option_reduce_risk')
        s4_b2_effect_on_risk = request.POST.get('4_2_effect_on_risk')
        s4_b2_residual_risk = request.POST.get('4_2_residual_risk')
        s4_b2_measure_approved = request.POST.get('4_2_measure_approved')

        if s4_b2_likelihood and s4_b2_severity is not None:
            liklihood_dict['s4_b2'] = {
                's4_b2_likelihood': replace_values_likelihood(s4_b2_likelihood),
                's4_b2_severity': replace_values_severity(s4_b2_severity),
                's4_b2_option_reduce_risk': s4_b2_option_reduce_risk,
                's4_b2_overall_risk': overall_risk_logic(replace_values_likelihood(s4_b2_likelihood),
                                                         replace_values_severity(s4_b2_severity)),
                's4_b2_effect_on_risk': effect_on_risk_logic(s4_b2_effect_on_risk),
                's4_b2_residual_risk': residual_risk_logic(s4_b2_residual_risk),
                's4_b2_measure_approved': measure_approved_logic(s4_b2_measure_approved)

            }

        s4_b3_likelihood = request.POST.get('4_3_likelihood')  # Step4 box3
        s4_b3_severity = request.POST.get('4_3_severity')  # Step4 box3
        s4_b3_option_reduce_risk = request.POST.get('4_3_option_reduce_risk')
        s4_b3_effect_on_risk = request.POST.get('4_3_effect_on_risk')
        s4_b3_residual_risk = request.POST.get('4_3_residual_risk')
        s4_b3_measure_approved = request.POST.get('4_3_measure_approved')

        if s4_b3_likelihood and s4_b3_severity is not None:
            liklihood_dict['s4_b3'] = {
                's4_b3_likelihood': replace_values_likelihood(s4_b3_likelihood),
                's4_b3_severity': replace_values_severity(s4_b3_severity),
                's4_b3_option_reduce_risk': s4_b3_option_reduce_risk,
                's4_b3_overall_risk': overall_risk_logic(replace_values_likelihood(s4_b3_likelihood),
                                                         replace_values_severity(s4_b3_severity)),
                's4_b3_effect_on_risk': effect_on_risk_logic(s4_b3_effect_on_risk),
                's4_b3_residual_risk': residual_risk_logic(s4_b3_residual_risk),
                's4_b3_measure_approved': measure_approved_logic(s4_b3_measure_approved)

            }

        s4_b4_likelihood = request.POST.get('4_4_likelihood')  # Step4 box4
        s4_b4_severity = request.POST.get('4_4_severity')  # Step4 box4
        s4_b4_option_reduce_risk = request.POST.get('4_4_option_reduce_risk')
        s4_b4_effect_on_risk = request.POST.get('4_4_effect_on_risk')
        s4_b4_residual_risk = request.POST.get('4_4_residual_risk')
        s4_b4_measure_approved = request.POST.get('4_4_measure_approved')

        if s4_b4_likelihood and s4_b4_severity is not None:
            liklihood_dict['s4_b4'] = {
                's4_b4_likelihood': replace_values_likelihood(s4_b4_likelihood),
                's4_b4_severity': replace_values_severity(s4_b4_severity),
                's4_b4_option_reduce_risk': s4_b4_option_reduce_risk,
                's4_b4_overall_risk': overall_risk_logic(replace_values_likelihood(s4_b4_likelihood),
                                                         replace_values_severity(s4_b4_severity)),
                's4_b4_effect_on_risk': effect_on_risk_logic(s4_b4_effect_on_risk),
                's4_b4_residual_risk': residual_risk_logic(s4_b4_residual_risk),
                's4_b4_measure_approved': measure_approved_logic(s4_b4_measure_approved)

            }

        s4_b5_likelihood = request.POST.get('4_5_likelihood')  # Step4 box5
        s4_b5_severity = request.POST.get('4_5_severity')  # Step4 box5
        s4_b5_option_reduce_risk = request.POST.get('4_5_option_reduce_risk')
        s4_b5_effect_on_risk = request.POST.get('4_5_effect_on_risk')
        s4_b5_residual_risk = request.POST.get('4_5_residual_risk')
        s4_b5_measure_approved = request.POST.get('4_5_measure_approved')

        if s4_b5_likelihood and s4_b5_severity is not None:
            liklihood_dict['s4_b5'] = {
                's4_b5_likelihood': replace_values_likelihood(s4_b5_likelihood),
                's4_b5_severity': replace_values_severity(s4_b5_severity),
                's4_b5_option_reduce_risk': s4_b5_option_reduce_risk,
                's4_b5_overall_risk': overall_risk_logic(replace_values_likelihood(s4_b5_likelihood),
                                                         replace_values_severity(s4_b5_severity)),
                's4_b5_effect_on_risk': effect_on_risk_logic(s4_b5_effect_on_risk),
                's4_b5_residual_risk': residual_risk_logic(s4_b5_residual_risk),
                's4_b5_measure_approved': measure_approved_logic(s4_b5_measure_approved)

            }

        s4_b6_likelihood = request.POST.get('4_6_likelihood')  # Step4 box6
        s4_b6_severity = request.POST.get('4_6_severity')  # Step4 box6
        s4_b6_option_reduce_risk = request.POST.get('4_6_option_reduce_risk')
        s4_b6_effect_on_risk = request.POST.get('4_6_effect_on_risk')
        s4_b6_residual_risk = request.POST.get('4_6_residual_risk')
        s4_b6_measure_approved = request.POST.get('4_6_measure_approved')

        if s4_b6_likelihood and s4_b6_severity is not None:
            liklihood_dict['s4_b6'] = {
                's4_b6_likelihood': replace_values_likelihood(s4_b6_likelihood),
                's4_b6_severity': replace_values_severity(s4_b6_severity),
                's4_b6_option_reduce_risk': s4_b6_option_reduce_risk,
                's4_b6_overall_risk': overall_risk_logic(replace_values_likelihood(s4_b6_likelihood),
                                                         replace_values_severity(s4_b6_severity)),
                's4_b6_effect_on_risk': effect_on_risk_logic(s4_b6_effect_on_risk),
                's4_b6_residual_risk': residual_risk_logic(s4_b6_residual_risk),
                's4_b6_measure_approved': measure_approved_logic(s4_b6_measure_approved)

            }

        s4_b7_likelihood = request.POST.get('4_7_likelihood')  # Step4 box7
        s4_b7_severity = request.POST.get('4_7_severity')  # Step4 box7
        s4_b7_option_reduce_risk = request.POST.get('4_7_option_reduce_risk')
        s4_b7_effect_on_risk = request.POST.get('4_7_effect_on_risk')
        s4_b7_residual_risk = request.POST.get('4_7_residual_risk')
        s4_b7_measure_approved = request.POST.get('4_7_measure_approved')

        if s4_b7_likelihood and s4_b7_severity is not None:
            liklihood_dict['s4_b7'] = {
                's4_b7_likelihood': replace_values_likelihood(s4_b7_likelihood),
                's4_b7_severity': replace_values_severity(s4_b7_severity),
                's4_b7_option_reduce_risk': s4_b7_option_reduce_risk,
                's4_b7_overall_risk': overall_risk_logic(replace_values_likelihood(s4_b7_likelihood),
                                                         replace_values_severity(s4_b7_severity)),
                's4_b7_effect_on_risk': effect_on_risk_logic(s4_b7_effect_on_risk),
                's4_b7_residual_risk': residual_risk_logic(s4_b7_residual_risk),
                's4_b7_measure_approved': measure_approved_logic(s4_b7_measure_approved)

            }

        s4_b8_likelihood = request.POST.get('4_8_likelihood')  # Step4 box8
        s4_b8_severity = request.POST.get('4_8_severity')  # Step4 box8
        s4_b8_option_reduce_risk = request.POST.get('4_8_option_reduce_risk')
        s4_b8_effect_on_risk = request.POST.get('4_8_effect_on_risk')
        s4_b8_residual_risk = request.POST.get('4_8_residual_risk')
        s4_b8_measure_approved = request.POST.get('4_8_measure_approved')

        if s4_b8_likelihood and s4_b8_severity is not None:
            liklihood_dict['s4_b8'] = {
                's4_b8_likelihood': replace_values_likelihood(s4_b8_likelihood),
                's4_b8_severity': replace_values_severity(s4_b8_severity),
                's4_b8_option_reduce_risk': s4_b8_option_reduce_risk,
                's4_b8_overall_risk': overall_risk_logic(replace_values_likelihood(s4_b8_likelihood),
                                                         replace_values_severity(s4_b8_severity)),
                's4_b8_effect_on_risk': effect_on_risk_logic(s4_b8_effect_on_risk),
                's4_b8_residual_risk': residual_risk_logic(s4_b8_residual_risk),
                's4_b8_measure_approved': measure_approved_logic(s4_b8_measure_approved)

            }

        s4_b9_likelihood = request.POST.get('4_9_likelihood')  # Step4 box9
        s4_b9_severity = request.POST.get('4_9_severity')  # Step4 box9
        s4_b9_option_reduce_risk = request.POST.get('4_9_option_reduce_risk')
        s4_b9_effect_on_risk = request.POST.get('4_9_effect_on_risk')
        s4_b9_residual_risk = request.POST.get('4_9_residual_risk')
        s4_b9_measure_approved = request.POST.get('4_9_measure_approved')

        if s4_b9_likelihood and s4_b9_severity is not None:
            liklihood_dict['s4_b9'] = {
                's4_b9_likelihood': replace_values_likelihood(s4_b9_likelihood),
                's4_b9_severity': replace_values_severity(s4_b9_severity),
                's4_b9_option_reduce_risk': s4_b9_option_reduce_risk,
                's4_b9_overall_risk': overall_risk_logic(replace_values_likelihood(s4_b9_likelihood),
                                                         replace_values_severity(s4_b9_severity)),
                's4_b9_effect_on_risk': effect_on_risk_logic(s4_b9_effect_on_risk),
                's4_b9_residual_risk': residual_risk_logic(s4_b9_residual_risk),
                's4_b9_measure_approved': measure_approved_logic(s4_b9_measure_approved)

            }

        s4_b10_likelihood = request.POST.get('4_10_likelihood')  # Step4 box10
        s4_b10_severity = request.POST.get('4_10_severity')  # Step4 box10
        s4_b10_option_reduce_risk = request.POST.get('4_10_option_reduce_risk')
        s4_b10_effect_on_risk = request.POST.get('4_10_effect_on_risk')
        s4_b10_residual_risk = request.POST.get('4_10_residual_risk')
        s4_b10_measure_approved = request.POST.get('4_10_measure_approved')

        if s4_b10_likelihood and s4_b10_severity is not None:
            liklihood_dict['s4_b10'] = {
                's4_b10_likelihood': replace_values_likelihood(s4_b10_likelihood),
                's4_b10_severity': replace_values_severity(s4_b10_severity),
                's4_b10_option_reduce_risk': s4_b10_option_reduce_risk,
                's4_b10_overall_risk': overall_risk_logic(replace_values_likelihood(s4_b10_likelihood),
                                                         replace_values_severity(s4_b10_severity)),
                's4_b10_effect_on_risk': effect_on_risk_logic(s4_b10_effect_on_risk),
                's4_b10_residual_risk': residual_risk_logic(s4_b10_residual_risk),
                's4_b10_measure_approved': measure_approved_logic(s4_b10_measure_approved)

            }

        s4_b11_likelihood = request.POST.get('4_11_likelihood')  # Step4 box11
        s4_b11_severity = request.POST.get('4_11_severity')  # Step4 box11
        s4_b11_option_reduce_risk = request.POST.get('4_11_option_reduce_risk')
        s4_b11_effect_on_risk = request.POST.get('4_11_effect_on_risk')
        s4_b11_residual_risk = request.POST.get('4_11_residual_risk')
        s4_b11_measure_approved = request.POST.get('4_11_measure_approved')

        if s4_b11_likelihood and s4_b11_severity is not None:
            liklihood_dict['s4_b11'] = {
                's4_b11_likelihood': replace_values_likelihood(s4_b11_likelihood),
                's4_b11_severity': replace_values_severity(s4_b11_severity),
                's4_b11_option_reduce_risk': s4_b11_option_reduce_risk,
                's4_b11_overall_risk': overall_risk_logic(replace_values_likelihood(s4_b11_likelihood),
                                                          replace_values_severity(s4_b11_severity)),
                's4_b11_effect_on_risk': effect_on_risk_logic(s4_b11_effect_on_risk),
                's4_b11_residual_risk': residual_risk_logic(s4_b11_residual_risk),
                's4_b11_measure_approved': measure_approved_logic(s4_b11_measure_approved)

            }

        s4_b12_likelihood = request.POST.get('4_12_likelihood')  # Step4 box12
        s4_b12_severity = request.POST.get('4_12_severity')  # Step4 box12
        s4_b12_option_reduce_risk = request.POST.get('4_12_option_reduce_risk')
        s4_b12_effect_on_risk = request.POST.get('4_12_effect_on_risk')
        s4_b12_residual_risk = request.POST.get('4_12_residual_risk')
        s4_b12_measure_approved = request.POST.get('4_12_measure_approved')

        if s4_b12_likelihood and s4_b12_severity is not None:
            liklihood_dict['s4_b12'] = {
                's4_b12_likelihood': replace_values_likelihood(s4_b12_likelihood),
                's4_b12_severity': replace_values_severity(s4_b12_severity),
                's4_b12_option_reduce_risk': s4_b12_option_reduce_risk,
                's4_b12_overall_risk': overall_risk_logic(replace_values_likelihood(s4_b12_likelihood),
                                                          replace_values_severity(s4_b12_severity)),
                's4_b12_effect_on_risk': effect_on_risk_logic(s4_b12_effect_on_risk),
                's4_b12_residual_risk': residual_risk_logic(s4_b12_residual_risk),
                's4_b12_measure_approved': measure_approved_logic(s4_b12_measure_approved)

            }

        s4_b13_likelihood = request.POST.get('4_13_likelihood')  # Step4 box12
        s4_b13_severity = request.POST.get('4_13_severity')  # Step4 box12
        s4_b13_option_reduce_risk = request.POST.get('4_13_option_reduce_risk')
        s4_b13_effect_on_risk = request.POST.get('4_13_effect_on_risk')
        s4_b13_residual_risk = request.POST.get('4_13_residual_risk')
        s4_b13_measure_approved = request.POST.get('4_13_measure_approved')

        if s4_b13_likelihood and s4_b13_severity is not None:
            liklihood_dict['s4_b13'] = {
                's4_b13_likelihood': replace_values_likelihood(s4_b13_likelihood),
                's4_b13_severity': replace_values_severity(s4_b13_severity),
                's4_b13_option_reduce_risk': s4_b13_option_reduce_risk,
                's4_b13_overall_risk': overall_risk_logic(replace_values_likelihood(s4_b13_likelihood),
                                                          replace_values_severity(s4_b13_severity)),
                's4_b13_effect_on_risk': effect_on_risk_logic(s4_b13_effect_on_risk),
                's4_b13_residual_risk': residual_risk_logic(s4_b13_residual_risk),
                's4_b13_measure_approved': measure_approved_logic(s4_b13_measure_approved)

            }

        excel_question_dict = get_excel_risk_questions(request)
        text_to_excel(request, excel_question_dict, liklihood_dict)
        request.session['liklihood_dict'] = liklihood_dict
        context = {
            'risk_score1': risk_score1,
            'risk_score2': risk_score2,
            'risk_score3': risk_score3,
            'risk_score4': risk_score4,
            'risk_score5': risk_score5,
            'risk_score6': risk_score6,
            'risk_score7': risk_score7,
            'risk_score8': risk_score8,
            'form1_percentage': form1_percentage,
            'form2_percentage': form2_percentage,
            'form3_percentage': form3_percentage,
            'form4_percentage': form4_percentage,
            'form5_percentage': form5_percentage,
            'form6_percentage': form6_percentage,
            'form7_percentage': form7_percentage,
            'form8_percentage': form8_percentage,
            'total_no_of_risk': request.session.get('total_no_of_risk'),
            'name_of_controller': request.session.get('manager'),
            'name_of_dpo': request.session.get('name_of_DPO'),
            'title_of_dpo': request.session.get('title_of_DPO')
        }
        return render(request, 'risk_summary.html', context)


def risk_summary_box_5(request):
    if request.method == 'POST':
        risk_score1 = request.session.get('risk_score1')
        risk_score2 = request.session.get('risk_score2')
        risk_score3 = request.session.get('risk_score3')
        risk_score4 = request.session.get('risk_score4')
        risk_score5 = request.session.get('risk_score5')
        risk_score6 = request.session.get('risk_score6')
        risk_score7 = request.session.get('risk_score7')
        risk_score8 = request.session.get('risk_score8')
        form1_percentage = request.session.get('form1_percentage')
        form2_percentage = request.session.get('form2_percentage')
        form3_percentage = request.session.get('form3_percentage')
        form4_percentage = request.session.get('form4_percentage')
        form5_percentage = request.session.get('form5_percentage')
        form6_percentage = request.session.get('form6_percentage')
        form7_percentage = request.session.get('form7_percentage')
        form8_percentage = request.session.get('form8_percentage')

        s5_b1_likelihood = request.POST.get('5_1_likelihood')  # Step5 box1
        s5_b1_severity = request.POST.get('5_1_severity')  # Step5 box1
        s5_b1_option_reduce_risk = request.POST.get('5_1_option_reduce_risk')
        s5_b1_effect_on_risk = request.POST.get('5_1_effect_on_risk')
        s5_b1_residual_risk = request.POST.get('5_1_residual_risk')
        s5_b1_measure_approved = request.POST.get('5_1_measure_approved')

        if s5_b1_likelihood and s5_b1_severity is not None:
            liklihood_dict['s5_b1'] = {
                's5_b1_likelihood': replace_values_likelihood(s5_b1_likelihood),
                's5_b1_severity': replace_values_severity(s5_b1_severity),
                's5_b1_option_reduce_risk': s5_b1_option_reduce_risk,
                's5_b1_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b1_likelihood),
                                                          replace_values_severity(s5_b1_severity)),
                's5_b1_effect_on_risk': effect_on_risk_logic(s5_b1_effect_on_risk),
                's5_b1_residual_risk': residual_risk_logic(s5_b1_residual_risk),
                's5_b1_measure_approved': measure_approved_logic(s5_b1_measure_approved)

            }

        s5_b2_likelihood = request.POST.get('5_2_likelihood')  # Step5 box2
        s5_b2_severity = request.POST.get('5_2_severity')  # Step5 box2
        s5_b2_option_reduce_risk = request.POST.get('5_2_option_reduce_risk')
        s5_b2_effect_on_risk = request.POST.get('5_2_effect_on_risk')
        s5_b2_residual_risk = request.POST.get('5_2_residual_risk')
        s5_b2_measure_approved = request.POST.get('5_2_measure_approved')

        if s5_b2_likelihood and s5_b2_severity is not None:
            liklihood_dict['s5_b2'] = {
                's5_b2_likelihood': replace_values_likelihood(s5_b2_likelihood),
                's5_b2_severity': replace_values_severity(s5_b2_severity),
                's5_b2_option_reduce_risk': s5_b2_option_reduce_risk,
                's5_b2_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b2_likelihood),
                                                         replace_values_severity(s5_b2_severity)),
                's5_b2_effect_on_risk': effect_on_risk_logic(s5_b2_effect_on_risk),
                's5_b2_residual_risk': residual_risk_logic(s5_b2_residual_risk),
                's5_b2_measure_approved': measure_approved_logic(s5_b2_measure_approved)

            }

        s5_b3_likelihood = request.POST.get('5_3_likelihood')  # Step5 box3
        s5_b3_severity = request.POST.get('5_3_severity')  # Step5 box3
        s5_b3_option_reduce_risk = request.POST.get('5_3_option_reduce_risk')
        s5_b3_effect_on_risk = request.POST.get('5_3_effect_on_risk')
        s5_b3_residual_risk = request.POST.get('5_3_residual_risk')
        s5_b3_measure_approved = request.POST.get('5_3_measure_approved')

        if s5_b3_likelihood and s5_b3_severity is not None:
            liklihood_dict['s5_b3'] = {
                's5_b3_likelihood': replace_values_likelihood(s5_b3_likelihood),
                's5_b3_severity': replace_values_severity(s5_b3_severity),
                's5_b3_option_reduce_risk': s5_b3_option_reduce_risk,
                's5_b3_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b3_likelihood),
                                                         replace_values_severity(s5_b3_severity)),
                's5_b3_effect_on_risk': effect_on_risk_logic(s5_b3_effect_on_risk),
                's5_b3_residual_risk': residual_risk_logic(s5_b3_residual_risk),
                's5_b3_measure_approved': measure_approved_logic(s5_b3_measure_approved)

            }

        s5_b4_likelihood = request.POST.get('5_4_likelihood')  # Step5 box4
        s5_b4_severity = request.POST.get('5_4_severity')  # Step5 box4
        s5_b4_option_reduce_risk = request.POST.get('5_4_option_reduce_risk')
        s5_b4_effect_on_risk = request.POST.get('5_4_effect_on_risk')
        s5_b4_residual_risk = request.POST.get('5_4_residual_risk')
        s5_b4_measure_approved = request.POST.get('5_4_measure_approved')

        if s5_b4_likelihood and s5_b4_severity is not None:
            liklihood_dict['s5_b4'] = {
                's5_b4_likelihood': replace_values_likelihood(s5_b4_likelihood),
                's5_b4_severity': replace_values_severity(s5_b4_severity),
                's5_b4_option_reduce_risk': s5_b4_option_reduce_risk,
                's5_b4_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b4_likelihood),
                                                         replace_values_severity(s5_b4_severity)),
                's5_b4_effect_on_risk': effect_on_risk_logic(s5_b4_effect_on_risk),
                's5_b4_residual_risk': residual_risk_logic(s5_b4_residual_risk),
                's5_b4_measure_approved': measure_approved_logic(s5_b4_measure_approved)

            }

        s5_b5_likelihood = request.POST.get('5_5_likelihood')  # Step5 box5
        s5_b5_severity = request.POST.get('5_5_severity')  # Step5 box5
        s5_b5_option_reduce_risk = request.POST.get('5_5_option_reduce_risk')
        s5_b5_effect_on_risk = request.POST.get('5_5_effect_on_risk')
        s5_b5_residual_risk = request.POST.get('5_5_residual_risk')
        s5_b5_measure_approved = request.POST.get('5_5_measure_approved')

        if s5_b5_likelihood and s5_b5_severity is not None:
            liklihood_dict['s5_b5'] = {
                's5_b5_likelihood': replace_values_likelihood(s5_b5_likelihood),
                's5_b5_severity': replace_values_severity(s5_b5_severity),
                's5_b5_option_reduce_risk': s5_b5_option_reduce_risk,
                's5_b5_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b5_likelihood),
                                                         replace_values_severity(s5_b5_severity)),
                's5_b5_effect_on_risk': effect_on_risk_logic(s5_b5_effect_on_risk),
                's5_b5_residual_risk': residual_risk_logic(s5_b5_residual_risk),
                's5_b5_measure_approved': measure_approved_logic(s5_b5_measure_approved)

            }

        pdf_print_dict_section5 = {
            's5_b1': {
                's5_b1_likelihood': replace_values_likelihood(s5_b1_likelihood),
                's5_b1_severity': replace_values_severity(s5_b1_severity),
                's5_b1_residual_risk': residual_risk_logic(s5_b1_residual_risk),
                's5_b1_measure_approved': measure_approved_logic(s5_b1_measure_approved),
                's5_b1_option_reduce_risk': s5_b1_option_reduce_risk,
                's5_b1_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b1_likelihood),
                                                         replace_values_severity(s5_b1_severity)),
                's5_b1_effect_on_risk': effect_on_risk_logic(s5_b1_effect_on_risk),
                'f5_1_question': 'Have you assigned a person/team who will be responsible for Identifying all the risks?'
            },
            's5_b2': {
                's5_b2_likelihood': replace_values_likelihood(s5_b2_likelihood),
                's5_b2_severity': replace_values_severity(s5_b2_severity),
                's5_b2_residual_risk': residual_risk_logic(s5_b2_residual_risk),
                's5_b2_measure_approved': measure_approved_logic(s5_b2_measure_approved),
                's5_b2_option_reduce_risk': s5_b2_option_reduce_risk,
                's5_b2_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b2_likelihood),
                                                         replace_values_severity(s5_b2_severity)),
                's5_b2_effect_on_risk': effect_on_risk_logic(s5_b2_effect_on_risk),
                'f5_2_question': 'How matured is the process of documenting type and details of the risks from the proposed processing in your organization?'
            },
            's5_b3': {
                's5_b3_likelihood': replace_values_likelihood(s5_b3_likelihood),
                's5_b3_severity': replace_values_severity(s5_b3_severity),
                's5_b3_residual_risk': residual_risk_logic(s5_b3_residual_risk),
                's5_b3_measure_approved': measure_approved_logic(s5_b3_measure_approved),
                's5_b3_option_reduce_risk': s5_b3_option_reduce_risk,
                's5_b3_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b3_likelihood),
                                                         replace_values_severity(s5_b3_severity)),
                's5_b3_effect_on_risk': effect_on_risk_logic(s5_b3_effect_on_risk),
                'f5_3_question': 'How matured is the process of calculating likehlihood and severity of the risk from the proposed processing in your organization ?'
            },
            's5_b4': {
                's5_b4_likelihood': replace_values_likelihood(s5_b4_likelihood),
                's5_b4_severity': replace_values_severity(s5_b4_severity),
                's5_b4_residual_risk': residual_risk_logic(s5_b4_residual_risk),
                's5_b4_measure_approved': measure_approved_logic(s5_b4_measure_approved),
                's5_b4_option_reduce_risk': s5_b4_option_reduce_risk,
                's5_b4_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b4_likelihood),
                                                         replace_values_severity(s5_b4_severity)),
                's5_b4_effect_on_risk': effect_on_risk_logic(s5_b4_effect_on_risk),
                'f5_4_question': 'How matured is the process of finding residual risks defined in your organization?'
            },
            's5_b5': {
                's5_b5_likelihood': replace_values_likelihood(s5_b5_likelihood),
                's5_b5_severity': replace_values_severity(s5_b5_severity),
                's5_b5_residual_risk': residual_risk_logic(s5_b5_residual_risk),
                's5_b5_measure_approved': measure_approved_logic(s5_b5_measure_approved),
                's5_b5_option_reduce_risk': s5_b5_option_reduce_risk,
                's5_b5_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b5_likelihood),
                                                         replace_values_severity(s5_b5_severity)),
                's5_b5_effect_on_risk': effect_on_risk_logic(s5_b5_effect_on_risk),
                'f5_5_question': 'To what extent is the plan to quantify the impact of the risks defined in your organization?'
            },
        }
        request.session['pdf_print_dict_section5'] = pdf_print_dict_section5

        excel_question_dict = get_excel_risk_questions(request)
        text_to_excel(request, excel_question_dict, liklihood_dict)
        request.session['liklihood_dict'] = liklihood_dict

        data = {'step5_box1_likelihood': replace_values_likelihood(s5_b1_likelihood),
                'step5_box1_severity': replace_values_severity(s5_b1_severity),
                'f5_1_question': request.session.get('f5_1_question'),

                'step5_box2_likelihood': replace_values_likelihood(s5_b2_likelihood),
                'step5_box2_severity': replace_values_severity(s5_b2_severity),
                'f5_2_question': request.session.get('f5_2_question'),

                'step5_box3_likelihood': replace_values_likelihood(s5_b3_likelihood),
                'step5_box3_severity': replace_values_severity(s5_b3_severity),
                'f5_3_question': request.session.get('f5_3_question'),

                'step5_box4_likelihood': replace_values_likelihood(s5_b4_likelihood),
                'step5_box4_severity': replace_values_severity(s5_b4_severity),
                'f5_4_question': request.session.get('f5_4_question'),

                'step5_box5_likelihood': replace_values_likelihood(s5_b5_likelihood),
                'step5_box5_severity': replace_values_severity(s5_b5_severity),
                'f5_5_question': request.session.get('f5_5_question'),
                }

        request.session['final_excel_data'] = data
        context = {
            'risk_score1': risk_score1,
            'risk_score2': risk_score2,
            'risk_score3': risk_score3,
            'risk_score4': risk_score4,
            'risk_score5': risk_score5,
            'risk_score6': risk_score6,
            'risk_score7': risk_score7,
            'risk_score8': risk_score8,
            'form1_percentage': form1_percentage,
            'form2_percentage': form2_percentage,
            'form3_percentage': form3_percentage,
            'form4_percentage': form4_percentage,
            'form5_percentage': form5_percentage,
            'form6_percentage': form6_percentage,
            'form7_percentage': form7_percentage,
            'form8_percentage': form8_percentage,
            'final_pdf_dict': data,
            'total_no_of_risk': request.session.get('total_no_of_risk'),
            'name_of_controller': request.session.get('manager'),
            'name_of_dpo': request.session.get('name_of_DPO'),
            'title_of_dpo': request.session.get('title_of_DPO')
        }
        return render(request, 'risk_summary.html', context)


def risk_summary_box_6(request):
    if request.method == 'POST':
        risk_score1 = request.session.get('risk_score1')
        risk_score2 = request.session.get('risk_score2')
        risk_score3 = request.session.get('risk_score3')
        risk_score4 = request.session.get('risk_score4')
        risk_score5 = request.session.get('risk_score5')
        risk_score6 = request.session.get('risk_score6')
        risk_score7 = request.session.get('risk_score7')
        risk_score8 = request.session.get('risk_score8')
        form1_percentage = request.session.get('form1_percentage')
        form2_percentage = request.session.get('form2_percentage')
        form3_percentage = request.session.get('form3_percentage')
        form4_percentage = request.session.get('form4_percentage')
        form5_percentage = request.session.get('form5_percentage')
        form6_percentage = request.session.get('form6_percentage')
        form7_percentage = request.session.get('form7_percentage')
        form8_percentage = request.session.get('form8_percentage')

        s6_b1_likelihood = request.POST.get('6_1_likelihood')  # Step6 box1
        s6_b1_severity = request.POST.get('6_1_severity')  # Step6 box1
        s6_b1_option_reduce_risk = request.POST.get('6_1_option_reduce_risk')
        s6_b1_effect_on_risk = request.POST.get('6_1_effect_on_risk')
        s6_b1_residual_risk = request.POST.get('6_1_residual_risk')
        s6_b1_measure_approved = request.POST.get('6_1_measure_approved')

        if s6_b1_likelihood and s6_b1_severity is not None:
            liklihood_dict['s6_b1'] = {
                's6_b1_likelihood': replace_values_likelihood(s6_b1_likelihood),
                's6_b1_severity': replace_values_severity(s6_b1_severity),
                's6_b1_option_reduce_risk': s6_b1_option_reduce_risk,
                's6_b1_overall_risk': overall_risk_logic(replace_values_likelihood(s6_b1_likelihood),
                                                         replace_values_severity(s6_b1_severity)),
                's6_b1_effect_on_risk': effect_on_risk_logic(s6_b1_effect_on_risk),
                's6_b1_residual_risk': residual_risk_logic(s6_b1_residual_risk),
                's6_b1_measure_approved': measure_approved_logic(s6_b1_measure_approved)

            }

        s6_b2_likelihood = request.POST.get('6_2_likelihood')  # Step6 box2
        s6_b2_severity = request.POST.get('6_2_severity')  # Step6 box2
        s6_b2_option_reduce_risk = request.POST.get('6_2_option_reduce_risk')
        s6_b2_effect_on_risk = request.POST.get('6_2_effect_on_risk')
        s6_b2_residual_risk = request.POST.get('6_2_residual_risk')
        s6_b2_measure_approved = request.POST.get('6_2_measure_approved')

        if s6_b2_likelihood and s6_b2_severity is not None:
            liklihood_dict['s6_b2'] = {
                's6_b2_likelihood': replace_values_likelihood(s6_b2_likelihood),
                's6_b2_severity': replace_values_severity(s6_b2_severity),
                's6_b2_option_reduce_risk': s6_b2_option_reduce_risk,
                's6_b2_overall_risk': overall_risk_logic(replace_values_likelihood(s6_b2_likelihood),
                                                         replace_values_severity(s6_b2_severity)),
                's6_b2_effect_on_risk': effect_on_risk_logic(s6_b2_effect_on_risk),
                's6_b2_residual_risk': residual_risk_logic(s6_b2_residual_risk),
                's6_b2_measure_approved': measure_approved_logic(s6_b2_measure_approved)

            }

        s6_b3_likelihood = request.POST.get('6_3_likelihood')  # Step6 box3
        s6_b3_severity = request.POST.get('6_3_severity')  # Step6 box3
        s6_b3_option_reduce_risk = request.POST.get('6_3_option_reduce_risk')
        s6_b3_effect_on_risk = request.POST.get('6_3_effect_on_risk')
        s6_b3_residual_risk = request.POST.get('6_3_residual_risk')
        s6_b3_measure_approved = request.POST.get('6_3_measure_approved')

        if s6_b3_likelihood and s6_b3_severity is not None:
            liklihood_dict['s6_b3'] = {
                's6_b3_likelihood': replace_values_likelihood(s6_b3_likelihood),
                's6_b3_severity': replace_values_severity(s6_b3_severity),
                's6_b3_option_reduce_risk': s6_b3_option_reduce_risk,
                's6_b3_overall_risk': overall_risk_logic(replace_values_likelihood(s6_b3_likelihood),
                                                         replace_values_severity(s6_b3_severity)),
                's6_b3_effect_on_risk': effect_on_risk_logic(s6_b3_effect_on_risk),
                's6_b3_residual_risk': residual_risk_logic(s6_b3_residual_risk),
                's6_b3_measure_approved': measure_approved_logic(s6_b3_measure_approved)

            }

        s6_b4_likelihood = request.POST.get('6_4_likelihood')  # Step6 box4
        s6_b4_severity = request.POST.get('6_4_severity')  # Step6 box4
        s6_b4_option_reduce_risk = request.POST.get('6_4_option_reduce_risk')
        s6_b4_effect_on_risk = request.POST.get('6_4_effect_on_risk')
        s6_b4_residual_risk = request.POST.get('6_4_residual_risk')
        s6_b4_measure_approved = request.POST.get('6_4_measure_approved')

        if s6_b4_likelihood and s6_b4_severity is not None:
            liklihood_dict['s6_b4'] = {
                's6_b4_likelihood': replace_values_likelihood(s6_b4_likelihood),
                's6_b4_severity': replace_values_severity(s6_b4_severity),
                's6_b4_option_reduce_risk': s6_b4_option_reduce_risk,
                's6_b4_overall_risk': overall_risk_logic(replace_values_likelihood(s6_b4_likelihood),
                                                         replace_values_severity(s6_b4_severity)),
                's6_b4_effect_on_risk': effect_on_risk_logic(s6_b4_effect_on_risk),
                's6_b4_residual_risk': residual_risk_logic(s6_b4_residual_risk),
                's6_b4_measure_approved': measure_approved_logic(s6_b4_measure_approved)

            }

        s6_b5_likelihood = request.POST.get('6_5_likelihood')  # Step6 box4
        s6_b5_severity = request.POST.get('6_5_severity')  # Step6 box4
        s6_b5_option_reduce_risk = request.POST.get('6_5_option_reduce_risk')
        s6_b5_effect_on_risk = request.POST.get('6_5_effect_on_risk')
        s6_b5_residual_risk = request.POST.get('6_5_residual_risk')
        s6_b5_measure_approved = request.POST.get('6_5_measure_approved')

        if s6_b5_likelihood and s6_b5_severity is not None:
            liklihood_dict['s6_b5'] = {
                's6_b5_likelihood': replace_values_likelihood(s6_b5_likelihood),
                's6_b5_severity': replace_values_severity(s6_b5_severity),
                's6_b5_option_reduce_risk': s6_b5_option_reduce_risk,
                's6_b5_overall_risk': overall_risk_logic(replace_values_likelihood(s6_b5_likelihood),
                                                         replace_values_severity(s6_b5_severity)),
                's6_b5_effect_on_risk': effect_on_risk_logic(s6_b5_effect_on_risk),
                's6_b5_residual_risk': residual_risk_logic(s6_b5_residual_risk),
                's6_b5_measure_approved': measure_approved_logic(s6_b5_measure_approved)

            }

        excel_question_dict = get_excel_risk_questions(request)
        text_to_excel(request, excel_question_dict, liklihood_dict)
        request.session['liklihood_dict'] = liklihood_dict
        context = {
            'risk_score1': risk_score1,
            'risk_score2': risk_score2,
            'risk_score3': risk_score3,
            'risk_score4': risk_score4,
            'risk_score5': risk_score5,
            'risk_score6': risk_score6,
            'risk_score7': risk_score7,
            'risk_score8': risk_score8,
            'form1_percentage': form1_percentage,
            'form2_percentage': form2_percentage,
            'form3_percentage': form3_percentage,
            'form4_percentage': form4_percentage,
            'form5_percentage': form5_percentage,
            'form6_percentage': form6_percentage,
            'form7_percentage': form7_percentage,
            'form8_percentage': form8_percentage,
            'total_no_of_risk': request.session.get('total_no_of_risk'),
            'name_of_controller': request.session.get('manager'),
            'name_of_dpo': request.session.get('name_of_DPO'),
            'title_of_dpo': request.session.get('title_of_DPO')
        }
        return render(request, 'risk_summary.html', context)


def risk_summary_box_7(request):
    if request.method == 'POST':
        risk_score1 = request.session.get('risk_score1')
        risk_score2 = request.session.get('risk_score2')
        risk_score3 = request.session.get('risk_score3')
        risk_score4 = request.session.get('risk_score4')
        risk_score5 = request.session.get('risk_score5')
        risk_score6 = request.session.get('risk_score6')
        risk_score7 = request.session.get('risk_score7')
        risk_score8 = request.session.get('risk_score8')
        form1_percentage = request.session.get('form1_percentage')
        form2_percentage = request.session.get('form2_percentage')
        form3_percentage = request.session.get('form3_percentage')
        form4_percentage = request.session.get('form4_percentage')
        form5_percentage = request.session.get('form5_percentage')
        form6_percentage = request.session.get('form6_percentage')
        form7_percentage = request.session.get('form7_percentage')
        form8_percentage = request.session.get('form8_percentage')

        s7_b1_likelihood = request.POST.get('7_1_likelihood')  # Step7 box1
        s7_b1_severity = request.POST.get('7_1_severity')  # Step7 box1
        s7_b1_option_reduce_risk = request.POST.get('7_1_option_reduce_risk')
        s7_b1_effect_on_risk = request.POST.get('7_1_effect_on_risk')
        s7_b1_residual_risk = request.POST.get('7_1_residual_risk')
        s7_b1_measure_approved = request.POST.get('7_1_measure_approved')

        if s7_b1_likelihood and s7_b1_severity is not None:
            liklihood_dict['s7_b1'] = {
                's7_b1_likelihood': replace_values_likelihood(s7_b1_likelihood),
                's7_b1_severity': replace_values_severity(s7_b1_severity),
                's7_b1_option_reduce_risk': s7_b1_option_reduce_risk,
                's7_b1_overall_risk': overall_risk_logic(replace_values_likelihood(s7_b1_likelihood),
                                                         replace_values_severity(s7_b1_severity)),
                's7_b1_effect_on_risk': effect_on_risk_logic(s7_b1_effect_on_risk),
                's7_b1_residual_risk': residual_risk_logic(s7_b1_residual_risk),
                's7_b1_measure_approved': measure_approved_logic(s7_b1_measure_approved)

            }

        s7_b2_likelihood = request.POST.get('7_2_likelihood')  # Step7 box2
        s7_b2_severity = request.POST.get('7_2_severity')  # Step7 box2
        s7_b2_option_reduce_risk = request.POST.get('7_2_option_reduce_risk')
        s7_b2_effect_on_risk = request.POST.get('7_2_effect_on_risk')
        s7_b2_residual_risk = request.POST.get('7_2_residual_risk')
        s7_b2_measure_approved = request.POST.get('7_2_measure_approved')

        if s7_b2_likelihood and s7_b2_severity is not None:
            liklihood_dict['s7_b2'] = {
                's7_b2_likelihood': replace_values_likelihood(s7_b2_likelihood),
                's7_b2_severity': replace_values_severity(s7_b2_severity),
                's7_b2_option_reduce_risk': s7_b2_option_reduce_risk,
                's7_b2_overall_risk': overall_risk_logic(replace_values_likelihood(s7_b2_likelihood),
                                                         replace_values_severity(s7_b2_severity)),
                's7_b2_effect_on_risk': effect_on_risk_logic(s7_b2_effect_on_risk),
                's7_b2_residual_risk': residual_risk_logic(s7_b2_residual_risk),
                's7_b2_measure_approved': measure_approved_logic(s7_b2_measure_approved)

            }
        excel_question_dict = get_excel_risk_questions(request)
        text_to_excel(request, excel_question_dict, liklihood_dict)
        request.session['liklihood_dict'] = liklihood_dict
        context = {
            'risk_score1': risk_score1,
            'risk_score2': risk_score2,
            'risk_score3': risk_score3,
            'risk_score4': risk_score4,
            'risk_score5': risk_score5,
            'risk_score6': risk_score6,
            'risk_score7': risk_score7,
            'risk_score8': risk_score8,
            'form1_percentage': form1_percentage,
            'form2_percentage': form2_percentage,
            'form3_percentage': form3_percentage,
            'form4_percentage': form4_percentage,
            'form5_percentage': form5_percentage,
            'form6_percentage': form6_percentage,
            'form7_percentage': form7_percentage,
            'form8_percentage': form8_percentage,
            'total_no_of_risk': request.session.get('total_no_of_risk'),
            'name_of_controller': request.session.get('manager'),
            'name_of_dpo': request.session.get('name_of_DPO'),
            'title_of_dpo': request.session.get('title_of_DPO')
        }
        return render(request, 'risk_summary.html', context)


def risk_summary_box_8(request):
    if request.method == 'POST':
        risk_score1 = request.session.get('risk_score1')
        risk_score2 = request.session.get('risk_score2')
        risk_score3 = request.session.get('risk_score3')
        risk_score4 = request.session.get('risk_score4')
        risk_score5 = request.session.get('risk_score5')
        risk_score6 = request.session.get('risk_score6')
        risk_score7 = request.session.get('risk_score7')
        risk_score8 = request.session.get('risk_score8')
        form1_percentage = request.session.get('form1_percentage')
        form2_percentage = request.session.get('form2_percentage')
        form3_percentage = request.session.get('form3_percentage')
        form4_percentage = request.session.get('form4_percentage')
        form5_percentage = request.session.get('form5_percentage')
        form6_percentage = request.session.get('form6_percentage')
        form7_percentage = request.session.get('form7_percentage')
        form8_percentage = request.session.get('form8_percentage')

        s8_b1_likelihood = request.POST.get('8_1_likelihood')  # Step8 box1
        s8_b1_severity = request.POST.get('8_1_severity')  # Step8 box1
        s8_b1_option_reduce_risk = request.POST.get('8_1_option_reduce_risk')
        s8_b1_effect_on_risk = request.POST.get('8_1_effect_on_risk')
        s8_b1_residual_risk = request.POST.get('8_1_residual_risk')
        s8_b1_measure_approved = request.POST.get('8_1_measure_approved')

        if s8_b1_likelihood and s8_b1_severity is not None:
            liklihood_dict['s8_b1'] = {
                's8_b1_likelihood': replace_values_likelihood(s8_b1_likelihood),
                's8_b1_severity': replace_values_severity(s8_b1_severity),
                's8_b1_option_reduce_risk': s8_b1_option_reduce_risk,
                's8_b1_overall_risk': overall_risk_logic(replace_values_likelihood(s8_b1_likelihood),
                                                         replace_values_severity(s8_b1_severity)),
                's8_b1_effect_on_risk': effect_on_risk_logic(s8_b1_effect_on_risk),
                's8_b1_residual_risk': residual_risk_logic(s8_b1_residual_risk),
                's8_b1_measure_approved': measure_approved_logic(s8_b1_measure_approved)

            }

        s8_b2_likelihood = request.POST.get('8_2_likelihood')  # Step8 box2
        s8_b2_severity = request.POST.get('8_2_severity')  # Step8 box2
        s8_b2_option_reduce_risk = request.POST.get('8_2_option_reduce_risk')
        s8_b2_effect_on_risk = request.POST.get('8_2_effect_on_risk')
        s8_b2_residual_risk = request.POST.get('8_2_residual_risk')
        s8_b2_measure_approved = request.POST.get('8_2_measure_approved')

        if s8_b2_likelihood and s8_b2_severity is not None:
            liklihood_dict['s8_b2'] = {
                's8_b2_likelihood': replace_values_likelihood(s8_b2_likelihood),
                's8_b2_severity': replace_values_severity(s8_b2_severity),
                's8_b2_option_reduce_risk': s8_b2_option_reduce_risk,
                's8_b2_overall_risk': overall_risk_logic(replace_values_likelihood(s8_b2_likelihood),
                                                         replace_values_severity(s8_b2_severity)),
                's8_b2_effect_on_risk': effect_on_risk_logic(s8_b2_effect_on_risk),
                's8_b2_residual_risk': residual_risk_logic(s8_b2_residual_risk),
                's8_b2_measure_approved': measure_approved_logic(s8_b2_measure_approved)

            }

        s8_b3_likelihood = request.POST.get('8_3_likelihood')  # Step8 box2
        s8_b3_severity = request.POST.get('8_3_severity')  # Step8 box2
        s8_b3_option_reduce_risk = request.POST.get('8_3_option_reduce_risk')
        s8_b3_effect_on_risk = request.POST.get('8_3_effect_on_risk')
        s8_b3_residual_risk = request.POST.get('8_3_residual_risk')
        s8_b3_measure_approved = request.POST.get('8_3_measure_approved')

        if s8_b3_likelihood and s8_b3_severity is not None:
            liklihood_dict['s8_b3'] = {
                's8_b3_likelihood': replace_values_likelihood(s8_b3_likelihood),
                's8_b3_severity': replace_values_severity(s8_b3_severity),
                's8_b3_option_reduce_risk': s8_b3_option_reduce_risk,
                's8_b3_overall_risk': overall_risk_logic(replace_values_likelihood(s8_b3_likelihood),
                                                         replace_values_severity(s8_b3_severity)),
                's8_b3_effect_on_risk': effect_on_risk_logic(s8_b3_effect_on_risk),
                's8_b3_residual_risk': residual_risk_logic(s8_b3_residual_risk),
                's8_b3_measure_approved': measure_approved_logic(s8_b3_measure_approved)

            }

        s8_b4_likelihood = request.POST.get('8_4_likelihood')  # Step8 box2
        s8_b4_severity = request.POST.get('8_4_severity')  # Step8 box2
        s8_b4_option_reduce_risk = request.POST.get('8_4_option_reduce_risk')
        s8_b4_effect_on_risk = request.POST.get('8_4_effect_on_risk')
        s8_b4_residual_risk = request.POST.get('8_4_residual_risk')
        s8_b4_measure_approved = request.POST.get('8_4_measure_approved')

        if s8_b4_likelihood and s8_b4_severity is not None:
            liklihood_dict['s8_b4'] = {
                's8_b4_likelihood': replace_values_likelihood(s8_b4_likelihood),
                's8_b4_severity': replace_values_severity(s8_b4_severity),
                's8_b4_option_reduce_risk': s8_b4_option_reduce_risk,
                's8_b4_overall_risk': overall_risk_logic(replace_values_likelihood(s8_b4_likelihood),
                                                         replace_values_severity(s8_b4_severity)),
                's8_b4_effect_on_risk': effect_on_risk_logic(s8_b4_effect_on_risk),
                's8_b4_residual_risk': residual_risk_logic(s8_b4_residual_risk),
                's8_b4_measure_approved': measure_approved_logic(s8_b4_measure_approved)

            }

        excel_question_dict = get_excel_risk_questions(request)
        text_to_excel(request, excel_question_dict, liklihood_dict)
        request.session['liklihood_dict'] = liklihood_dict
        context = {
            'risk_score1': risk_score1,
            'risk_score2': risk_score2,
            'risk_score3': risk_score3,
            'risk_score4': risk_score4,
            'risk_score5': risk_score5,
            'risk_score6': risk_score6,
            'risk_score7': risk_score7,
            'risk_score8': risk_score8,
            'form1_percentage': form1_percentage,
            'form2_percentage': form2_percentage,
            'form3_percentage': form3_percentage,
            'form4_percentage': form4_percentage,
            'form5_percentage': form5_percentage,
            'form6_percentage': form6_percentage,
            'form7_percentage': form7_percentage,
            'form8_percentage': form8_percentage,
            'total_no_of_risk': request.session.get('total_no_of_risk'),
            'name_of_controller': request.session.get('manager'),
            'name_of_dpo': request.session.get('name_of_DPO'),
            'title_of_dpo': request.session.get('title_of_DPO')
        }
        return render(request, 'risk_summary.html', context)


@login_required
def risk_summary_details(request):
    if request.method == 'POST':
        # Section 1
        s1_b1_likelihood = request.POST.get('1_1_likelihood') # Step1_Box1
        s1_b1_severity = request.POST.get('1_1_severity')
        request.session['s1_b1_likelihood'] = s1_b1_likelihood
        request.session['s1_b1_severity'] = s1_b1_severity

        s1_b2_likelihood = request.POST.get('1_2_likelihood')  # Step1_Box2
        s1_b2_severity = request.POST.get('1_2_severity')
        request.session['s1_b2_likelihood'] = s1_b2_likelihood
        request.session['s1_b2_severity'] = s1_b2_severity

        s1_b3_likelihood = request.POST.get('1_3_likelihood')  # Step1_Box3
        s1_b3_severity = request.POST.get('1_3_severity')
        request.session['s1_b3_likelihood'] = s1_b3_likelihood
        request.session['s1_b3_severity'] = s1_b3_severity

        s1_b4_likelihood = request.POST.get('1_4_likelihood')  # Step1_Box4
        s1_b4_severity = request.POST.get('1_4_severity')
        request.session['s1_b4_likelihood'] = s1_b4_likelihood
        request.session['s1_b4_severity'] = s1_b4_severity

        s1_b5_likelihood = request.POST.get('1_5_likelihood')  # Step1_Box5
        s1_b5_severity = request.POST.get('1_5_severity')
        request.session['s1_b5_likelihood'] = s1_b5_likelihood
        request.session['s1_b5_severity'] = s1_b5_severity

        s1_b6_likelihood = request.POST.get('1_6_likelihood')  # Step1_Box6
        s1_b6_severity = request.POST.get('1_6_severity')
        request.session['s1_b6_likelihood'] = s1_b6_likelihood
        request.session['s1_b6_severity'] = s1_b6_severity


        # db_objects = RiskSummaryForm.objects.all()
        # form1_data = RiskSummaryForm(s1_b1_likelihood=s1_b1_likelihood,
        #                              s1_b2_likelihood=s1_b2_likelihood,
        #                              s1_b3_likelihood=s1_b3_likelihood,
        #                              s1_b4_likelihood=s1_b4_likelihood,
        #                              s1_b5_likelihood=s1_b5_likelihood,
        #                              s1_b6_likelihood=s1_b6_likelihood,
        #
        #                              s1_b1_severity=s1_b1_likelihood,
        #                              s1_b2_severity=s1_b2_likelihood,
        #                              s1_b3_severity=s1_b3_likelihood,
        #                              s1_b4_severity=s1_b4_likelihood,
        #                              s1_b5_severity=s1_b5_likelihood,
        #                              s1_b6_severity=s1_b6_likelihood,
        #                              )
        #
        # form1_data.save()

        # Section 2

        s2_b1_likelihood = request.POST.get('2_1_likelihood')  # Step1_Box1
        s2_b1_severity = request.POST.get('2_1_severity')
        request.session['s2_b1_likelihood'] = s2_b1_likelihood
        request.session['s2_b1_severity'] = s2_b1_severity

        s2_b2_likelihood = request.POST.get('2_2_likelihood')  # Step1_Box2
        s2_b2_severity = request.POST.get('2_2_severity')
        request.session['s2_b2_likelihood'] = s2_b2_likelihood
        request.session['s2_b2_severity'] = s2_b2_severity

        s2_b3_likelihood = request.POST.get('2_3_likelihood')  # Step1_Box3
        s2_b3_severity = request.POST.get('2_3_severity')
        request.session['s2_b3_likelihood'] = s2_b3_likelihood
        request.session['s2_b3_severity'] = s2_b3_severity

        s2_b4_likelihood = request.POST.get('2_4_likelihood')  # Step1_Box4
        s2_b4_severity = request.POST.get('2_4_severity')
        request.session['s2_b4_likelihood'] = s2_b4_likelihood
        request.session['s2_b4_severity'] = s2_b4_severity

        s2_b5_likelihood = request.POST.get('2_5_likelihood')  # Step1_Box5
        s2_b5_severity = request.POST.get('2_5_severity')
        request.session['s2_b5_likelihood'] = s2_b5_likelihood
        request.session['s2_b5_severity'] = s2_b5_severity

        s2_b6_likelihood = request.POST.get('2_6_likelihood')  # Step1_Box6
        s2_b6_severity = request.POST.get('2_6_severity')
        request.session['s2_b6_likelihood'] = s2_b6_likelihood
        request.session['s2_b6_severity'] = s2_b6_severity

        s2_b7_likelihood = request.POST.get('2_7_likelihood')  # Step1_Box6
        s2_b7_severity = request.POST.get('2_7_severity')
        request.session['s2_b7_likelihood'] = s2_b7_likelihood
        request.session['s2_b7_severity'] = s2_b7_severity

        s2_b8_likelihood = request.POST.get('2_8_likelihood')  # Step1_Box6
        s2_b8_severity = request.POST.get('2_8_severity')
        request.session['s2_b8_likelihood'] = s2_b8_likelihood
        request.session['s2_b8_severity'] = s2_b8_severity

        s2_b9_likelihood = request.POST.get('2_9_likelihood')  # Step1_Box6
        s2_b9_severity = request.POST.get('2_9_severity')
        request.session['s2_b9_likelihood'] = s2_b9_likelihood
        request.session['s2_b9_severity'] = s2_b9_severity



        # Step 2 start
        s2_b1_likelihood = request.POST.get('2_1_likelihood') # Step2 box1
        s2_b1_severity = request.POST.get('2_1_severity') # Step2 box1
        request.session['s2_b1_likelihood'] = s2_b1_likelihood
        request.session['s2_b1_severity'] = s2_b1_severity


        # Test ###################



        ##########################
        # Step 5 start
        s5_b1_likelihood = request.POST.get('5_1_likelihood')  # Step5 box1
        s5_b1_severity = request.POST.get('5_1_severity')  # Step5 box1
        s5_b1_option_reduce_risk = request.POST.get('5_1_option_reduce_risk')
        s5_b1_effect_on_risk = request.POST.get('5_1_effect_on_risk')
        s5_b1_residual_risk = request.POST.get('5_1_residual_risk')
        s5_b1_measure_approved = request.POST.get('5_1_measure_approved')
        s5_b1_comments_box = request.POST.get('5_1_comments_box')

        s5_b1_effect_on_risk_logic = effect_on_risk_logic(s5_b1_effect_on_risk)
        s5_b1_residual_risk_logic = residual_risk_logic(s5_b1_residual_risk)
        s5_b1_measure_approved_logic = measure_approved_logic(s5_b1_measure_approved)

        request.session['s5_b1_likelihood'] = s5_b1_likelihood
        request.session['s5_b1_severity'] = s5_b1_severity
        request.session['s5_b1_option_reduce_risk'] = s5_b1_option_reduce_risk
        request.session['s5_b1_residual_risk'] = s5_b1_residual_risk_logic
        request.session['s5_b1_measure_approved'] = s5_b1_measure_approved_logic
        request.session['s5_b1_comments_box'] = s5_b1_comments_box
        request.session['s5_b1_effect_on_risk'] = s5_b1_effect_on_risk_logic


        s5_b2_likelihood = request.POST.get('5_2_likelihood')  # Step5 box2
        s5_b2_severity = request.POST.get('5_2_severity')  # Step5 box2
        s5_b2_option_reduce_risk = request.POST.get('5_2_option_reduce_risk')
        s5_b2_effect_on_risk = request.POST.get('5_2_effect_on_risk')
        s5_b2_residual_risk = request.POST.get('5_2_residual_risk')
        s5_b2_measure_approved = request.POST.get('5_2_measure_approved')
        s5_b2_comments_box = request.POST.get('5_2_comments_box')

        s5_b2_effect_on_risk_logic = effect_on_risk_logic(s5_b2_effect_on_risk)
        s5_b2_residual_risk_logic = residual_risk_logic(s5_b2_residual_risk)
        s5_b2_measure_approved_logic = measure_approved_logic(s5_b2_measure_approved)

        request.session['s5_b2_likelihood'] = s5_b2_likelihood
        request.session['s5_b2_severity'] = s5_b2_severity
        request.session['s5_b2_option_reduce_risk'] = s5_b2_option_reduce_risk
        request.session['s5_b2_residual_risk'] = s5_b2_residual_risk_logic
        request.session['s5_b2_measure_approved'] = s5_b2_measure_approved_logic
        request.session['s5_b2_comments_box'] = s5_b2_comments_box
        request.session['s5_b2_effect_on_risk'] = s5_b2_effect_on_risk_logic


        s5_b3_likelihood = request.POST.get('5_3_likelihood')  # Step5 box3
        s5_b3_severity = request.POST.get('5_3_severity')  # Step5 box3
        s5_b3_option_reduce_risk = request.POST.get('5_3_option_reduce_risk')
        s5_b3_effect_on_risk = request.POST.get('5_3_effect_on_risk')
        s5_b3_residual_risk = request.POST.get('5_3_residual_risk')
        s5_b3_measure_approved = request.POST.get('5_3_measure_approved')
        s5_b3_comments_box = request.POST.get('5_3_comments_box')

        s5_b3_effect_on_risk_logic = effect_on_risk_logic(s5_b3_effect_on_risk)
        s5_b3_residual_risk_logic = residual_risk_logic(s5_b3_residual_risk)
        s5_b3_measure_approved_logic = measure_approved_logic(s5_b3_measure_approved)

        request.session['s5_b3_likelihood'] = s5_b3_likelihood
        request.session['s5_b3_severity'] = s5_b3_severity
        request.session['s5_b3_option_reduce_risk'] = s5_b3_option_reduce_risk
        request.session['s5_b3_residual_risk'] = s5_b3_residual_risk_logic
        request.session['s5_b3_measure_approved'] = s5_b3_measure_approved_logic
        request.session['s5_b3_comments_box'] = s5_b3_comments_box
        request.session['s5_b3_effect_on_risk'] = s5_b3_effect_on_risk_logic

        s5_b4_likelihood = request.POST.get('5_4_likelihood')  # Step5 box4
        s5_b4_severity = request.POST.get('5_4_severity')  # Step5 box4
        s5_b4_option_reduce_risk = request.POST.get('5_4_option_reduce_risk')
        s5_b4_effect_on_risk = request.POST.get('5_4_effect_on_risk')
        s5_b4_residual_risk = request.POST.get('5_4_residual_risk')
        s5_b4_measure_approved = request.POST.get('5_4_measure_approved')
        s5_b4_comments_box = request.POST.get('5_4_comments_box')

        s5_b4_effect_on_risk_logic = effect_on_risk_logic(s5_b4_effect_on_risk)
        s5_b4_residual_risk_logic = residual_risk_logic(s5_b4_residual_risk)
        s5_b4_measure_approved_logic = measure_approved_logic(s5_b4_measure_approved)

        request.session['s5_b4_likelihood'] = s5_b4_likelihood
        request.session['s5_b4_severity'] = s5_b4_severity
        request.session['s5_b4_option_reduce_risk'] = s5_b4_option_reduce_risk
        request.session['s5_b4_residual_risk'] = s5_b4_residual_risk_logic
        request.session['s5_b4_measure_approved'] = s5_b4_measure_approved_logic
        request.session['s5_b4_comments_box'] = s5_b4_comments_box
        request.session['s5_b4_effect_on_risk'] = s5_b4_effect_on_risk_logic

        s5_b5_likelihood = request.POST.get('5_5_likelihood')  # Step5 box5
        s5_b5_severity = request.POST.get('5_5_severity')  # Step5 box5
        s5_b5_option_reduce_risk = request.POST.get('5_5_option_reduce_risk')
        s5_b5_effect_on_risk = request.POST.get('5_5_effect_on_risk')
        s5_b5_residual_risk = request.POST.get('5_5_residual_risk')
        s5_b5_measure_approved = request.POST.get('5_5_measure_approved')
        s5_b5_comments_box = request.POST.get('5_5_comments_box')

        s5_b5_effect_on_risk_logic = effect_on_risk_logic(s5_b5_effect_on_risk)
        s5_b5_residual_risk_logic = residual_risk_logic(s5_b5_residual_risk)
        s5_b5_measure_approved_logic = measure_approved_logic(s5_b5_measure_approved)

        request.session['s5_b5_likelihood'] = s5_b5_likelihood
        request.session['s5_b5_severity'] = s5_b5_severity
        request.session['s5_b5_option_reduce_risk'] = s5_b5_option_reduce_risk
        request.session['s5_b5_residual_risk'] = s5_b5_residual_risk_logic
        request.session['s5_b5_measure_approved'] = s5_b5_measure_approved_logic
        request.session['s5_b5_comments_box'] = s5_b5_comments_box
        request.session['s5_b5_effect_on_risk'] = s5_b5_effect_on_risk_logic

        pdf_print_dict_section5 = {
            's5_b1': {
                's5_b1_likelihood': replace_values_likelihood(s5_b1_likelihood),
                's5_b1_severity': replace_values_severity(s5_b1_severity),
                's5_b1_residual_risk': s5_b1_residual_risk_logic,
                's5_b1_measure_approved': s5_b1_measure_approved_logic,
                's5_b1_option_reduce_risk': s5_b1_option_reduce_risk,
                's5_b1_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b1_likelihood), replace_values_severity(s5_b1_severity)),
                's5_b1_effect_on_risk': s5_b1_effect_on_risk_logic,
                'f5_1_question': 'Have you assigned a person/team who will be responsible for Identifying all the risks?'
            },
            's5_b2': {
                's5_b2_likelihood': replace_values_likelihood(s5_b2_likelihood),
                's5_b2_severity': replace_values_severity(s5_b2_severity),
                's5_b2_residual_risk': s5_b2_residual_risk_logic,
                's5_b2_measure_approved': s5_b2_measure_approved_logic,
                's5_b2_option_reduce_risk': s5_b2_option_reduce_risk,
                's5_b2_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b2_likelihood), replace_values_severity(s5_b2_severity)),
                's5_b2_effect_on_risk': s5_b2_effect_on_risk_logic,
                'f5_2_question': 'How matured is the process of documenting type and details of the risks from the proposed processing in your organization?'

            },
            's5_b3': {
                's5_b3_likelihood': replace_values_likelihood(s5_b3_likelihood),
                's5_b3_severity': replace_values_severity(s5_b3_severity),
                's5_b3_residual_risk': s5_b3_residual_risk_logic,
                's5_b3_measure_approved': s5_b3_measure_approved_logic,
                's5_b3_option_reduce_risk': s5_b3_option_reduce_risk,
                's5_b3_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b3_likelihood), replace_values_severity(s5_b3_severity)),
                's5_b3_effect_on_risk': s5_b3_effect_on_risk_logic,
                'f5_3_question': 'How matured is the process of calculating likehlihood and severity of the risk from the proposed processing in your organization?'

            },
            's5_b4': {
                's5_b4_likelihood': replace_values_likelihood(s5_b4_likelihood),
                's5_b4_severity': replace_values_severity(s5_b4_severity),
                's5_b4_residual_risk': s5_b4_residual_risk_logic,
                's5_b4_measure_approved': s5_b4_measure_approved_logic,
                's5_b4_option_reduce_risk': s5_b4_option_reduce_risk,
                's5_b4_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b4_likelihood), replace_values_severity(s5_b4_severity)),
                's5_b4_effect_on_risk': s5_b4_effect_on_risk_logic,
                'f5_4_question': 'How matured is the process of finding residual risks defined in your organization?'

            },
            's5_b5': {
                's5_b5_likelihood': replace_values_likelihood(s5_b5_likelihood),
                's5_b5_severity': replace_values_severity(s5_b5_severity),
                's5_b5_residual_risk': s5_b5_residual_risk_logic,
                's5_b5_measure_approved': s5_b5_measure_approved_logic,
                's5_b5_effect_on_risk': s5_b5_effect_on_risk_logic,
                's5_b5_overall_risk': overall_risk_logic(replace_values_likelihood(s5_b5_likelihood), replace_values_severity(s5_b5_severity)),
                's5_b5_option_reduce_risk': s5_b5_option_reduce_risk,
                'f5_5_question': 'To what extent is the plan to quantify the impact of the risks defined in your organization?'

            }
        }
        request.session['pdf_print_dict_section5'] = pdf_print_dict_section5

        # Creating final excel and pdf dictionary
        # Section 1
        value_excel_dict = {}
        f1_1 = request.session.get('f1_1')
        if f1_1 == '1' or f1_1 == '2':
            f1_1_question = 'To what extent is the purpose of the processing defined in your organization?; ' + replace_value_1_2(f1_1)
            request.session['f1_1_question'] = f1_1_question
            value_excel_dict['f1_1'] = f1_1_question

        f1_2 = request.session.get('f1_2')
        if f1_2 == '1' or f1_2 == '2':
            f1_2_question = 'To what extent is the aim of the project defined in your organization?; ' + replace_value_1_2(f1_2)
            request.session['f1_2_question'] = f1_2_question
            value_excel_dict['f1_2'] = f1_2_question

        f1_3 = request.session.get('f1_3')
        if f1_3 == '1' or f1_3 == '2':
            f1_3_question = 'To what extent is method of processing defined in your organization?; ' + replace_value_1_2(f1_3)
            request.session['f1_3_question'] = f1_3_question
            value_excel_dict['f1_3'] = f1_3_question

        f1_4 = request.session.get('f1_4')
        if f1_4 == '1' or f1_4 == '2':
            f1_4_question = 'To what extent are the expected benefit(s) from the processing defined?; ' + replace_value_1_2(f1_4)
            request.session['f1_4_question'] = f1_4_question
            value_excel_dict['f1_4'] = f1_4_question

        f1_5 = request.session.get('f1_5')
        if f1_5 == '1' or f1_5 == '2':
            f1_5_question = 'To what extent is the list of departments who will benefit from this processing defined in your organization?; ' + replace_value_1_2(f1_5)
            request.session['f1_5_question'] = f1_5_question
            value_excel_dict['f1_5'] = f1_5_question

        f1_6 = request.session.get('f1_6')
        if f1_6 == '1' or f1_6 == '2':
            f1_6_question = 'To what extent is the list of members / teams who will be involved in the processing defined in your organization?; ' + replace_value_1_2(f1_6)
            request.session['f1_6_question'] = f1_6_question
            value_excel_dict['f1_6'] = f1_6_question

        number_of_risk = request.session.get('risk_score1')[1]

        request.session['f5_1_question'] = 'Have you assigned a person/team who will be responsible for Identifying all the risks?'

        # Section 2
        f2_1 = request.session.get('f2_1')
        if f2_1 == '1' or f2_1 == '2':
            f2_1_question = 'How matured is the process of collecting, using, storing and deleting data in your organization ?; '+ replace_value_1_2(f2_1)
            value_excel_dict['f2_1'] = f2_1_question

        f2_2 = request.session.get('f2_2')
        if f2_2 == '1' or f2_2 == '2':
            f2_2_question = 'To what extent are the data source(s) used to initiate processing defined in your organization ?; '+ replace_value_1_2(f2_2)
            value_excel_dict['f2_2'] = f2_2_question

        f2_3 = request.session.get('f2_3')
        if f2_3 == '1' or f2_3 == '2':
            f2_3_question = 'Do you have a list of people who will have access to this data?; '+ replace_value_1_2(f2_3)
            value_excel_dict['f2_3'] = f2_3_question

        f2_4 = request.session.get('f2_4')
        if f2_4 == '1' or f2_4 == '2':
            f2_4_question = 'To what extent is the types of High-risk processing defined for this data processing?; '+ replace_value_1_2(f2_4)
            value_excel_dict['f2_4'] = f2_4_question

        f2_5 = request.session.get('f2_5')
        if f2_5 == '1' or f2_5 == '2':
            f2_5_question = 'Do you comply by legal requirements to collect the data (Do you have consent of the data subjects) ?; '+ replace_value_1_2(f2_5)
            value_excel_dict['f2_5'] = f2_5_question

        f2_6 = request.session.get('f2_6')
        if f2_6 == '1' or f2_6 == '2':
            f2_6_question = 'How many data subjects are likely to be affected by the project?; '+ replace_value_1_2(f2_6)
            value_excel_dict['f2_6'] = f2_6_question

        f2_7 = request.session.get('f2_7')
        if f2_7 == '1' or f2_7 == '2':
            f2_7_question = 'Where is data stored ?; '+ replace_value_1_2(f2_7)
            value_excel_dict['f2_7'] = f2_7_question

        f2_8 = request.session.get('f2_8')
        if f2_8 == '1' or f2_8 == '2':
            f2_8_question = 'Do you have appropriate measures to destroy data after use?; '+ replace_value_1_2(f2_8)
            value_excel_dict['f2_8'] = f2_8_question

        f2_9 = request.session.get('f2_9')
        if f2_9 == '1' or f2_9 == '2':
            f2_9_question = 'To what extent is the Data retention policy defined in your organization?; '+ replace_value_1_2(f2_9)
            value_excel_dict['f2_9'] = f2_9_question


        # Section 3
        f3_1 = request.session.get('f3_1')
        if f3_1 == '1' or f3_1 == '2':
            f3_1_question = 'Who all will be involved in the consultation process ?'
            value_excel_dict['f3_1'] = f3_1_question

        f3_2 = request.session.get('f3_2')
        if f3_2 == '1' or f3_2 == '2':
            f3_2_question = 'Have you designed a Consultation Process . This Consultation process will involve seeking views of Data subjects or their representatives on the intended processing ?'
            value_excel_dict['f3_2'] = f3_2_question


        f3_3 = request.session.get('f3_3')
        if f3_3 == '1' or f3_3 == '1':
            f3_3_question = 'Is there a process in your organization to choose the people for consultation on the proposed data processing ?'
            value_excel_dict['f3_3'] = f3_3_question


        f3_4 = request.session.get('f3_4')
        if f3_4 == '1' or f3_4 == '2':
            f3_4_question = 'How matured is the process in your organization of evaluating data processors, information security experts or other staff as consultants ?'
            value_excel_dict['f3_4'] = f3_4_question


        f3_5 = request.session.get('f3_5')
        if f3_5 == '1' or f3_5 == '2':
            f3_5_question = 'How matured is the process in your organization of reaching out to the consultants for consultation by data subjects, project teams etc ?'
            value_excel_dict['f3_5'] = f3_5_question

        f3_6 = request.session.get('f3_6')
        if f3_6 == '1' or f3_6 == '2':
            f3_6_question = 'To what extent is your process defined for mapping roles of the selected consultant(s) ?'
            value_excel_dict['f3_6'] = f3_6_question

        f3_7 = request.session.get('f3_7')
        if f3_7 == '1' or f3_7 == '2':
            f3_7_question = 'Do you have a list of members, other than designated consultant, who will also involve from the organization in the consultation process ?'
            value_excel_dict['f3_7'] = f3_7_question

        f3_8 = request.session.get('f3_8')
        if f3_8 == '1' or f3_8 == '2':
            f3_8_question = 'Do you have a process to monitor the consultation process ?'
            value_excel_dict['f3_8'] = f3_8_question

        f3_9 = request.session.get('f3_9')
        if f3_9 == '1' or f3_9 == '2':
            f3_9_question = 'How matured is the process in your organization to deliberate on the issues raised by the consultants or data subjects ?'
            value_excel_dict['f3_9'] = f3_9_question

        f3_10 = request.session.get('f3_10')
        if f3_10 == '1' or f3_10 == '2':
            f3_10_question = 'Do you have a proper forum where these consultants can provide assurances and rasie their concerns ?'
            value_excel_dict['f3_10'] = f3_10_question

        f3_11 = request.session.get('f3_11')
        if f3_11 == '1' or f3_11 == '2':
            f3_11_question = 'Have you assigned a person who will be responsible for keeping a check that all the issues raised by these consultants have been resolved ?'
            value_excel_dict['f3_11'] = f3_11_question

        f3_12 = request.session.get('f3_12')
        if f3_12 == '1' or f3_12 == '2':
            f3_12_question = 'If Response to the above question is Yes, Please provide the name of the assigned personnel.'
            value_excel_dict['f3_12'] = f3_12_question


        # Section 4
        f4_2 = request.session.get('f4_2')
        if f4_2 == '1' or f4_2 == '2':
            f4_2_question = 'How matured is the process of defining the legality of data processing with respect to the data subjects in your organization ?'
            value_excel_dict['f4_2'] = f4_2_question

        f4_3 = request.session.get('f4_3')
        if f4_3 == '1' or f4_3 == '2':
            f4_3_question = 'Is there a process to evaluate the outcome of the processing with respect to the desired goals ?'
            value_excel_dict['f4_3'] = f4_3_question

        f4_4 = request.session.get('f4_4')
        if f4_4 == '1' or f4_4 == '2':
            f4_4_question = 'Will the data processing achieve the desired goal ?'
            value_excel_dict['f4_1'] = f4_4_question

        f4_5 = request.session.get('f4_5')
        if f4_5 == '1' or f4_5 == '2':
            f4_5_question = 'Is there an alternate approach/s to achieve the same outcome ?'
            value_excel_dict['f4_5'] = f4_5_question

        f4_6 = request.session.get('f4_6')
        if f4_6 == '1' or f4_6 == '2':
            f4_6_question = 'How matured is the process to explore the alternate approaches (less intrusive measures) in order to achieve the same results ?'
            value_excel_dict['f4_6'] = f4_6_question

        f4_7 = request.session.get('f4_7')
        if f4_7 == '1' or f4_7 == '2':
            f4_7_question = 'What is the maturity of defined data quality KPIs and metrics in your organization to monitor data quality and integrity ?'
            value_excel_dict['f4_7'] = f4_7_question

        f4_8 = request.session.get('f4_8')
        if f4_8 == '1' or f4_8 == '2':
            f4_8_question = 'To what extent is the information that will be given to the data subjects defined in your organization ?'
            value_excel_dict['f4_8'] = f4_8_question

        f4_9 = request.session.get('f4_9')
        if f4_9 == '1' or f4_9 == '2':
            f4_9_question = 'How matured is the process in your organization of upholding the Data Subject Rights while processing their data ?'
            value_excel_dict['f4_9'] = f4_9_question

        f4_10 = request.session.get('f4_10')
        if f4_10 == '1' or f4_10 == '2':
            f4_10_question = 'How matured is the process to monitor how compliant designated processing entities are while preforming the proposed processing ?'
            value_excel_dict['f4_10'] = f4_10_question

        f4_11 = request.session.get('f4_11')
        if f4_11 == '1' or f4_11 == '2':
            f4_11_question = 'How matured is your process is to ensure that data processing is not used for out of scope requirement in order to prevent function creep ?'
            value_excel_dict['f4_11'] = f4_11_question

        f4_12 = request.session.get('f4_12')
        if f4_12 == '1' or f4_12 == '2':
            f4_12_question = 'To what extent are the measures defined to ensure processors comply to the scope of the project during data processing ?'
            value_excel_dict['f4_12'] = f4_12_question

        f4_13 = request.session.get('f4_13')
        if f4_13 == '1' or f4_13 == '2':
            f4_13_question = 'To what extend the measures defined to safeguard any international transfers ?'
            value_excel_dict['f4_13'] = f4_13_question


        # Section 5
        f5_1 = request.session.get('f5_1')
        if f5_1 == '1' or f5_1 == '2':
            f5_1_question = 'Have you assigned a person/team who will be responsible for Identifying all the risks ?'
            request.session['f5_1_question'] = f5_1_question
            value_excel_dict['f5_1'] = f5_1_question

        f5_2 = request.session.get('f5_2')
        if f5_2 == '1' or f5_2 == '2':
            f5_2_question = 'How matured is the process of documenting type and details of the risks from the proposed processing in your organization?; '
            request.session['f5_2_question'] = f5_2_question
            value_excel_dict['f5_2'] = f5_2_question

        f5_3 = request.session.get('f5_3')
        if f5_3 == '1' or f5_3 == '2':
            f5_3_question = 'How matured is the process of calculating likehlihood and severity of the risk from the proposed processing in your organization?; '
            request.session['f5_3_question'] = f5_3_question
            value_excel_dict['f5_3'] = f5_3_question

        f5_4 = request.session.get('f5_4')
        if f5_4 == '1' or f5_4 == '2':
            f5_4_question = 'How matured is the process of finding residual risks < define what is residual risks > defined in your organization?; '
            request.session['f5_4_question'] = f5_4_question
            value_excel_dict['f5_4'] = f5_4_question

        f5_5 = request.session.get('f5_5')
        if f5_5 == '1' or f5_5 == '2':
            f5_5_question = 'To what extent is the plan to quantify the impact of the risks defined in your organization?; '
            request.session['f5_5_question'] = f5_5_question
            value_excel_dict['f5_5'] = f5_5_question

        # Section 6
        f6_1 = request.session.get('f6_1')
        if f6_1 == '1' or f6_1 == '2':
            f6_1_question = 'Have you assigned a person/ team who will be responsible for describing the risk mitigation measures ?'
            value_excel_dict['f6_1'] = f6_1_question

        f6_2 = request.session.get('f6_2')
        if f6_2 == '1' or f6_2 == '2':
            f6_2_question = 'Have you assigned a person/ team who will be responsible who will be responsible for defining the impact of risk ?'
            value_excel_dict['f6_2'] = f6_2_question

        f6_3 = request.session.get('f6_3')
        if f6_3 == '1' or f6_3 == '2':
            f6_3_question = 'To what extent is the plan to reduce or eliminate risk defined in your organization ?'
            value_excel_dict['f6_3'] = f6_3_question

        f6_4 = request.session.get('f6_4')
        if f6_4 == '1' or f6_4 == '2':
            f6_4_question = 'Have you assigned a person who will be responsible for planning and leading the risk mitigation action plan ?'
            value_excel_dict['f6_4'] = f6_4_question

        f6_5 = request.session.get('f6_5')
        if f6_5 == '1' or f6_1 == '2':
            f6_5_question = 'Have you assigned a person who will be responsible for auditing the plan ?'
            value_excel_dict['f6_5'] = f6_5_question

        # Section 7
        f7_1 = request.session.get('f7_1')
        if f7_1 == '1' or f7_1 == '2':
            f7_1_question = 'Have you assigned a person who will be responsible for examining the residual risks ?'
            value_excel_dict['f7_1'] = f7_1_question

        f7_2 = request.session.get('f7_2')
        if f7_2 == '1' or f7_2 == '2':
            f7_2_question = 'Comments provided by Consultation team '
            value_excel_dict['f7_2'] = f7_2_question


        # Section 8
        f8_1 = request.session.get('f8_1')
        if f8_1 == '1' or f8_1 == '2':
            f8_1_question = 'Have you assigned a person who will be responsible for collecting and analysing the DPIA outcomes ?'
            value_excel_dict['f8_1'] = f8_1_question

        f8_2 = request.session.get('f8_2')
        if f8_2 == '1' or f8_2 == '2':
            f8_2_question = 'Have you assigned a person who will be responsible for planning intergration of DPIA outcomes into the project plan ?'
            value_excel_dict['f8_2'] = f8_2_question

        f8_3 = request.session.get('f8_3')
        if f8_3 == '1' or f8_3 == '2':
            f8_3_question = 'To what extent is your policy for assigning time period to Implement the DPIA outcomes defined in your organization ?'
            value_excel_dict['f8_3'] = f8_3_question

        f8_4 = request.session.get('f8_4')
        if f8_4 == '1' or f8_4 == '2':
            f8_4_question = 'To what extent are the KPIs which will depict the success of Integration defined in your organization ?'
            value_excel_dict['f8_4'] = f8_4_question



        total_risks = request.session.get('total_no_of_risk')


        data = {'step1_box1_likelihood': replace_values_likelihood(s1_b1_likelihood),
                'step1_box1_severity': replace_values_severity(s1_b1_severity),
                'f1_1_question': request.session.get('f1_1_question'),

                'step1_box2_likelihood': replace_values_likelihood(s1_b2_likelihood),
                'step1_box2_severity': replace_values_severity(s1_b2_severity),
                'f1_2_question': request.session.get('f1_2_question'),

                'step1_box3_likelihood': replace_values_likelihood(s1_b3_likelihood),
                'step1_box3_severity': replace_values_severity(s1_b3_severity),
                'f1_3_question': request.session.get('f1_3_question'),

                'step1_box4_likelihood': replace_values_likelihood(s1_b4_likelihood),
                'step1_box4_severity': replace_values_severity(s1_b4_severity),
                'f1_4_question': request.session.get('f1_4_question'),

                'step1_box5_likelihood': replace_values_likelihood(s1_b5_likelihood),
                'step1_box5_severity': replace_values_severity(s1_b5_severity),
                'f1_5_question': request.session.get('f1_5_question'),

                'step1_box6_likelihood': replace_values_likelihood(s1_b6_likelihood),
                'step1_box6_severity': replace_values_severity(s1_b6_severity),
                'f1_6_question': request.session.get('f1_6_question'),

                # 5th Box Questions
                'step5_box1_likelihood': replace_values_likelihood(s5_b1_likelihood),
                'step5_box1_severity': replace_values_severity(s5_b1_severity),
                'f5_2_question': request.session.get('f5_2_question'),

                'step5_box2_likelihood': replace_values_likelihood(s5_b2_likelihood),
                'step5_box2_severity': replace_values_severity(s5_b2_severity),
                'f5_3_question': request.session.get('f5_3_question'),

                'step5_box3_likelihood': replace_values_likelihood(s5_b3_likelihood),
                'step5_box3_severity': replace_values_severity(s5_b3_severity),
                'f5_4_question': request.session.get('f5_4_question'),

                'step5_box4_likelihood': replace_values_likelihood(s5_b4_likelihood),
                'step5_box4_severity': replace_values_severity(s5_b4_severity),
                'f5_5_question': request.session.get('f5_5_question'),
                }

        request.session['final_excel_data'] = data
        request.session['value_excel_dict'] = value_excel_dict

        risk_score1 = request.session.get('risk_score1')
        risk_score2 = request.session.get('risk_score2')
        risk_score3 = request.session.get('risk_score3')
        risk_score4 = request.session.get('risk_score4')
        risk_score5 = request.session.get('risk_score5')
        risk_score6 = request.session.get('risk_score6')
        risk_score7 = request.session.get('risk_score7')
        risk_score8 = request.session.get('risk_score8')
        form1_percentage = request.session.get('form1_percentage')
        form2_percentage = request.session.get('form2_percentage')
        form3_percentage = request.session.get('form3_percentage')
        form4_percentage = request.session.get('form4_percentage')
        form5_percentage = request.session.get('form5_percentage')
        form6_percentage = request.session.get('form6_percentage')
        form7_percentage = request.session.get('form7_percentage')
        form8_percentage = request.session.get('form8_percentage')

        context = {
            'title': request.session.get('summary_title'),
            'score_count': request.session.get('-'),
            'box_num': request.session.get('box_num'),
            'total_no_of_risk': request.session.get('total_no_of_risk'),
            'risk_score1': risk_score1,
            'risk_score2': risk_score2,
            'risk_score3': risk_score3,
            'risk_score4': risk_score4,
            'risk_score5': risk_score5,
            'risk_score6': risk_score6,
            'risk_score7': risk_score7,
            'risk_score8': risk_score8,
            'form1_percentage': form1_percentage,
            'form2_percentage': form2_percentage,
            'form3_percentage': form3_percentage,
            'form4_percentage': form4_percentage,
            'form5_percentage': form5_percentage,
            'form6_percentage': form6_percentage,
            'form7_percentage': form7_percentage,
            'form8_percentage': form8_percentage,
        }
        return render(request, 'risk_summary.html', context)

    else:
        f1_1 = request.session.get('f1_1')
        f1_2 = request.session.get('f1_2')
        f1_3 = request.session.get('f1_3')
        f1_4 = request.session.get('f1_4')
        f1_5 = request.session.get('f1_5')
        f1_6 = request.session.get('f1_6')

        excel_dict = {}
        counter = 0
        if f1_1 == '1' or f1_1 == '2':
            counter+=1
            excel_dict['{}'.format(counter)] = 'To what extent is the purpose of the processing defined in your organization?'
        if f1_2 == '1' or f1_2 == '2':
            counter+=1
            excel_dict['{}'.format(counter)] = 'To what extent is the aim of the project defined in your organization?'
        if f1_3 == '1' or f1_2 == '2':
            counter+=1
            excel_dict['{}'.format(counter)] = 'To what extent is method of processing defined in your organization?'
        if f1_4 == '1' or f1_4 == '2':
            counter+=1
            excel_dict['{}'.format(counter)] = 'To what extent are the expected benefit(s) from the processing defined?'
        if f1_5 == '1' or f1_5 == '2':
            counter+=1
            excel_dict['{}'.format(counter)] = 'To what extent is the list of departments who will benefit from this processing defined in your organization?'
        if f1_6 == '1' or f1_6 == '2':
            counter += 1
            excel_dict['{}'.format(
                counter)] = 'To what extent is the list of departments who will benefit from this processing defined in your organization?'

        import re
        title = request.GET.get('search').split('?')[0]
        find_num = re.compile("(\d+)")
        score_str = request.GET.get('search').split('?score=')[1]
        num = find_num.match(score_str).groups()
        score = num[0]
        box_num = request.GET.get('search').split('?score=')[1].split('number=')[1]
        score_count = [i for i in range(1, int(score) + 1)]


        request.session['summary_title'] = title
        request.session['summary_score'] = score
        request.session['box_num'] = box_num
        request.session['score_count'] = score_count
        risk_score1 = request.session.get('risk_score1')
        risk_score2 = request.session.get('risk_score2')
        risk_score3 = request.session.get('risk_score3')
        risk_score4 = request.session.get('risk_score4')
        risk_score5 = request.session.get('risk_score5')
        risk_score6 = request.session.get('risk_score6')
        risk_score7 = request.session.get('risk_score7')
        risk_score8 = request.session.get('risk_score8')
        form1_percentage = request.session.get('form1_percentage')
        form2_percentage = request.session.get('form2_percentage')
        form3_percentage = request.session.get('form3_percentage')
        form4_percentage = request.session.get('form4_percentage')
        form5_percentage = request.session.get('form5_percentage')
        form6_percentage = request.session.get('form6_percentage')
        form7_percentage = request.session.get('form7_percentage')
        form8_percentage = request.session.get('form8_percentage')
        excel_risk_questions = get_excel_risk_questions(request)


        context = {
            'title': title,
            'score': score,
            'box_num': box_num,
            'excel_risk_questions': excel_risk_questions,
            'excel_dict': excel_dict,
            'score_count': score_count,
            'risk_score1': risk_score1,
            'risk_score2': risk_score2,
            'risk_score3': risk_score3,
            'risk_score4': risk_score4,
            'risk_score5': risk_score5,
            'risk_score6': risk_score6,
            'risk_score7': risk_score7,
            'risk_score8': risk_score8,
            'form1_percentage': form1_percentage,
            'form2_percentage': form2_percentage,
            'form3_percentage': form3_percentage,
            'form4_percentage': form4_percentage,
            'form5_percentage': form5_percentage,
            'form6_percentage': form6_percentage,
            'form7_percentage': form7_percentage,
            'form8_percentage': form8_percentage,
            'name_of_controller': request.session.get('manager'),
            'name_of_dpo': request.session.get('name_of_DPO'),
            'title_of_dpo': request.session.get('title_of_DPO')
        }
        return render(request, 'risk_summary_details.html', context)


def replace_value_1_2(value):
    if value is not None:
        if value == '1':
            ex_value = 'Absent'
        else:
            ex_value = 'Basic'
        return ex_value


def replace_value_init_define(value):
    if value is not None:
        if value == '1':
            ex_value = 'Initial'
        else:
            ex_value = 'Defined'
        return ex_value

def replace_value_yes_no(value):
    if value is not None:
        if value == '0':
            ex_value = 'Yes'
        else:
            ex_value = 'No'
        return ex_value

def replace_value_process_no_process(value):
    if value is not None:
        if value == '0':
            ex_value = 'No Process'
        else:
            ex_value = 'Recommended'
        return ex_value


def replace_values_likelihood(value):
    if value is None:
        return None
    if value == '1':
        text = 'Unlikely'
    elif value == '2':
        text = 'Possibly'
    else:
        text = 'Highly Likely'
    return text


def replace_values_severity(value):
    if value is None:
        return None
    if value == '1':
        text = 'Minor'
    elif value == '2':
        text = 'Moderate'
    else:
        text = 'Significant'
    return text


def effect_on_risk_logic(value):
    if value is None:
        return None
    logic = ''
    if value == '1':
        logic = 'Eliminated'
    elif value == '2':
        logic = 'Reduced'
    else:
        logic = 'Accepted'
    return logic


def residual_risk_logic(value):
    if value is None:
        return None
    logic = ''
    if value == '1':
        logic = 'Low'
    elif value == '2':
        logic = 'Medium'
    else:
        logic = 'High'
    return logic


def measure_approved_logic(value):
    if value is None:
        return None
    logic = ''
    if value == '0':
        logic = 'Done'
    else:
        logic = 'Not Done'
    return logic


def text_to_excel(request, excel_data, liklihood_dict):
    # load excel file
    workbook = load_workbook(filename="media/excel/summary_report_blank.xlsx")

    # open workbook
    sheet = workbook.active

    # modify the desired cell
    sheet["C7"] = request.session.get('title')
    sheet["C8"] = request.session.get('author')
    sheet["C9"] = request.session.get('name_of_DPO')
    sheet["C10"] = request.session.get('date')
    sheet["C11"] = request.session.get('status')

    c = 16
    for val in excel_data.items():
        sheet['C{}'.format(c)] = val[1]
        c+=1

    risk_score1 = request.session.get('risk_score1')
    risk_score2 = request.session.get('risk_score2')
    risk_score3 = request.session.get('risk_score3')
    risk_score4 = request.session.get('risk_score4')
    risk_score5 = request.session.get('risk_score5')
    risk_score6 = request.session.get('risk_score6')
    risk_score7 = request.session.get('risk_score7')
    risk_score8 = request.session.get('risk_score8')


    # Liklihood data on excel
    liklihood_final_list = []
    if liklihood_dict.get('s1_b1') is not None:
        for i in range(1, risk_score1[1] + 1):
            liklihood_final_list.append(liklihood_dict['s1_b{}'.format(i)]['s1_b{}_likelihood'.format(i)])

    if liklihood_dict.get('s2_b1') is not None:
        for i in range(1, risk_score2[1] + 1):
            liklihood_final_list.append(liklihood_dict['s2_b{}'.format(i)]['s2_b{}_likelihood'.format(i)])

    if liklihood_dict.get('s3_b1') is not None:
        for i in range(1, risk_score3[1] + 1):
            liklihood_final_list.append(liklihood_dict['s3_b{}'.format(i)]['s3_b{}_likelihood'.format(i)])

    if liklihood_dict.get('s4_b1') is not None:
        for i in range(1, risk_score4[1] + 1):
            liklihood_final_list.append(liklihood_dict['s4_b{}'.format(i)]['s4_b{}_likelihood'.format(i)])

    if liklihood_dict.get('s5_b1') is not None:
        for i in range(1, risk_score5[1] + 1):
            liklihood_final_list.append(liklihood_dict['s5_b{}'.format(i)]['s5_b{}_likelihood'.format(i)])

    if liklihood_dict.get('s6_b1') is not None:
        for i in range(1, risk_score6[1] + 1):
            liklihood_final_list.append(liklihood_dict['s6_b{}'.format(i)]['s6_b{}_likelihood'.format(i)])

    if liklihood_dict.get('s7_b1') is not None:
        for i in range(1, risk_score7[1] + 1):
            liklihood_final_list.append(liklihood_dict['s7_b{}'.format(i)]['s7_b{}_likelihood'.format(i)])

    if liklihood_dict.get('s8_b1') is not None:
        for i in range(1, risk_score8[1] + 1):
            liklihood_final_list.append(liklihood_dict['s8_b{}'.format(i)]['s8_b{}_likelihood'.format(i)])

    set_count = 16
    for count in liklihood_final_list:
        sheet['E{}'.format(set_count)] = count
        set_count += 1


    # Severity data on excel

    severity_final_list = []
    if liklihood_dict.get('s1_b1') is not None:
        for i in range(1, risk_score1[1] + 1):
            severity_final_list.append(liklihood_dict['s1_b{}'.format(i)]['s1_b{}_severity'.format(i)])

    if liklihood_dict.get('s2_b1') is not None:
        for i in range(1, risk_score2[1] + 1):
            severity_final_list.append(liklihood_dict['s2_b{}'.format(i)]['s2_b{}_severity'.format(i)])

    if liklihood_dict.get('s3_b1') is not None:
        for i in range(1, risk_score3[1] + 1):
            severity_final_list.append(liklihood_dict['s3_b{}'.format(i)]['s3_b{}_severity'.format(i)])

    if liklihood_dict.get('s4_b1') is not None:
        for i in range(1, risk_score4[1] + 1):
            severity_final_list.append(liklihood_dict['s4_b{}'.format(i)]['s4_b{}_severity'.format(i)])

    if liklihood_dict.get('s5_b1') is not None:
        for i in range(1, risk_score5[1] + 1):
            severity_final_list.append(liklihood_dict['s5_b{}'.format(i)]['s5_b{}_severity'.format(i)])

    if liklihood_dict.get('s6_b1') is not None:
        for i in range(1, risk_score6[1] + 1):
            severity_final_list.append(liklihood_dict['s6_b{}'.format(i)]['s6_b{}_severity'.format(i)])

    if liklihood_dict.get('s7_b1') is not None:
        for i in range(1, risk_score7[1] + 1):
            severity_final_list.append(liklihood_dict['s7_b{}'.format(i)]['s7_b{}_severity'.format(i)])

    if liklihood_dict.get('s8_b1') is not None:
        for i in range(1, risk_score8[1] + 1):
            severity_final_list.append(liklihood_dict['s8_b{}'.format(i)]['s8_b{}_severity'.format(i)])

    set_count = 16
    for count in severity_final_list:
        sheet['F{}'.format(set_count)] = count
        set_count += 1

    # Overall Risk data on excel

    overall_risk_final_list = []
    if liklihood_dict.get('s1_b1') is not None:
        for i in range(1, risk_score1[1] + 1):
            overall_risk_final_list.append(liklihood_dict['s1_b{}'.format(i)]['s1_b{}_overall_risk'.format(i)])

    if liklihood_dict.get('s2_b1') is not None:
        for i in range(1, risk_score2[1] + 1):
            overall_risk_final_list.append(liklihood_dict['s2_b{}'.format(i)]['s2_b{}_overall_risk'.format(i)])

    if liklihood_dict.get('s3_b1') is not None:
        for i in range(1, risk_score3[1] + 1):
            overall_risk_final_list.append(liklihood_dict['s3_b{}'.format(i)]['s3_b{}_overall_risk'.format(i)])

    if liklihood_dict.get('s4_b1') is not None:
        for i in range(1, risk_score4[1] + 1):
            overall_risk_final_list.append(liklihood_dict['s4_b{}'.format(i)]['s4_b{}_overall_risk'.format(i)])

    if liklihood_dict.get('s5_b1') is not None:
        for i in range(1, risk_score5[1] + 1):
            overall_risk_final_list.append(liklihood_dict['s5_b{}'.format(i)]['s5_b{}_overall_risk'.format(i)])

    if liklihood_dict.get('s6_b1') is not None:
        for i in range(1, risk_score6[1] + 1):
            overall_risk_final_list.append(liklihood_dict['s6_b{}'.format(i)]['s6_b{}_overall_risk'.format(i)])

    if liklihood_dict.get('s7_b1') is not None:
        for i in range(1, risk_score7[1] + 1):
            overall_risk_final_list.append(liklihood_dict['s7_b{}'.format(i)]['s7_b{}_overall_risk'.format(i)])

    if liklihood_dict.get('s8_b1') is not None:
        for i in range(1, risk_score8[1] + 1):
            overall_risk_final_list.append(liklihood_dict['s8_b{}'.format(i)]['s8_b{}_overall_risk'.format(i)])

    set_count = 16
    for count in overall_risk_final_list:
        sheet['G{}'.format(set_count)] = count
        set_count += 1

    # Option Reduce Risk data

    opt_redu_risk_list = []
    if liklihood_dict.get('s1_b1') is not None:
        for i in range(1, risk_score1[1] + 1):
            opt_redu_risk_list.append(liklihood_dict['s1_b{}'.format(i)]['s1_b{}_option_reduce_risk'.format(i)])

    if liklihood_dict.get('s2_b1') is not None:
        for i in range(1, risk_score2[1] + 1):
            opt_redu_risk_list.append(liklihood_dict['s2_b{}'.format(i)]['s2_b{}_option_reduce_risk'.format(i)])

    if liklihood_dict.get('s3_b1') is not None:
        for i in range(1, risk_score3[1] + 1):
            opt_redu_risk_list.append(liklihood_dict['s3_b{}'.format(i)]['s3_b{}_option_reduce_risk'.format(i)])

    if liklihood_dict.get('s4_b1') is not None:
        for i in range(1, risk_score4[1] + 1):
            opt_redu_risk_list.append(liklihood_dict['s4_b{}'.format(i)]['s4_b{}_option_reduce_risk'.format(i)])

    if liklihood_dict.get('s5_b1') is not None:
        for i in range(1, risk_score5[1] + 1):
            opt_redu_risk_list.append(liklihood_dict['s5_b{}'.format(i)]['s5_b{}_option_reduce_risk'.format(i)])

    if liklihood_dict.get('s6_b1') is not None:
        for i in range(1, risk_score6[1] + 1):
            opt_redu_risk_list.append(liklihood_dict['s6_b{}'.format(i)]['s6_b{}_option_reduce_risk'.format(i)])

    if liklihood_dict.get('s7_b1') is not None:
        for i in range(1, risk_score7[1] + 1):
            opt_redu_risk_list.append(liklihood_dict['s7_b{}'.format(i)]['s7_b{}_option_reduce_risk'.format(i)])

    if liklihood_dict.get('s8_b1') is not None:
        for i in range(1, risk_score8[1] + 1):
            opt_redu_risk_list.append(liklihood_dict['s8_b{}'.format(i)]['s8_b{}_option_reduce_risk'.format(i)])

    set_count = 16
    for count in opt_redu_risk_list:
        sheet['I{}'.format(set_count)] = count
        set_count += 1

    # Effect on Risk data on excel
    effect_risk_list = []
    if liklihood_dict.get('s1_b1') is not None:
        for i in range(1, risk_score1[1] + 1):
            effect_risk_list.append(liklihood_dict['s1_b{}'.format(i)]['s1_b{}_effect_on_risk'.format(i)])

        if liklihood_dict.get('s2_b1') is not None:
            for i in range(1, risk_score2[1] + 1):
                effect_risk_list.append(liklihood_dict['s2_b{}'.format(i)]['s2_b{}_effect_on_risk'.format(i)])

        if liklihood_dict.get('s3_b1') is not None:
            for i in range(1, risk_score3[1] + 1):
                effect_risk_list.append(liklihood_dict['s3_b{}'.format(i)]['s3_b{}_effect_on_risk'.format(i)])

        if liklihood_dict.get('s4_b1') is not None:
            for i in range(1, risk_score4[1] + 1):
                effect_risk_list.append(liklihood_dict['s4_b{}'.format(i)]['s4_b{}_effect_on_risk'.format(i)])

        if liklihood_dict.get('s5_b1') is not None:
            for i in range(1, risk_score5[1] + 1):
                effect_risk_list.append(liklihood_dict['s5_b{}'.format(i)]['s5_b{}_effect_on_risk'.format(i)])

        if liklihood_dict.get('s6_b1') is not None:
            for i in range(1, risk_score6[1] + 1):
                effect_risk_list.append(liklihood_dict['s6_b{}'.format(i)]['s6_b{}_effect_on_risk'.format(i)])

        if liklihood_dict.get('s7_b1') is not None:
            for i in range(1, risk_score7[1] + 1):
                effect_risk_list.append(liklihood_dict['s7_b{}'.format(i)]['s7_b{}_effect_on_risk'.format(i)])

        if liklihood_dict.get('s8_b1') is not None:
            for i in range(1, risk_score8[1] + 1):
                effect_risk_list.append(liklihood_dict['s8_b{}'.format(i)]['s8_b{}_effect_on_risk'.format(i)])

        set_count = 16
        for count in effect_risk_list:
            sheet['K{}'.format(set_count)] = count
            set_count += 1


        # Residual Risk data on excel

        residual_risk_list = []
        if liklihood_dict.get('s1_b1') is not None:
            for i in range(1, risk_score1[1] + 1):
                residual_risk_list.append(liklihood_dict['s1_b{}'.format(i)]['s1_b{}_residual_risk'.format(i)])

        if liklihood_dict.get('s2_b1') is not None:
            for i in range(1, risk_score2[1] + 1):
                residual_risk_list.append(liklihood_dict['s2_b{}'.format(i)]['s2_b{}_residual_risk'.format(i)])

        if liklihood_dict.get('s3_b1') is not None:
            for i in range(1, risk_score3[1] + 1):
                residual_risk_list.append(liklihood_dict['s3_b{}'.format(i)]['s3_b{}_residual_risk'.format(i)])

        if liklihood_dict.get('s4_b1') is not None:
            for i in range(1, risk_score4[1] + 1):
                residual_risk_list.append(liklihood_dict['s4_b{}'.format(i)]['s4_b{}_residual_risk'.format(i)])

        if liklihood_dict.get('s5_b1') is not None:
            for i in range(1, risk_score5[1] + 1):
                residual_risk_list.append(liklihood_dict['s5_b{}'.format(i)]['s5_b{}_residual_risk'.format(i)])

        if liklihood_dict.get('s6_b1') is not None:
            for i in range(1, risk_score6[1] + 1):
                residual_risk_list.append(liklihood_dict['s6_b{}'.format(i)]['s6_b{}_residual_risk'.format(i)])

        if liklihood_dict.get('s7_b1') is not None:
            for i in range(1, risk_score7[1] + 1):
                residual_risk_list.append(liklihood_dict['s7_b{}'.format(i)]['s7_b{}_residual_risk'.format(i)])

        if liklihood_dict.get('s8_b1') is not None:
            for i in range(1, risk_score8[1] + 1):
                residual_risk_list.append(liklihood_dict['s8_b{}'.format(i)]['s8_b{}_residual_risk'.format(i)])

        set_count = 16
        for count in residual_risk_list:
            sheet['L{}'.format(set_count)] = count
            set_count += 1

        # Measured Approved data on excel

        measured_risk_list = []
        if liklihood_dict.get('s1_b1') is not None:
            for i in range(1, risk_score1[1] + 1):
                measured_risk_list.append(liklihood_dict['s1_b{}'.format(i)]['s1_b{}_measure_approved'.format(i)])

        if liklihood_dict.get('s2_b1') is not None:
            for i in range(1, risk_score2[1] + 1):
                measured_risk_list.append(liklihood_dict['s2_b{}'.format(i)]['s2_b{}_measure_approved'.format(i)])

        if liklihood_dict.get('s3_b1') is not None:
            for i in range(1, risk_score3[1] + 1):
                measured_risk_list.append(liklihood_dict['s3_b{}'.format(i)]['s3_b{}_measure_approved'.format(i)])

        if liklihood_dict.get('s4_b1') is not None:
            for i in range(1, risk_score4[1] + 1):
                measured_risk_list.append(liklihood_dict['s4_b{}'.format(i)]['s4_b{}_measure_approved'.format(i)])

        if liklihood_dict.get('s5_b1') is not None:
            for i in range(1, risk_score5[1] + 1):
                measured_risk_list.append(liklihood_dict['s5_b{}'.format(i)]['s5_b{}_measure_approved'.format(i)])

        if liklihood_dict.get('s6_b1') is not None:
            for i in range(1, risk_score6[1] + 1):
                measured_risk_list.append(liklihood_dict['s6_b{}'.format(i)]['s6_b{}_measure_approved'.format(i)])

        if liklihood_dict.get('s7_b1') is not None:
            for i in range(1, risk_score7[1] + 1):
                measured_risk_list.append(liklihood_dict['s7_b{}'.format(i)]['s7_b{}_measure_approved'.format(i)])

        if liklihood_dict.get('s8_b1') is not None:
            for i in range(1, risk_score8[1] + 1):
                measured_risk_list.append(liklihood_dict['s8_b{}'.format(i)]['s8_b{}_measure_approved'.format(i)])

        set_count = 16
        for count in measured_risk_list:
            sheet['M{}'.format(set_count)] = count
            set_count += 1

        response_list = request.session.get('response_list')
        set_count = 16
        for count in response_list:
            sheet['D{}'.format(set_count)] = count
            set_count += 1

    row_count = sheet.max_row
    column_count = sheet.max_column
    print('Row Count: ', row_count)
    print('Column Count: ', column_count)

    total_no_of_risk = request.session.get('total_no_of_risk')
    sheet.delete_rows(total_no_of_risk + 16, 50)

    # save the file
    workbook.save(filename="media/excel/summary_report_filled.xlsx")


def dpia_status(input_data):
    status = 'Mandatory'
    if input_data.get('data_processing_project') == '0':
        if input_data.get('select_data_process') == '4':
            status = 'Recommended'
        elif input_data.get('select_data_process') != '4' or \
                input_data.get('select_data_process') == '3' or \
                input_data.get('select_data_process') == '2' or \
            input_data.get('select_data_process') == '1':
            status = 'Mandatory'
        else:
            status = 'Not Required'
    else:
        status = 'Not Required'
    return status


def screening(request):
    if request.method == 'GET':
        session_dict = get_session_data()
        count = 1
        context = {}
        session_data = False
        for i in session_dict:
            if request.session['title'] == session_dict.get('session_dict_{}'.format(count)).get('title') \
                    and request.session['date'] == session_dict.get('session_dict_{}'.format(count)).get('date'):
                context['value_dict'] = session_dict.get('session_dict_{}'.format(count))
                session_data = True
            count += 1
        if not session_data:
            context['value_dict'] = 0
        context['session_dict'] = session_dict
        return render(request, 'screening.html', context)
    if request.method == 'POST':
        request.session['name_of_organization'] = request.POST.get("name_of_organization")
        request.session['industry'] = request.POST.get("industry")
        request.session['scope_of_service_project'] = request.POST.get("scope_of_service_project")
        request.session['data_protection_officer'] = request.POST.get("data_protection_officer")
        request.session['name_of_DPO'] = request.POST.get("name_of_DPO")
        request.session['title_of_DPO'] = request.POST.get("title_of_DPO")

        name_of_organization = request.session['name_of_organization']
        industry = request.session['industry']
        scope_of_service_project = request.session['scope_of_service_project']
        data_protection_officer = request.session['data_protection_officer']
        name_of_DPO = request.session['name_of_DPO']
        title_of_DPO = request.session['title_of_DPO']

        input_data = {
            'name_of_organization': name_of_organization,
            'industry': industry,
            'scope_of_service_project': scope_of_service_project,
            'data_protection_officer': data_protection_officer,
            'name_of_DPO': name_of_DPO,
            'title_of_DPO': title_of_DPO,
        }
        session_dict = get_session_data()
        count = 1
        context = {}
        session_data = False
        for i in session_dict:
            if request.session['title'] == session_dict.get('session_dict_{}'.format(count)).get('title')\
                    and request.session['date'] == session_dict.get('session_dict_{}'.format(count)).get('date'):
                context['value_dict'] = session_dict.get('session_dict_{}'.format(count))
                session_data = True
            count += 1
        if not session_data:
            context['value_dict'] = 0
        context['input_data'] = input_data
        context['session_dict'] = session_dict
        context['dpia_status'] = dpia_status(input_data)
        return render(request, 'screening1.html', context)
    return render(request, 'screening.html')


def screening1(request):
    if request.method == 'POST':
        request.session['data_processing_project'] = request.POST.get("data_processing_project")
        request.session['select_data_process'] = request.POST.get("select_data_process")
        request.session['processing_data'] = request.POST.get("processing_data")
        request.session['data_processing_involve'] = request.POST.get("data_processing_involve")
        request.session['automated_decision_making'] = request.POST.get("automated_decision_making")
        request.session['systematic_monitoring'] = request.POST.get("systematic_monitoring")
        request.session['process_data_on_large_scale'] = request.POST.get("process_data_on_large_scale")
        request.session['data_processing_involve_reusing_old_dataset'] = request.POST.get("data_processing_involve_reusing_old_dataset")
        request.session['vulnerable_data_subjects'] = request.POST.get("vulnerable_data_subjects")
        request.session['data_processing_involve_innovative_technologies'] = request.POST.get("data_processing_involve_innovative_technologies")
        request.session['data_processing_involve_sharing_data_outside_european_union'] = request.POST.get("data_processing_involve_sharing_data_outside_european_union")
        request.session['data_processing_involve_collection_personal_information'] = request.POST.get("data_processing_involve_collection_personal_information")
        request.session['data_processing_involve_third_party'] = request.POST.get("data_processing_involve_third_party")
        request.session['data_processing_involve_change_information_is_stored_secured'] = request.POST.get("data_processing_involve_change_information_is_stored_secured")
        request.session['data_procc_involve_chg_personal_data_currently_collected'] = request.POST.get("data_procc_involve_chg_personal_data_currently_collected")
        request.session['conducted_DPIA_for_similar_scope_of_service'] = request.POST.get("conducted_DPIA_for_similar_scope_of_service")

        data_processing_project = request.session['data_processing_project']
        select_data_process = request.session['select_data_process']
        processing_data = request.session['processing_data']
        data_processing_involve = request.session['data_processing_involve']
        automated_decision_making = request.session['automated_decision_making']
        systematic_monitoring = request.session['systematic_monitoring']
        process_data_on_large_scale = request.session['process_data_on_large_scale']
        data_processing_involve_reusing_old_dataset = request.session['data_processing_involve_reusing_old_dataset']
        vulnerable_data_subjects = request.session['vulnerable_data_subjects']
        data_processing_involve_innovative_technologies = request.session['data_processing_involve_innovative_technologies']
        data_processing_involve_sharing_data_outside_european_union = request.session['data_processing_involve_sharing_data_outside_european_union']
        data_processing_involve_collection_personal_information = request.session['data_processing_involve_collection_personal_information']
        data_processing_involve_third_party = request.session['data_processing_involve_third_party']
        data_processing_involve_change_information_is_stored_secured = request.session['data_processing_involve_change_information_is_stored_secured']
        data_procc_involve_chg_personal_data_currently_collected = request.session['data_procc_involve_chg_personal_data_currently_collected']
        conducted_DPIA_for_similar_scope_of_service = request.session['conducted_DPIA_for_similar_scope_of_service']

        input_data = {
            'data_processing_project': data_processing_project,
            'select_data_process': select_data_process,
            'processing_data': processing_data,
            'data_processing_involve': data_processing_involve,
            'automated_decision_making': automated_decision_making,
            'systematic_monitoring': systematic_monitoring,
            'process_data_on_large_scale': process_data_on_large_scale,
            'data_processing_involve_reusing_old_dataset': data_processing_involve_reusing_old_dataset,
            'vulnerable_data_subjects': vulnerable_data_subjects,
            'data_processing_involve_innovative_technologies': data_processing_involve_innovative_technologies,
            'data_processing_involve_sharing_data_outside_european_union': data_processing_involve_sharing_data_outside_european_union,
            'data_processing_involve_collection_personal_information': data_processing_involve_collection_personal_information,
            'data_processing_involve_third_party': data_processing_involve_third_party,
            'data_processing_involve_change_information_is_stored_secured': data_processing_involve_change_information_is_stored_secured,
            'data_procc_involve_chg_personal_data_currently_collected': data_procc_involve_chg_personal_data_currently_collected,
            'conducted_DPIA_for_similar_scope_of_service': conducted_DPIA_for_similar_scope_of_service,

        }
        session_dict = get_session_data()
        context = {
            'input_data': input_data,
            'dpia_status': dpia_status(input_data),
            'session_dict': session_dict,
        }
        return render(request, 'result.html', context)
    return render(request, 'screening.html')


class RiskCalculations:
    def __init__(self, input_data, request):
        self.input_data = input_data
        self.risk_score = 0
        self.number_of_risks = 0
        self.request = request
        self.form1 = 6
        self.form2 = 9
        self.form3 = 12
        self.form4 = 13
        self.form5 = 5
        self.form6 = 5
        self.form7 = 2
        self.form8 = 4

    # Risk Summary Form 1
    def risk_calculation_f1_1(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f1_1') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f1_1') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f1_1') == '3':
            risk_score += 1
        if self.input_data.get('f1_1') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        self.request.session['f1_1'] = self.input_data.get('f1_1')
        self.request.session['risk_score_f1_1'] = total_risk[0]
        self.request.session['number_of_risks_f1_1'] = total_risk[1]
        return total_risk

    def risk_calculation_f1_2(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f1_2') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f1_2') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f1_2') == '3':
            risk_score += 1
        if self.input_data.get('f1_2') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        self.request.session['f1_2'] = self.input_data.get('f1_2')
        self.request.session['risk_score_f1_2'] = total_risk[0]
        self.request.session['number_of_risks_f1_2'] = total_risk[1]
        return total_risk

    def risk_calculation_f1_3(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f1_3') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f1_3') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f1_3') == '3':
            risk_score += 1
        if self.input_data.get('f1_3') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        self.request.session['f1_3'] = self.input_data.get('f1_3')
        self.request.session['risk_score_f1_3'] = total_risk[0]
        self.request.session['number_of_risks_f1_3'] = total_risk[1]
        return total_risk

    def risk_calculation_f1_4(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f1_4') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f1_4') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f1_4') == '3':
            risk_score += 1
        if self.input_data.get('f1_4') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        self.request.session['f1_4'] = self.input_data.get('f1_4')
        self.request.session['risk_score_f1_4'] = total_risk[0]
        self.request.session['number_of_risks_f1_4'] = total_risk[1]
        return total_risk

    def risk_calculation_f1_5(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f1_5') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f1_5') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f1_5') == '3':
            risk_score += 1
        if self.input_data.get('f1_5') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        self.request.session['f1_5'] = self.input_data.get('f1_5')
        self.request.session['risk_score_f1_5'] = total_risk[0]
        self.request.session['number_of_risks_f1_5'] = total_risk[1]
        return total_risk

    def risk_calculation_f1_6(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f1_6') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f1_6') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f1_6') == '3':
            risk_score += 1
        if self.input_data.get('f1_6') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        self.request.session['f1_6'] = self.input_data.get('f1_6')
        self.request.session['risk_score_f1_6'] = total_risk[0]
        self.request.session['number_of_risks_f1_6'] = total_risk[1]
        return total_risk

    def risk_calculation_f1_all(self):
        f1_1 = self.risk_calculation_f1_1()
        f1_2 = self.risk_calculation_f1_2()
        f1_3 = self.risk_calculation_f1_3()
        f1_4 = self.risk_calculation_f1_4()
        f1_5 = self.risk_calculation_f1_5()
        f1_6 = self.risk_calculation_f1_6()
        risk_score = f1_1[0] + f1_2[0] + f1_3[0] + f1_4[0] + f1_5[0] + f1_6[0]
        number_of_risks = f1_1[1] + f1_2[1] + f1_3[1] + f1_4[1] + f1_5[1] + f1_6[1]
        total_risk = [risk_score, number_of_risks]
        return total_risk

    # Risk Summary Form 2
    def risk_calculation_f2_1(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f2_1') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f2_1') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f2_1') == '3':
            risk_score += 1
        if self.input_data.get('f2_1') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        self.request.session['f2_1'] = self.input_data.get('f2_1')
        self.request.session['risk_score_f2_1'] = total_risk[0]
        self.request.session['number_of_risks_f2_1'] = total_risk[1]
        return total_risk

    def risk_calculation_f2_2(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f2_2') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f2_2') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f2_2') == '3':
            risk_score += 1
        if self.input_data.get('f2_2') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        self.request.session['f2_2'] = self.input_data.get('f2_2')
        self.request.session['risk_score_f2_2'] = total_risk[0]
        self.request.session['number_of_risks_f2_2'] = total_risk[1]
        return total_risk

    def risk_calculation_f2_3(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f2_3') == '0':
            risk_score += 0
            number_of_risks += 0
        elif self.input_data.get('f2_3') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        self.request.session['f2_3'] = self.input_data.get('f2_3')
        self.request.session['risk_score_f2_1'] = total_risk[0]
        self.request.session['number_of_risks_f2_1'] = total_risk[1]
        return total_risk

    def risk_calculation_f2_4(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f2_4') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f2_4') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f2_4') == '3':
            risk_score += 1
        if self.input_data.get('f2_4') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        self.request.session['f2_1'] = self.input_data.get('f2_1')
        self.request.session['risk_score_f2_1'] = total_risk[0]
        self.request.session['number_of_risks_f2_1'] = total_risk[1]
        return total_risk

    def risk_calculation_f2_5(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f2_5') == '0':
            risk_score += 0
            number_of_risks += 0
        elif self.input_data.get('f2_5') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        self.request.session['f2_1'] = self.input_data.get('f2_1')
        self.request.session['risk_score_f2_1'] = total_risk[0]
        self.request.session['number_of_risks_f2_1'] = total_risk[1]
        return total_risk

    def risk_calculation_f2_6(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f2_6') == '1':
            number_of_risks += 0
        elif self.input_data.get('f2_6') == '2':
            risk_score += 1
            number_of_risks += 1
        elif self.input_data.get('f2_6') == '3':
            risk_score += 2
            number_of_risks += 1
        if self.input_data.get('f2_6') == '4':
            risk_score += 3
            number_of_risks += 1
        if self.input_data.get('f2_6') == '5':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        self.request.session['f2_1'] = self.input_data.get('f2_1')
        self.request.session['risk_score_f2_1'] = total_risk[0]
        self.request.session['number_of_risks_f2_1'] = total_risk[1]
        return total_risk

    def risk_calculation_f2_7(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f2_7') == '1':
            number_of_risks += 1
        elif self.input_data.get('f2_7') == '2':
            number_of_risks += 1
        elif self.input_data.get('f2_7') == '3':
            number_of_risks += 1
        elif self.input_data.get('f2_7') == '4':
            number_of_risks += 1
        if self.input_data.get('f2_7') == '5':
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        self.request.session['f2_1'] = self.input_data.get('f2_1')
        self.request.session['risk_score_f2_1'] = total_risk[0]
        self.request.session['number_of_risks_f2_1'] = total_risk[1]
        return total_risk

    def risk_calculation_f2_8(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f2_8') == '0':
            risk_score += 0
            number_of_risks += 0
        elif self.input_data.get('f2_8') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        self.request.session['f2_1'] = self.input_data.get('f2_1')
        self.request.session['risk_score_f2_1'] = total_risk[0]
        self.request.session['number_of_risks_f2_1'] = total_risk[1]
        return total_risk

    def risk_calculation_f2_9(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f2_9') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f2_9') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f2_9') == '3':
            risk_score += 1
        if self.input_data.get('f2_9') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        self.request.session['f2_1'] = self.input_data.get('f2_1')
        self.request.session['risk_score_f2_1'] = total_risk[0]
        self.request.session['number_of_risks_f2_1'] = total_risk[1]
        return total_risk

    def risk_calculation_f2_all(self):
        f2_1 = self.risk_calculation_f2_1()
        f2_2 = self.risk_calculation_f2_2()
        f2_3 = self.risk_calculation_f2_3()
        f2_4 = self.risk_calculation_f2_4()
        f2_5 = self.risk_calculation_f2_5()
        f2_6 = self.risk_calculation_f2_6()
        f2_7 = self.risk_calculation_f2_7()
        f2_8 = self.risk_calculation_f2_8()
        f2_9 = self.risk_calculation_f2_9()
        risk_score = f2_1[0] + f2_2[0] + f2_3[0] + f2_4[0] + f2_5[0] + f2_6[0] + f2_7[0] + f2_8[0] + f2_9[0]
        number_of_risks = f2_1[1] + f2_2[1] + f2_3[1] + f2_4[1] + f2_5[1] + f2_6[1] + f2_7[1] + f2_8[1] + f2_9[1]
        total_risk = [risk_score, number_of_risks]
        return total_risk

    # Risk Summary Form 3
    def risk_calculation_f3_1(self):
        risk_score = 0
        number_of_risks = 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f3_2(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f3_2') == '0':
            risk_score += 0
        elif self.input_data.get('f3_2') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f3_3(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f3_3') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f3_3') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f3_3') == '3':
            risk_score += 1
        if self.input_data.get('f3_3') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f3_4(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f3_4') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f3_4') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f3_4') == '3':
            risk_score += 1
        if self.input_data.get('f3_4') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f3_5(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f3_5') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f3_5') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f3_5') == '3':
            risk_score += 1
        if self.input_data.get('f3_5') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f3_6(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f3_6') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f3_6') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f3_6') == '3':
            risk_score += 1
        if self.input_data.get('f3_6') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f3_7(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f3_7') == '0':
            risk_score += 0
        elif self.input_data.get('f3_7') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f3_8(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f3_8') == '0':
            risk_score += 0
        elif self.input_data.get('f3_8') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f3_9(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f3_9') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f3_9') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f3_9') == '3':
            risk_score += 1
        if self.input_data.get('f3_9') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f3_10(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f3_10') == '0':
            risk_score += 0
        elif self.input_data.get('f3_10') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f3_11(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f3_11') == '0':
            risk_score += 0
        elif self.input_data.get('f3_11') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f3_12(self):
        risk_score = 0
        number_of_risks = 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f3_all(self):
        f3_1 = self.risk_calculation_f3_1()
        f3_2 = self.risk_calculation_f3_2()
        f3_3 = self.risk_calculation_f3_3()
        f3_4 = self.risk_calculation_f3_4()
        f3_5 = self.risk_calculation_f3_5()
        f3_6 = self.risk_calculation_f3_6()
        f3_7 = self.risk_calculation_f3_7()
        f3_8 = self.risk_calculation_f3_8()
        f3_9 = self.risk_calculation_f3_9()
        f3_10 = self.risk_calculation_f3_10()
        f3_11 = self.risk_calculation_f3_11()
        f3_12 = self.risk_calculation_f3_12()

        risk_score = f3_1[0] + f3_2[0] + f3_3[0] + f3_4[0] + f3_5[0] + f3_6[0] + f3_7[0] + f3_8[0] + f3_9[0] + f3_10[0] + f3_11[0] + f3_12[0]
        number_of_risks = f3_1[1] + f3_2[1] + f3_3[1] + f3_4[1] + f3_5[1] + f3_6[1] + f3_7[1] + f3_8[1] + f3_9[1] + f3_10[1] + f3_11[1] + f3_12[1]
        total_risk = [risk_score, number_of_risks]
        return total_risk

    # Risk Summary Form 4
    def risk_calculation_f4_1(self):
        risk_score = 0
        number_of_risks = 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f4_2(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f4_2') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f4_2') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f4_2') == '3':
            risk_score += 1
        if self.input_data.get('f4_2') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f4_3(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f4_3') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f4_3') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f4_3') == '3':
            risk_score += 1
        if self.input_data.get('f4_3') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f4_4(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f4_4') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f4_4') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f4_4') == '3':
            risk_score += 1
        if self.input_data.get('f4_4') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f4_5(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f4_5') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f4_5') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f4_5') == '3':
            risk_score += 1
        if self.input_data.get('f4_5') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f4_6(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f4_6') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f4_6') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f4_6') == '3':
            risk_score += 1
        if self.input_data.get('f4_6') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f4_7(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f4_7') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f4_7') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f4_7') == '3':
            risk_score += 1
        if self.input_data.get('f4_7') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f4_8(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f4_8') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f4_8') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f4_8') == '3':
            risk_score += 1
        if self.input_data.get('f4_8') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f4_9(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f4_9') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f4_9') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f4_9') == '3':
            risk_score += 1
        if self.input_data.get('f4_9') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f4_10(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f4_10') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f4_10') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f4_10') == '3':
            risk_score += 1
        if self.input_data.get('f4_10') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f4_11(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f4_11') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f4_11') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f4_11') == '3':
            risk_score += 1
        if self.input_data.get('f4_11') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f4_12(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f4_12') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f4_12') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f4_12') == '3':
            risk_score += 1
        if self.input_data.get('f4_12') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f4_13(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f4_13') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f4_13') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f4_13') == '3':
            risk_score += 1
        if self.input_data.get('f4_13') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f4_all(self):
        f4_1 = self.risk_calculation_f4_1()
        f4_2 = self.risk_calculation_f4_2()
        f4_3 = self.risk_calculation_f4_3()
        f4_4 = self.risk_calculation_f4_4()
        f4_5 = self.risk_calculation_f4_5()
        f4_6 = self.risk_calculation_f4_6()
        f4_7 = self.risk_calculation_f4_7()
        f4_8 = self.risk_calculation_f4_8()
        f4_9 = self.risk_calculation_f4_9()
        f4_10 = self.risk_calculation_f4_10()
        f4_11 = self.risk_calculation_f4_11()
        f4_12 = self.risk_calculation_f4_12()
        f4_13 = self.risk_calculation_f4_13()


        risk_score = f4_1[0] + f4_2[0] + f4_3[0] + f4_4[0] + f4_5[0] + f4_6[0] + f4_7[0] + f4_8[0] + f4_9[0] + + f4_10[0] + f4_11[0] + f4_12[0] + f4_13[0]
        number_of_risks = f4_1[1] + f4_2[1] + f4_3[1] + f4_4[1] + f4_5[1] + f4_6[1] + f4_7[1] + f4_8[1] + f4_9[1] + f4_10[1] + f4_11[1] + f4_12[1] + f4_13[1]
        total_risk = [risk_score, number_of_risks]
        return total_risk

    # Risk Summary Form 5
    def risk_calculation_f5_1(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f5_1') == '0':
            risk_score = 0
            number_of_risks = 0
        elif self.input_data.get('f5_1') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f5_2(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f5_2') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f5_2') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f5_2') == '3':
            risk_score += 1
        if self.input_data.get('f5_2') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f5_3(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f5_3') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f5_3') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f5_3') == '3':
            risk_score += 1
        if self.input_data.get('f5_3') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f5_4(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f5_4') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f5_4') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f5_4') == '3':
            risk_score += 1
        if self.input_data.get('f5_4') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f5_5(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f5_5') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f5_5') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f5_5') == '3':
            risk_score += 1
        if self.input_data.get('f5_5') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f5_all(self):
        f5_1 = self.risk_calculation_f5_1()
        f5_2 = self.risk_calculation_f5_2()
        f5_3 = self.risk_calculation_f5_3()
        f5_4 = self.risk_calculation_f5_4()
        f5_5 = self.risk_calculation_f5_5()

        risk_score = f5_1[0] + f5_2[0] + f5_3[0] + f5_4[0] + f5_5[0]
        number_of_risks = f5_1[1] + f5_2[1] + f5_3[1] + f5_4[1] + f5_5[1]
        total_risk = [risk_score, number_of_risks]
        return total_risk

    # Risk Summary Form 6
    def risk_calculation_f6_1(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f6_1') == '0':
            risk_score += 0
        elif self.input_data.get('f6_1') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f6_2(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f6_2') == '0':
            risk_score += 0
        elif self.input_data.get('f6_2') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f6_3(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f3_3') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f3_3') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f3_3') == '3':
            risk_score += 1
        if self.input_data.get('f3_3') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f6_4(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f6_4') == '0':
            risk_score += 0
        elif self.input_data.get('f6_4') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f6_5(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f6_5') == '0':
            risk_score += 0
        elif self.input_data.get('f6_5') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f6_all(self):
        f6_1 = self.risk_calculation_f6_1()
        f6_2 = self.risk_calculation_f6_2()
        f6_3 = self.risk_calculation_f6_3()
        f6_4 = self.risk_calculation_f6_4()
        f6_5 = self.risk_calculation_f6_5()
        risk_score = f6_1[0] + f6_2[0] + f6_3[0] + f6_4[0] + f6_5[0]
        number_of_risks = f6_1[1] + f6_2[1] + f6_3[1] + f6_4[1] + f6_5[1]
        total_risk = [risk_score, number_of_risks]
        return total_risk

    # Risk Summary Form 7
    def risk_calculation_f7_1(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f7_1') == '0':
            risk_score = 0
            number_of_risks = 0
        elif self.input_data.get('f7_1') == '1':
            risk_score += 1
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f7_2(self):
        risk_score = 0
        number_of_risks = 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f7_all(self):
        f7_1 = self.risk_calculation_f7_1()
        f7_2 = self.risk_calculation_f7_2()
        risk_score = f7_1[0] + f7_2[0]
        number_of_risks = f7_1[1] + f7_2[1]
        total_risk = [risk_score, number_of_risks]
        return total_risk

    # Risk Summary Form 8
    def risk_calculation_f8_1(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f8_1') == '0':
            risk_score += 0

        elif self.input_data.get('f8_1') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f8_2(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f8_2') == '0':
            risk_score += 0
            number_of_risks += 1
        elif self.input_data.get('f8_2') == '1':
            risk_score += 3
            number_of_risks += 1
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f8_3(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f8_3') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f8_3') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f8_3') == '3':
            risk_score += 1
        if self.input_data.get('f8_3') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f8_4(self):
        risk_score = 0
        number_of_risks = 0
        if self.input_data.get('f8_4') == '1':
            risk_score += 3
            number_of_risks += 1
        elif self.input_data.get('f8_4') == '2':
            risk_score += 2
            number_of_risks += 1
        elif self.input_data.get('f8_4') == '3':
            risk_score += 1
        if self.input_data.get('f8_4') == '4':
            risk_score += 0
        total_risk = [risk_score, number_of_risks]
        return total_risk

    def risk_calculation_f8_all(self):
        f8_1 = self.risk_calculation_f8_1()
        f8_2 = self.risk_calculation_f8_2()
        f8_3 = self.risk_calculation_f8_3()
        f8_4 = self.risk_calculation_f8_4()
        risk_score = f8_1[0] + f8_2[0] + f8_3[0] + f8_4[0]
        number_of_risks = f8_1[1] + f8_2[1] + f8_3[1] + f8_4[1]
        total_risk = [risk_score, number_of_risks]
        return total_risk


def dpia_screening(request):
    if request.method == 'GET':
        session_dict = get_session_data()
        count = 1
        context = {}
        session_data = False
        for i in session_dict:
            if request.session['title'] == session_dict.get('session_dict_{}'.format(count)).get('title') \
                    and request.session['date'] == session_dict.get('session_dict_{}'.format(count)).get('date'):
                context['value_dict'] = session_dict.get('session_dict_{}'.format(count))
                context['approve_or_reject'] = request.session.get('approve_or_reject')
                session_data = True
            count += 1
        if not session_data:
            context['value_dict'] = 0
        context['session_dict'] = session_dict
        return render(request, 'dpia_screening.html', context)
    if request.method == 'POST':
        request.session['f1_1'] = request.POST.get("f1_1")
        request.session['f1_2'] = request.POST.get("f1_2")
        request.session['f1_3'] = request.POST.get("f1_3")
        request.session['f1_4'] = request.POST.get("f1_4")
        request.session['f1_5'] = request.POST.get("f1_5")
        request.session['f1_6'] = request.POST.get("f1_6")
        f1_1 = request.session.get('f1_1')
        f1_2 = request.session.get('f1_2')
        f1_3 = request.session.get('f1_3')
        f1_4 = request.session.get('f1_4')
        f1_5 = request.session.get('f1_5')
        f1_6 = request.session.get('f1_6')

        request.session['f2_1'] = request.POST.get("f2_1")
        request.session['f2_2'] = request.POST.get("f2_2")
        request.session['f2_3'] = request.POST.get("f2_3")
        request.session['f2_4'] = request.POST.get("f2_4")
        request.session['f2_5'] = request.POST.get("f2_5")
        request.session['f2_6'] = request.POST.get("f2_6")
        request.session['f2_7'] = request.POST.get("f2_7")
        request.session['f2_8'] = request.POST.get("f2_8")
        request.session['f2_9'] = request.POST.get("f2_9")
        f2_1 = request.session.get('f2_1')
        f2_2 = request.session.get('f2_2')
        f2_3 = request.session.get('f2_3')
        f2_4 = request.session.get('f2_4')
        f2_5 = request.session.get('f2_5')
        f2_6 = request.session.get('f2_6')
        f2_7 = request.session.get('f2_7')
        f2_8 = request.session.get('f2_8')
        f2_9 = request.session.get('f2_9')

        request.session['f3_1'] = request.POST.get("f3_1")
        request.session['f3_2'] = request.POST.get("f3_2")
        request.session['f3_3'] = request.POST.get("f3_3")
        request.session['f3_4'] = request.POST.get("f3_4")
        request.session['f3_5'] = request.POST.get("f3_5")
        request.session['f3_6'] = request.POST.get("f3_6")
        request.session['f3_7'] = request.POST.get("f3_7")
        request.session['f3_8'] = request.POST.get("f3_8")
        request.session['f3_9'] = request.POST.get("f3_9")
        request.session['f3_10'] = request.POST.get("f3_10")
        request.session['f3_11'] = request.POST.get("f3_11")
        request.session['f3_12'] = request.POST.get("f3_12")
        f3_1 = request.session.get('f3_1')
        f3_2 = request.session.get('f3_2')
        f3_3 = request.session.get('f3_3')
        f3_4 = request.session.get('f3_4')
        f3_5 = request.session.get('f3_5')
        f3_6 = request.session.get('f3_6')
        f3_7 = request.session.get('f3_7')
        f3_8 = request.session.get('f3_8')
        f3_9 = request.session.get('f3_9')
        f3_10 = request.session.get('f3_10')
        f3_11 = request.session.get('f3_11')
        f3_12 = request.session.get('f3_12')

        request.session['f4_1'] = request.POST.get("f4_1")
        request.session['f4_2'] = request.POST.get("f4_2")
        request.session['f4_3'] = request.POST.get("f4_3")
        request.session['f4_4'] = request.POST.get("f4_4")
        request.session['f4_5'] = request.POST.get("f4_5")
        request.session['f4_6'] = request.POST.get("f4_6")
        request.session['f4_7'] = request.POST.get("f4_7")
        request.session['f4_8'] = request.POST.get("f4_8")
        request.session['f4_9'] = request.POST.get("f4_9")
        request.session['f4_10'] = request.POST.get("f4_10")
        request.session['f4_11'] = request.POST.get("f4_11")
        request.session['f4_12'] = request.POST.get("f4_12")
        request.session['f4_13'] = request.POST.get("f4_13")
        f4_1 = request.session.get('f4_1')
        f4_2 = request.session.get('f4_2')
        f4_3 = request.session.get('f4_3')
        f4_4 = request.session.get('f4_4')
        f4_5 = request.session.get('f4_5')
        f4_6 = request.session.get('f4_6')
        f4_7 = request.session.get('f4_7')
        f4_8 = request.session.get('f4_8')
        f4_9 = request.session.get('f4_9')
        f4_10 = request.session.get('f4_10')
        f4_11 = request.session.get('f4_11')
        f4_12 = request.session.get('f4_12')
        f4_13 = request.session.get('f4_13')

        request.session['f5_1'] = request.POST.get("f5_1")
        request.session['f5_2'] = request.POST.get("f5_2")
        request.session['f5_3'] = request.POST.get("f5_3")
        request.session['f5_4'] = request.POST.get("f5_4")
        request.session['f5_5'] = request.POST.get("f5_5")
        f5_1 = request.session.get('f5_1')
        f5_2 = request.session.get('f5_2')
        f5_3 = request.session.get('f5_3')
        f5_4 = request.session.get('f5_4')
        f5_5 = request.session.get('f5_5')

        request.session['f6_1'] = request.POST.get("f6_1")
        request.session['f6_2'] = request.POST.get("f6_2")
        request.session['f6_3'] = request.POST.get("f6_3")
        request.session['f6_4'] = request.POST.get("f6_4")
        request.session['f6_5'] = request.POST.get("f6_5")
        f6_1 = request.session.get('f6_1')
        f6_2 = request.session.get('f6_2')
        f6_3 = request.session.get('f6_3')
        f6_4 = request.session.get('f6_4')
        f6_5 = request.session.get('f6_5')

        request.session['f7_1'] = request.POST.get("f7_1")
        request.session['f7_2'] = request.POST.get("f7_2")
        f7_1 = request.session.get('f7_1')
        f7_2 = request.session.get('f7_2')

        request.session['f8_1'] = request.POST.get("f8_1")
        request.session['f8_2'] = request.POST.get("f8_2")
        request.session['f8_3'] = request.POST.get("f8_3")
        request.session['f8_4'] = request.POST.get("f8_4")
        f8_1 = request.session.get('f8_1')
        f8_2 = request.session.get('f8_2')
        f8_3 = request.session.get('f8_3')
        f8_4 = request.session.get('f8_4')

        input_data = {
            'f1_1': f1_1,
            'f1_2': f1_2,
            'f1_3': f1_3,
            'f1_4': f1_4,
            'f1_5': f1_5,
            'f1_6': f1_6,

            'f2_1': f2_1,
            'f2_2': f2_2,
            'f2_3': f2_3,
            'f2_4': f2_4,
            'f2_5': f2_5,
            'f2_6': f2_6,
            'f2_7': f2_7,
            'f2_8': f2_8,
            'f2_9': f2_9,

            'f3_1': f3_1,
            'f3_2': f3_2,
            'f3_3': f3_3,
            'f3_4': f3_4,
            'f3_5': f3_5,
            'f3_6': f3_6,
            'f3_7': f3_7,
            'f3_8': f3_8,
            'f3_9': f3_9,
            'f3_10': f3_10,
            'f3_11': f3_11,
            'f3_12': f3_12,

            'f4_1': f4_1,
            'f4_2': f4_2,
            'f4_3': f4_3,
            'f4_4': f4_4,
            'f4_5': f4_5,
            'f4_6': f4_6,
            'f4_7': f4_7,
            'f4_8': f4_8,
            'f4_9': f4_9,
            'f4_10': f4_10,
            'f4_11': f4_11,
            'f4_12': f4_12,
            'f4_13': f4_13,

            'f5_1': f5_1,
            'f5_2': f5_2,
            'f5_3': f5_3,
            'f5_4': f5_4,
            'f5_5': f5_5,

            'f6_1': f6_1,
            'f6_2': f6_2,
            'f6_3': f6_3,
            'f6_4': f6_4,
            'f6_5': f6_5,

            'f7_1': f7_1,
            'f7_2': f7_2,

            'f8_1': f8_1,
            'f8_2': f8_2,
            'f8_3': f8_3,
            'f8_4': f8_4,
        }
        db_objects = Master.objects.all()
        db_dict_list = []
        session_title = request.session.get('title')
        session_status = request.session.get('status')
        session_date = request.session.get('date')
        for instance in db_objects:
            db_dict_list.append(instance.__dict__)
        for item in db_dict_list:
            if item['title'] == session_title and item['status'] == session_status \
                    and item['date'] == session_date or item['title'] == session_title \
                    and item['status'] == session_status or item['title'] == session_title \
                    and item['date'] == session_date:
                record = Master.objects.filter(title=session_title)
                record.delete()

        assessment_db_data = Master(username=request.session.get('username'),
                                    title=request.session.get('title'),
                        author=request.session.get('author'),
                        department=request.session.get('department'),
                        role=request.session.get('role'),
                        manager=request.session.get('manager'),
                        status=request.session.get('status'),
                                    date=request.session.get('date'),
                        name_of_organization=request.session.get('name_of_organization'),
                                    industry=request.session.get('industry'),
                        scope_of_service_project=request.session.get('scope_of_service_project'),
                        data_protection_officer=request.session.get('data_protection_officer'),
                        name_of_DPO=request.session.get('name_of_DPO'),
                                    title_of_DPO=request.session.get('title_of_DPO'),
                        data_processing_project=request.session.get('data_processing_project'),
                        select_data_process=request.session.get('select_data_process'),
                        processing_data=request.session.get('processing_data'),
                        data_processing_involve=request.session.get('data_processing_involve'),
                        automated_decision_making=request.session.get('automated_decision_making'),
                        systematic_monitoring=request.session.get('systematic_monitoring'),
                        process_data_on_large_scale=request.session.get('process_data_on_large_scale'),
                        data_processing_involve_reusing_old_dataset=request.session.get('data_processing_involve_reusing_old_dataset'),
                        vulnerable_data_subjects=request.session.get('vulnerable_data_subjects'),
                        data_processing_involve_innovative_technologies=request.session.get('data_processing_involve_innovative_technologies'),
                        data_processing_involve_sharing_data_outside_european_union=request.session.get('data_processing_involve_sharing_data_outside_european_union'),
                        data_processing_involve_collection_personal_information=request.session.get('data_processing_involve_collection_personal_information'),
                        data_processing_involve_third_party=request.session.get('data_processing_involve_third_party'),
                        data_processing_involve_change_information_is_stored_secured=request.session.get('data_processing_involve_change_information_is_stored_secured'),
                        data_procc_involve_chg_personal_data_currently_collected=request.session.get('data_procc_involve_chg_personal_data_currently_collected'),
                        conducted_DPIA_for_similar_scope_of_service=request.session.get('conducted_DPIA_for_similar_scope_of_service'),
                                    f1_1=f1_1, f1_2=f1_2, f1_3=f1_3, f1_4=f1_4, f1_5=f1_5, f1_6=f1_6,
                                    f2_1=f2_1, f2_2=f2_2, f2_3=f2_3, f2_4=f2_4, f2_5=f2_5, f2_6=f2_6, f2_7=f2_7, f2_8=f2_8, f2_9=f2_9,
                                    f3_1=f3_1, f3_2=f3_2, f3_3=f3_3, f3_4=f3_4, f3_5=f3_5, f3_6=f3_6, f3_7=f3_7, f3_8=f3_8, f3_9=f3_9, f3_10=f3_10, f3_11=f3_11, f3_12=f3_12,
                                    f4_1=f4_1, f4_2=f4_2, f4_3=f4_3, f4_4=f4_4, f4_5=f4_5, f4_6=f4_6, f4_7=f4_7, f4_8=f4_8, f4_9=f4_9, f4_10=f4_10, f4_11=f4_11, f4_12=f4_12, f4_13=f4_13,
                                    f5_1=f5_1, f5_2=f5_2, f5_3=f5_3, f5_4=f5_4, f5_5=f5_5,
                                    f6_1=f6_1, f6_2=f6_2, f6_3=f6_3, f6_4=f6_4, f6_5=f6_5,
                                    f7_1=f7_1, f7_2=f7_2,
                                    f8_1=f8_1, f8_2=f8_2, f8_3=f8_3, f8_4=f8_4
                                    )
        assessment_db_data.save()

        table = RiskCalculations(input_data, request)
        form1_percentage = round((table.risk_calculation_f1_all()[1] / table.form1) * 100, 2)
        request.session['form1_percentage'] = form1_percentage
        form2_percentage = round((table.risk_calculation_f2_all()[1] / table.form2) * 100, 2)
        request.session['form2_percentage'] = form2_percentage
        form3_percentage = round((table.risk_calculation_f3_all()[1] / table.form3) * 100, 2)
        request.session['form3_percentage'] = form3_percentage
        form4_percentage = round((table.risk_calculation_f4_all()[1] / table.form4) * 100, 2)
        request.session['form4_percentage'] = form4_percentage
        form5_percentage = round((table.risk_calculation_f5_all()[1] / table.form5) * 100, 2)
        request.session['form5_percentage'] = form5_percentage
        form6_percentage = round((table.risk_calculation_f6_all()[1] / table.form6) * 100, 2)
        request.session['form6_percentage'] = form6_percentage
        form7_percentage = round((table.risk_calculation_f7_all()[1] / table.form7) * 200, 2)
        request.session['form7_percentage'] = form7_percentage
        form8_percentage = round((table.risk_calculation_f8_all()[1] / table.form8) * 100, 2)
        request.session['form8_percentage'] = form8_percentage

        request.session['risk_score1'] = table.risk_calculation_f1_all()
        request.session['risk_score2'] = table.risk_calculation_f2_all()
        request.session['risk_score3'] = table.risk_calculation_f3_all()
        request.session['risk_score4'] = table.risk_calculation_f4_all()
        request.session['risk_score5'] = table.risk_calculation_f5_all()
        request.session['risk_score6'] = table.risk_calculation_f6_all()
        request.session['risk_score7'] = table.risk_calculation_f7_all()
        request.session['risk_score8'] = table.risk_calculation_f8_all()
        request.session['total_no_of_risk'] = table.risk_calculation_f1_all()[1] + table.risk_calculation_f2_all()[1] + \
                                              table.risk_calculation_f3_all()[1] + table.risk_calculation_f4_all()[1] + \
                                              table.risk_calculation_f5_all()[1] + table.risk_calculation_f6_all()[1] + \
                                              table.risk_calculation_f7_all()[1] + table.risk_calculation_f8_all()[1]

        session_dict = get_session_data()
        count = 1
        context = {}
        session_data = False
        for i in session_dict:
            if request.session['title'] == session_dict.get('session_dict_{}'.format(count)).get('title') \
                    and request.session['date'] == session_dict.get('session_dict_{}'.format(count)).get('date'):
                context['value_dict'] = session_dict.get('session_dict_{}'.format(count))
                request.session['value_dict_temp'] = session_dict.get('session_dict_{}'.format(count))
                context['approve_or_reject'] = request.session.get('approve_or_reject')
                session_data = True
            count += 1
        if not session_data:
            context['value_dict'] = 0
        context = {
            'input_data': input_data,
            'value_dict': request.session.get('value_dict_temp'),
            'name_of_controller': request.session.get('manager'),
            'title_of_dpo': request.session.get('title_of_DPO'),
            'name_of_dpo': request.session.get('name_of_DPO'),
            'risk_score1': request.session.get('risk_score1'),
            'risk_score2': request.session.get('risk_score2'),
            'risk_score3': request.session.get('risk_score3'),
            'risk_score4': request.session.get('risk_score4'),
            'risk_score5': request.session.get('risk_score5'),
            'risk_score6': request.session.get('risk_score6'),
            'risk_score7': request.session.get('risk_score7'),
            'risk_score8': request.session.get('risk_score8'),
            'form1_percentage': form1_percentage,
            'form2_percentage': form2_percentage,
            'form3_percentage': form3_percentage,
            'form4_percentage': form4_percentage,
            'form5_percentage': form5_percentage,
            'form6_percentage': form6_percentage,
            'form7_percentage': form7_percentage,
            'form8_percentage': form8_percentage,
            'total_no_of_risk': request.session.get('total_no_of_risk'),
            'session_dict': session_dict,
            'proceed_status': '1'
        }
        return render(request, 'dpia_screening.html', context)
    else:
        return render(request, 'dpia_screening.html')


def click_risk_summary(request):
    session_dict = get_session_data()
    count = 1
    context = {}
    session_data = False
    for i in session_dict:
        if request.session['title'] == session_dict.get('session_dict_{}'.format(count)).get('title') \
                and request.session['date'] == session_dict.get('session_dict_{}'.format(count)).get('date'):
            context['value_dict'] = session_dict.get('session_dict_{}'.format(count))
            request.session['value_dict_temp'] = session_dict.get('session_dict_{}'.format(count))
            context['approve_or_reject'] = request.session.get('approve_or_reject')
            session_data = True
        count += 1
    if not session_data:
        context['value_dict'] = 0
    context = {
        'value_dict': request.session.get('value_dict_temp'),
        'name_of_controller': request.session.get('manager'),
        'title_of_dpo': request.session.get('title_of_DPO'),
        'name_of_dpo': request.session.get('name_of_DPO'),
        'risk_score1': request.session.get('risk_score1'),
        'risk_score2': request.session.get('risk_score2'),
        'risk_score3': request.session.get('risk_score3'),
        'risk_score4': request.session.get('risk_score4'),
        'risk_score5': request.session.get('risk_score5'),
        'risk_score6': request.session.get('risk_score6'),
        'risk_score7': request.session.get('risk_score7'),
        'risk_score8': request.session.get('risk_score8'),
        'form1_percentage': request.session.get('form1_percentage'),
        'form2_percentage': request.session.get('form2_percentage'),
        'form3_percentage': request.session.get('form3_percentage'),
        'form4_percentage': request.session.get('form4_percentage'),
        'form5_percentage': request.session.get('form5_percentage'),
        'form6_percentage': request.session.get('form6_percentage'),
        'form7_percentage': request.session.get('form7_percentage'),
        'form8_percentage': request.session.get('form8_percentage'),
        'total_no_of_risk': request.session.get('total_no_of_risk'),
        'session_dict': session_dict,
        'proceed_status': True
    }
    return render(request, 'risk_summary.html', context)

    
    
class PDF(FPDF):
    pass
    # nothing happens when it is executed.


def get_pdf(request):

    # Page 1 ===========================================

    name_of_org = request.session.get('name_of_organization')
    project_name = request.session.get('title')
    date = request.session.get('date')
    name_of_dpo = request.session.get('name_of_DPO')
    name_of_data_owner = request.session.get('title_of_DPO')

    f1_1 = request.session.get('f1_1')
    f1_2 = request.session.get('f1_2')
    f1_3 = request.session.get('f1_3')
    f1_4 = request.session.get('f1_4')
    f1_5 = request.session.get('f1_5')
    f1_6 = request.session.get('f1_6')

    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=letter)
    can.setFont("Helvetica", 10)

    can.drawString(370, 593, name_of_org)
    can.drawString(370, 579, project_name)
    can.drawString(370, 565, date)
    can.drawString(370, 552, name_of_dpo)
    can.drawString(370, 538,  name_of_data_owner)

    # can.drawCenterdString(10, 50, 'text')
    flow_obj = []
    styles = getSampleStyleSheet()
    text1 = request.session.get('f1_final_text')
    p_text = Paragraph(text1, style=styles['Normal'])
    flow_obj.append(p_text)
    frame = Frame(85, 190, 448, 220)
    frame.addFromList(flow_obj, can)

    can.save()

    new_pdf = PdfFileReader(packet)
    # read your existing PDF
    existing_pdf = PdfFileReader("media/Final_GDPR_report.pdf")
    output = PdfFileWriter()

    page = existing_pdf.getPage(0)
    page.mergePage(new_pdf.getPage(0))
    output.addPage(page)
    # finally, write "output" to a real file
    outputStream = open("pdf/page1.pdf", "wb")
    output.write(outputStream)
    outputStream.close()


# Page 2 ===================================================

    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=letter)
    can.setFontSize(10)

    flow_obj2_1 = []
    flow_obj2_2_1 = []
    styles2 = getSampleStyleSheet()
    text1 = request.session.get('box2_section2_text1')
    text2 = request.session.get('box2_section2_text2')
    p_text = Paragraph(text1, style=styles2['Normal'])
    p2_text = Paragraph(text2, style=styles2['Normal'])
    flow_obj2_1.append(p_text)
    flow_obj2_2_1.append(p2_text)

    frame = Frame(85, 430, 448, 180)
    frame.addFromList(flow_obj2_1, can)

    frame = Frame(85, 150, 448, 180)
    frame.addFromList(flow_obj2_2_1, can)
    can.save()

    new_pdf2 = PdfFileReader(packet)
    # read your existing PDF
    existing_pdf = PdfFileReader("media/Final_GDPR_report.pdf")
    output2 = PdfFileWriter()

    page = existing_pdf.getPage(1)
    page.mergePage(new_pdf2.getPage(0))
    output2.addPage(page)
    # finally, write "output" to a real file
    outputStream = open("pdf/page2.pdf", "wb")
    output2.write(outputStream)
    outputStream.close()

    # Page 3 ===================================================

    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=letter)
    can.setFontSize(10)

    flow_obj3_1 = []
    flow_obj3_2_1 = []
    styles3 = getSampleStyleSheet()
    text3_1 = request.session.get('box2_section2_text3')
    text3_2 = request.session.get('text_box_2_4')
    p3_1_text = Paragraph(text3_1, style=styles3['Normal'])
    p3_2_text = Paragraph(text3_2, style=styles3['Normal'])
    flow_obj3_1.append(p3_1_text)
    flow_obj3_2_1.append(p3_2_text)

    frame = Frame(85, 420, 448, 180)
    frame.addFromList(flow_obj3_1, can)

    frame = Frame(85, 100, 448, 180)
    frame.addFromList(flow_obj3_2_1, can)

    can.save()

    new_pdf3 = PdfFileReader(packet)
    # read your existing PDF
    existing_pdf = PdfFileReader("media/Final_GDPR_report.pdf")
    output3 = PdfFileWriter()

    page = existing_pdf.getPage(2)
    page.mergePage(new_pdf3.getPage(0))
    output3.addPage(page)
    # finally, write "output" to a real file
    outputStream = open("pdf/page3.pdf", "wb")
    output3.write(outputStream)
    outputStream.close()

    # Page 4 ===================================================

    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=letter)
    can.setFontSize(10)

    # can.drawString(370, 577, "Digital Finance")
    # can.drawString(370, 562, "14 Jan 2022")
    # can.drawString(370, 549, "Kumar Das")

    f4_1 = request.session.get('f4_1')
    f4_2 = request.session.get('f4_2')
    f4_3 = request.session.get('f4_3')
    f4_4 = request.session.get('f4_4')
    f4_5 = request.session.get('f4_5')
    f4_6 = request.session.get('f4_6')
    f4_7 = request.session.get('f4_7')
    f4_8 = request.session.get('f4_8')
    f4_9 = request.session.get('f4_9')
    f4_10 = request.session.get('f4_10')
    f4_11 = request.session.get('f4_11')
    f4_12 = request.session.get('f4_12')
    f4_13 = request.session.get('f4_13')

    if f4_1 == '1':
        f4_1_text = 'The legal basis for processing are - *** Select all from the list ; separate by comma if more than 1 ***'
    else:
        f4_1_text = 'The legal basis for processing are - *** Select all from the list ; separate by comma if more than 1 ***'

    if f4_2 == '1':
        f4_2_text = ' For this scope of service , we have not defined the legality of data processing.'
    elif f4_2 == '2':
        f4_2_text = ' For this scope of service , our organization recommends project lead to define the legality of data processing.'
    elif f4_2 == '3':
        f4_2_text = ' For this scope of service , it is important for the project lead to define the legality of data processing.'
    else:
        f4_2_text = ' For this scope of service , our organization recommends project lead to define the legality of data processing.'

    if f4_3 == '0':
        f4_3_text = ' Yes, data processing will achieve the desired goals.'
    else:
        f4_3_text = ' No, The data processing will not achieve the desired goals.'

    if f4_4 == '0':
        f4_4_text = ' Yes, there are alternate approaches to achieve the same outcome.'
    else:
        f4_4_text = 'No, we do not have any alternate approach to achieve the same outcome.'

    if f4_5 == '1':
        f4_5_text = ' We have not have any process to explore the  alternate approaches which can be used to achieve the same results.'
    elif f4_5 == '2':
        f4_5_text = ' We explore a couple of more approaches to achieve the same results. All these approaches are documented in the final report.'
    elif f4_5 == '3':
        f4_5_text = ' We explore all the possible approaches which can achieve the same results and measure their intrusive , post which project team shares it with lead to finalize the final approach to process data.'
    else:
        f4_5_text = ' We explore all the possible approaches which can achieve the same results and measure their intrusive , post which project team check the feasibility before finalizing the final approach to process data.'

    if f4_6 == '1':
        f4_6_text = " Our organization does not currently have any KPI's to monitor data quality and integrity for this specific scope of service."
    elif f4_6 == '2':
        f4_6_text = " We have a list of KPI's in our organization which is used to monitor data quality and integrity for all data related services."
    elif f4_6 == '3':
        f4_6_text = " Project lead decided the list of  KPI's which will be used to monitor data quality and integrity for this specific scope of service."
    else:
        f4_6_text = " Our organization has tailored a list of KPI's which will be used to monitor data quality and integrity for this specific scope of service. This list is prepared by the team and reviewed by the lead."

    if f4_7 == '1':
        f4_7_text = " Data Subjects are not provided any information after collection of the data."
    elif f4_7 == '2':
        f4_7_text = " Data Subjects are informed about the intent of processing data. "
    elif f4_7 == '3':
        f4_7_text = " Data Subjects are informed about the intent of processing data and also notifed if there is any change in the scope of service. "
    else:
        f4_7_text = " Data Subjects are informed about the intent of processing data and also notify if there is any change in the scope of service. Also , they will be notified once the processing is done and data can be archived."

    if f4_8 == '1':
        f4_8_text = " Currently , we do not have process to uphold data subject rights while processing data."
    elif f4_8 == '2':
        f4_8_text = " Our processes related  to data subject rights recommends the data processing team  to uphold these rights and aim for complete compliance."
    elif f4_8 == '3':
        f4_8_text = " We make it essential for the team to upload all the Data subjects rights and make sure that the team comply to it throughout the processing."
    else:
        f4_8_text = " Data subjects rights are given utmost importance while designing data processing plan. We make sure that Data Subject rights are upheld throughout the processing."

    if f4_9 == '1':
        f4_9_text = " At present , we do not have a process to monitor the compliance of the designated entitites while data processing."
    elif f4_9 == '2':
        f4_9_text = " We have a generic process to monitor the compliance of the designated processing entitities. This monitoring process considers a few of the KPI's to measure the compliance."
    elif f4_9 == '3':
        f4_9_text = " We have a robust process to monitor the compliance of the designated processing entitie which aims to achieve full compliance."
    else:
        f4_9_text = "Compliance of the designated processing entites is a must and aim is always to achieve 100% compliance. We have a robust process to monitor the compliance of the designated processing entities."

    if f4_10 == '1':
        f4_10_text = " We do not have a process to ensure any out of scope data processing activities."
    elif f4_10 == '2':
        f4_10_text = " We have a process which ensures that we do not deviate from the decided scope of service while processing data."
    elif f4_10 == '3':
        f4_10_text = " We make sure that we stick to the decided scope of service and have no deviations from it while processing. We do not accept any out of scope requirements and prevent any function creep."
    else:
        f4_10_text = " We make sure that we stick to the decided scope of service and have no deviations from it while processing. We document all the ad-hoc requirements and plan it out for second phase rather than adding it to on-going processing with aim to prevent function creep."

    if f4_11 == '1':
        f4_11_text = 'We do not have a process to ensure any out of scope data processing activities.'
    elif f4_11 == '2':
        f4_11_text = 'We have a process which ensures that we do not deviate from the decided scope of service while processing data.'
    elif f4_11 == '3':
        f4_11_text = 'We make sure that we stick to the decided scope of service and have no deviations from it while processing. We do not accept any out of scope requirements and prevent any function creep.'
    else:
        f4_11_text = 'We make sure that we stick to the decided scope of service and have no deviations from it while processing. We document all the ad-hoc requirements and plan it out for second phase rather than adding it to on-going processing with aim to prevent function creep.'

    if f4_12 == '1':
        f4_12_text = 'For this scope of service , we do not have any measures to ensure that processors comply with the scope of project.'
    elif f4_12 == '2':
        f4_12_text = 'Data processors are informed about the boundaries of the scope and asked to esnure compliance while processing data.'
    elif f4_12 == '3':
        f4_12_text = 'We have well defined controls in place to ensure that data processors are aware about the scope of service and they comply by them.'
    else:
        f4_12_text = 'We have well defined controls in place to ensure that data processors are aware about the scope of service and they comply by them. Processors are aslo asked to prvoide full disclosure about their prcoessing activites and ensure that they document all the data transfers.'

    if f4_13 == '1':
        f4_13_text = 'There is no measures defined to safegaurd any international transfer of the data.'
    elif f4_13 == '2':
        f4_13_text = 'We have a few controls to keep the data safe in the designated geography and prevent them from any international transfer.'
    elif   f4_13 == '3':
        f4_13_text = 'We keep a strict check on any international transfer of the data and have several safegaurds aiming to do same. We have efficient controls to keep the data safe in the designated geography.'
    else:
        f4_13_text = 'We keep a strict check on any international transfer of the data and have several safegaurds aiming to do same. We also make sure that all the data controllers and processors who have access to our data also have efficient controls to keep the data safe in the designated geography.'



    flow_obj4 = []
    styles4 = getSampleStyleSheet()
    final_f4_text = f4_1_text + f4_2_text + f4_3_text + f4_4_text + f4_5_text + f4_6_text + f4_7_text + f4_8_text + f4_9_text + f4_10_text + f4_11_text + f4_12_text + f4_13_text
    text4 = request.session.get('f4_final_text')
    p4_text = Paragraph(text4, style=styles4['Normal'])
    flow_obj4.append(p4_text)
    frame4 = Frame(85, 26, 448, 240)
    frame4.addFromList(flow_obj4, can)

    flow_obj4_2 = []
    text5 = request.session.get('f3_final_text')
    p4_text = Paragraph(text5, style=styles4['Normal'])
    flow_obj4_2.append(p4_text)
    frame4 = Frame(85, 440, 448, 170)
    frame4.addFromList(flow_obj4_2, can)

    can.save()

    new_pdf4 = PdfFileReader(packet)
    # read your existing PDF
    existing_pdf = PdfFileReader("media/Final_GDPR_report.pdf")
    output4 = PdfFileWriter()

    page = existing_pdf.getPage(3)
    page.mergePage(new_pdf4.getPage(0))
    output4.addPage(page)
    # finally, write "output" to a real file
    outputStream = open("pdf/page4.pdf", "wb")
    output4.write(outputStream)
    outputStream.close()

    # Page 5 ===================================================

    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=letter)
    can.setFontSize(10)
    styles5 = getSampleStyleSheet()
    final_text_form5_f1 = []
    final_text_form5_f2 = []
    final_text_form5_f3 = []
    final_text_form5_f4 = []
    final_text_form5_f5 = []
    final_text_form5_f6 = []

    # Section 5 Question 1
    # Text form5_1
    pdf_print_dict_section5 = request.session.get('pdf_print_dict_section5', None)
    if pdf_print_dict_section5:
        if pdf_print_dict_section5.get('s5_b1').get('s5_b1_likelihood') is not None:
            form5_f1_text = pdf_print_dict_section5.get('s5_b1').get('f5_1_question')
            form5_f1_text_approved = Paragraph(form5_f1_text, style=styles5['Normal'])
            final_text_form5_f2.append(form5_f1_text_approved)

            frame5 = Frame(85, 450, 240, 55,)
            frame5.addFromList(final_text_form5_f2, can)

            final_text_form5_f2_1 = []
            final_text_form5_f2_2 = []
            final_text_form5_f2_3 = []

            # Text form5_f2_1
            form5_f2_1_text = pdf_print_dict_section5.get('s5_b1').get('s5_b1_likelihood')
            form5_f2_1_text_approved = Paragraph(form5_f2_1_text, style=styles5['Normal'])
            final_text_form5_f2_1.append(form5_f2_1_text_approved)

            frame5 = Frame(356, 460, 60, 45)
            frame5.addFromList(final_text_form5_f2_1, can)

            # Text form5_f2_2
            form5_f2_2_text = pdf_print_dict_section5.get('s5_b1').get('s5_b1_severity')
            form5_f2_2_text_approved = Paragraph(form5_f2_2_text, style=styles5['Normal'])
            final_text_form5_f2_2.append(form5_f2_2_text_approved)

            frame5 = Frame(422, 460, 60, 45)
            frame5.addFromList(final_text_form5_f2_2, can)

            # Text form5_f2_3
            form5_f2_3_text = pdf_print_dict_section5.get('s5_b1').get('s5_b1_residual_risk')
            form5_f2_3_text_approved = Paragraph(form5_f2_3_text, style=styles5['Normal'])
            final_text_form5_f2_3.append(form5_f2_3_text_approved)

            frame5 = Frame(482, 460, 60, 45)
            frame5.addFromList(final_text_form5_f2_3, can)


    # Text form5_3


    if pdf_print_dict_section5:
        if pdf_print_dict_section5.get('s5_b2').get('s5_b2_likelihood') is not None:
            form5_f3_text = pdf_print_dict_section5.get('s5_b2').get('f5_2_question')
            form5_f3_text_approved = Paragraph(form5_f3_text, style=styles5['Normal'])
            final_text_form5_f3.append(form5_f3_text_approved)

            frame5 = Frame(85, 400, 240, 55)
            frame5.addFromList(final_text_form5_f3, can)

            final_text_form5_f3_1 = []
            final_text_form5_f3_2 = []
            final_text_form5_f3_3 = []

            # Text form5_f2_3
            form5_f3_1_text = pdf_print_dict_section5.get('s5_b2').get('s5_b2_likelihood')
            form5_f3_1_text_approved = Paragraph(form5_f3_1_text, style=styles5['Normal'])
            final_text_form5_f3_1.append(form5_f3_1_text_approved)

            frame5 = Frame(356, 410, 60, 45)
            frame5.addFromList(final_text_form5_f3_1, can)

            # Text form5_f3_2
            form5_f3_2_text = pdf_print_dict_section5.get('s5_b2').get('s5_b2_severity')
            form5_f3_2_text_approved = Paragraph(form5_f3_2_text, style=styles5['Normal'])
            final_text_form5_f3_2.append(form5_f3_2_text_approved)

            frame5 = Frame(422, 410, 60, 45)
            frame5.addFromList(final_text_form5_f3_2, can)

            # Text form5_f3_3
            form5_f3_3_text = pdf_print_dict_section5.get('s5_b2').get('s5_b2_residual_risk')
            form5_f3_3_text_approved = Paragraph(form5_f3_3_text, style=styles5['Normal'])
            final_text_form5_f3_3.append(form5_f3_3_text_approved)

            frame5 = Frame(482, 410, 60, 45)
            frame5.addFromList(final_text_form5_f3_3, can)


    # Text form5_4
    if pdf_print_dict_section5:
        if pdf_print_dict_section5.get('s5_b3').get('s5_b3_likelihood') is not None:
            form5_f4_text = pdf_print_dict_section5.get('s5_b3').get('f5_3_question')
            form5_f4_text_approved = Paragraph(form5_f4_text, style=styles5['Normal'])
            final_text_form5_f4.append(form5_f4_text_approved)

            frame5 = Frame(85, 350, 240, 55)
            frame5.addFromList(final_text_form5_f4, can)

            final_text_form5_f4_1 = []
            final_text_form5_f4_2 = []
            final_text_form5_f4_3 = []

            # Text form5_f4_1
            form5_f4_1_text = pdf_print_dict_section5.get('s5_b3').get('s5_b3_likelihood')
            form5_f4_1_text_approved = Paragraph(form5_f4_1_text, style=styles5['Normal'])
            final_text_form5_f4_1.append(form5_f4_1_text_approved)

            frame5 = Frame(356, 360, 60, 45)
            frame5.addFromList(final_text_form5_f4_1, can)

            # Text form5_f4_2
            form5_f4_2_text = pdf_print_dict_section5.get('s5_b3').get('s5_b3_severity')
            form5_f4_2_text_approved = Paragraph(form5_f4_2_text, style=styles5['Normal'])
            final_text_form5_f4_2.append(form5_f4_2_text_approved)

            frame5 = Frame(422, 360, 60, 45)
            frame5.addFromList(final_text_form5_f4_2, can)

            # Text form5_f4_3
            form5_f4_3_text = pdf_print_dict_section5.get('s5_b3').get('s5_b3_residual_risk')
            form5_f4_3_text_approved = Paragraph(form5_f4_3_text, style=styles5['Normal'])
            final_text_form5_f4_3.append(form5_f4_3_text_approved)

            frame5 = Frame(482, 360, 60, 45)
            frame5.addFromList(final_text_form5_f4_3, can)


    # Text form5_5


    if pdf_print_dict_section5:
        if pdf_print_dict_section5.get('s5_b4').get('s5_b4_likelihood') is not None:
            form5_f5_text = pdf_print_dict_section5.get('s5_b4').get('f5_4_question')
            form5_f5_text_approved = Paragraph(form5_f5_text, style=styles5['Normal'])
            final_text_form5_f5.append(form5_f5_text_approved)

            frame5 = Frame(85, 310, 240, 45)
            frame5.addFromList(final_text_form5_f5, can)

            final_text_form5_f5_1 = []
            final_text_form5_f5_2 = []
            final_text_form5_f5_3 = []

            # Text form5_f5_1
            form5_f5_1_text = pdf_print_dict_section5.get('s5_b4').get('s5_b4_likelihood')
            form5_f5_1_text_approved = Paragraph(form5_f5_1_text, style=styles5['Normal'])
            final_text_form5_f5_1.append(form5_f5_1_text_approved)

            frame5 = Frame(356, 310, 60, 45)
            frame5.addFromList(final_text_form5_f5_1, can)

            # Text form5_f5_2
            form5_f5_2_text = pdf_print_dict_section5.get('s5_b4').get('s5_b4_severity')
            form5_f5_2_text_approved = Paragraph(form5_f5_2_text, style=styles5['Normal'])
            final_text_form5_f5_2.append(form5_f5_2_text_approved)

            frame5 = Frame(422, 310, 60, 45)
            frame5.addFromList(final_text_form5_f5_2, can)

            # Text form5_f5_3
            form5_f5_3_text = pdf_print_dict_section5.get('s5_b4').get('s5_b4_residual_risk')
            form5_f5_3_text_approved = Paragraph(form5_f5_3_text, style=styles5['Normal'])
            final_text_form5_f5_3.append(form5_f5_3_text_approved)

            frame5 = Frame(482, 310, 60, 45)
            frame5.addFromList(final_text_form5_f5_3, can)

        # Text form5_6
        if pdf_print_dict_section5:
            if pdf_print_dict_section5.get('s5_b5').get('s5_b5_likelihood') is not None:
                form5_f6_text = pdf_print_dict_section5.get('s5_b5').get('f5_5_question')
                form5_f6_text_approved = Paragraph(form5_f6_text, style=styles5['Normal'])
                final_text_form5_f6.append(form5_f6_text_approved)

                frame5 = Frame(85, 270, 240, 45)
                frame5.addFromList(final_text_form5_f6, can)

                final_text_form5_f6_1 = []
                final_text_form5_f6_2 = []
                final_text_form5_f6_3 = []

                # Text form5_f5_1
                form5_f6_1_text = pdf_print_dict_section5.get('s5_b4').get('s5_b4_likelihood')
                form5_f6_1_text_approved = Paragraph(form5_f6_1_text, style=styles5['Normal'])
                final_text_form5_f6_1.append(form5_f6_1_text_approved)

                frame5 = Frame(356, 270, 60, 45)
                frame5.addFromList(final_text_form5_f6_1, can)

                # Text form5_f5_2
                form5_f6_2_text = pdf_print_dict_section5.get('s5_b4').get('s5_b4_severity')
                form5_f6_2_text_approved = Paragraph(form5_f5_2_text, style=styles5['Normal'])
                final_text_form5_f6_2.append(form5_f6_2_text_approved)

                frame5 = Frame(422, 270, 60, 45)
                frame5.addFromList(final_text_form5_f6_2, can)

                # Text form5_f5_3
                form5_f6_3_text = pdf_print_dict_section5.get('s5_b4').get('s5_b4_residual_risk')
                form5_f6_3_text_approved = Paragraph(form5_f6_3_text, style=styles5['Normal'])
                final_text_form5_f6_3.append(form5_f6_3_text_approved)

                frame5 = Frame(482, 270, 60, 45)
                frame5.addFromList(final_text_form5_f6_3, can)


    can.save()

    new_pdf5 = PdfFileReader(packet)
    # read your existing PDF
    existing_pdf = PdfFileReader("media/Final_GDPR_report.pdf")
    output5 = PdfFileWriter()

    page = existing_pdf.getPage(4)
    page.mergePage(new_pdf5.getPage(0))
    output5.addPage(page)
    # finally, write "output" to a real file
    outputStream = open("pdf/page5.pdf", "wb")
    output5.write(outputStream)
    outputStream.close()

    # Page 6 ===================================================

    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=letter)
    can.setFontSize(10)

    styles6 = getSampleStyleSheet()
    final_text_form6_f1 = []
    final_text_form6_f2 = []
    final_text_form6_f3 = []
    final_text_form6_f4 = []
    final_text_form6_f5 = []

    # Text form6_f1
    if pdf_print_dict_section5:
        if pdf_print_dict_section5.get('s5_b1').get('s5_b1_likelihood') is not None:
            form6_f1_text = '1.'
            form6_f1_text_approved = Paragraph(form6_f1_text, style=styles6['Normal'])
            final_text_form6_f1.append(form6_f1_text_approved)

            frame5 = Frame(82, 520, 60, 35)
            frame5.addFromList(final_text_form6_f1, can)

            form6_f1_1 = []
            form6_f1_2 = []
            form6_f1_3 = []
            form6_f1_4 = []
            form6_f1_5 = []

            # Text form6_f1_1
            form6_f1_1_text = pdf_print_dict_section5.get('s5_b1').get('s5_b1_option_reduce_risk')
            form6_f1_1_text_approved = Paragraph(form6_f1_1_text, style=styles6['Normal'])
            form6_f1_1.append(form6_f1_1_text_approved)

            frame5 = Frame(160, 520, 170, 35)
            frame5.addFromList(form6_f1_1, can)

            # Text form6_f1_2
            form6_f1_2_text = pdf_print_dict_section5.get('s5_b1').get('s5_b1_effect_on_risk')
            form6_f1_2_text_approved = Paragraph(form6_f1_2_text, style=styles6['Normal'])
            form6_f1_2.append(form6_f1_2_text_approved)

            frame5 = Frame(340, 520, 70, 35)
            frame5.addFromList(form6_f1_2, can)

            # Text form6_f1_3
            form6_f1_3_text = pdf_print_dict_section5.get('s5_b1').get('s5_b1_residual_risk')
            form6_f1_3_text_approved = Paragraph(form6_f1_3_text, style=styles6['Normal'])
            form6_f1_3.append(form6_f1_3_text_approved)

            frame5 = Frame(410, 520, 67, 35)
            frame5.addFromList(form6_f1_3, can)

            # Text form6_f1_4
            form6_f1_4_text = pdf_print_dict_section5.get('s5_b1').get('s5_b1_measure_approved')
            form6_f1_4_text_approved = Paragraph(form6_f1_4_text, style=styles6['Normal'])
            form6_f1_4.append(form6_f1_4_text_approved)

            frame5 = Frame(479, 520, 64, 35)
            frame5.addFromList(form6_f1_4, can)


    # Text form6_f2
    if pdf_print_dict_section5:
        if pdf_print_dict_section5.get('s5_b2').get('s5_b2_likelihood') is not None:
            form6_f2_text = '2.'
            form6_f2_text_approved = Paragraph(form6_f2_text, style=styles6['Normal'])
            final_text_form6_f2.append(form6_f2_text_approved)

            frame5 = Frame(82, 480, 60, 35)
            frame5.addFromList(final_text_form6_f2, can)

            form6_f2_1 = []
            form6_f2_2 = []
            form6_f2_3 = []
            form6_f2_4 = []
            form6_f2_5 = []
            # Text form6_f1_1
            form6_f2_1_text = pdf_print_dict_section5.get('s5_b2').get('s5_b2_option_reduce_risk')
            form6_f2_1_text_approved = Paragraph(form6_f2_1_text, style=styles6['Normal'])
            form6_f2_1.append(form6_f2_1_text_approved)

            frame5 = Frame(160, 480, 170, 35)
            frame5.addFromList(form6_f2_1, can)

            # Text form6_f1_2
            form6_f2_2_text = pdf_print_dict_section5.get('s5_b2').get('s5_b2_effect_on_risk')
            form6_f2_2_text_approved = Paragraph(form6_f2_2_text, style=styles6['Normal'])
            form6_f2_2.append(form6_f2_2_text_approved)

            frame5 = Frame(340, 480, 70, 35)
            frame5.addFromList(form6_f2_2, can)

            # Text form6_f1_3
            form6_f2_3_text = pdf_print_dict_section5.get('s5_b2').get('s5_b2_residual_risk')
            form6_f2_3_text_approved = Paragraph(form6_f2_3_text, style=styles6['Normal'])
            form6_f2_3.append(form6_f2_3_text_approved)

            frame5 = Frame(410, 480, 67, 35)
            frame5.addFromList(form6_f2_3, can)

            # Text form6_f1_4
            form6_f2_4_text = pdf_print_dict_section5.get('s5_b2').get('s5_b2_measure_approved')
            form6_f2_4_text_approved = Paragraph(form6_f2_4_text, style=styles6['Normal'])
            form6_f2_4.append(form6_f2_4_text_approved)

            frame5 = Frame(479, 480, 64, 35)
            frame5.addFromList(form6_f2_4, can)


    # Text form6_f3
    if pdf_print_dict_section5:
        if pdf_print_dict_section5.get('s5_b3').get('s5_b3_likelihood') is not None:
            form6_f3_text = '3.'
            form6_f3_text_approved = Paragraph(form6_f3_text, style=styles6['Normal'])
            final_text_form6_f3.append(form6_f3_text_approved)

            frame5 = Frame(82, 440, 60, 35)
            frame5.addFromList(final_text_form6_f3, can)


            form6_f3_1 = []
            form6_f3_2 = []
            form6_f3_3 = []
            form6_f3_4 = []
            form6_f3_5 = []
            # Text form6_f3_1
            form6_f3_1_text = pdf_print_dict_section5.get('s5_b3').get('s5_b3_option_reduce_risk')
            form6_f3_1_text_approved = Paragraph(form6_f3_1_text, style=styles6['Normal'])
            form6_f3_1.append(form6_f3_1_text_approved)

            frame5 = Frame(160, 440, 170, 35)
            frame5.addFromList(form6_f3_1, can)

            # Text form6_f3_2
            form6_f3_2_text = pdf_print_dict_section5.get('s5_b3').get('s5_b3_effect_on_risk')
            form6_f3_2_text_approved = Paragraph(form6_f3_2_text, style=styles6['Normal'])
            form6_f3_2.append(form6_f3_2_text_approved)

            frame5 = Frame(340, 440, 70, 35)
            frame5.addFromList(form6_f3_2, can)

            # Text form6_f3_3
            form6_f3_3_text = pdf_print_dict_section5.get('s5_b3').get('s5_b3_residual_risk')
            form6_f3_3_text_approved = Paragraph(form6_f3_3_text, style=styles6['Normal'])
            form6_f3_3.append(form6_f3_3_text_approved)

            frame5 = Frame(410, 440, 67, 35)
            frame5.addFromList(form6_f3_3, can)

            # Text form6_f3_4
            form6_f3_4_text = pdf_print_dict_section5.get('s5_b3').get('s5_b3_measure_approved')
            form6_f3_4_text_approved = Paragraph(form6_f3_4_text, style=styles6['Normal'])
            form6_f3_4.append(form6_f3_4_text_approved)

            frame5 = Frame(479, 440, 64, 35)
            frame5.addFromList(form6_f3_4, can)


    # Text form6_f4
    if pdf_print_dict_section5:
        if pdf_print_dict_section5.get('s5_b4').get('s5_b4_likelihood') is not None:
            form6_f4_text = '4.'
            form6_f4_text_approved = Paragraph(form6_f4_text, style=styles6['Normal'])
            final_text_form6_f4.append(form6_f4_text_approved)

            frame5 = Frame(82, 400, 60, 35)
            frame5.addFromList(final_text_form6_f4, can)

            form6_f4_1 = []
            form6_f4_2 = []
            form6_f4_3 = []
            form6_f4_4 = []
            form6_f4_5 = []
            # Text form6_f4_1
            form6_f4_1_text = pdf_print_dict_section5.get('s5_b4').get('s5_b4_option_reduce_risk')
            form6_f4_1_text_approved = Paragraph(form6_f4_1_text, style=styles6['Normal'])
            form6_f4_1.append(form6_f4_1_text_approved)

            frame5 = Frame(160, 400, 170, 35)
            frame5.addFromList(form6_f4_1, can)

            # Text form6_f4_2
            form6_f4_2_text = pdf_print_dict_section5.get('s5_b4').get('s5_b4_effect_on_risk')
            form6_f4_2_text_approved = Paragraph(form6_f4_2_text, style=styles6['Normal'])
            form6_f4_2.append(form6_f4_2_text_approved)

            frame5 = Frame(340, 400, 70, 35)
            frame5.addFromList(form6_f4_2, can)

            # Text form6_f4_3
            form6_f4_3_text = pdf_print_dict_section5.get('s5_b4').get('s5_b4_residual_risk')
            form6_f4_3_text_approved = Paragraph(form6_f4_3_text, style=styles6['Normal'])
            form6_f4_3.append(form6_f4_3_text_approved)

            frame5 = Frame(410, 400, 67, 35)
            frame5.addFromList(form6_f4_3, can)

            # Text form6_f4_4
            form6_f4_4_text = pdf_print_dict_section5.get('s5_b4').get('s5_b4_measure_approved')
            form6_f4_4_text_approved = Paragraph(form6_f4_4_text, style=styles6['Normal'])
            form6_f4_4.append(form6_f4_4_text_approved)

            frame5 = Frame(479, 400, 64, 35)
            frame5.addFromList(form6_f4_4, can)


    # Text form6_f5
    if pdf_print_dict_section5:
        if pdf_print_dict_section5.get('s5_b5').get('s5_b5_likelihood') is not None:
            form6_f5_text = '5'
            form6_f5_text_approved = Paragraph(form6_f5_text, style=styles6['Normal'])
            final_text_form6_f5.append(form6_f5_text_approved)

            frame5 = Frame(82, 360, 60, 35)
            frame5.addFromList(final_text_form6_f5, can)

            form6_f5_1 = []
            form6_f5_2 = []
            form6_f5_3 = []
            form6_f5_4 = []
            form6_f5_5 = []
            # Text form6_f5_1
            form6_f5_1_text = pdf_print_dict_section5.get('s5_b5').get('s5_b5_option_reduce_risk')
            form6_f5_1_text_approved = Paragraph(form6_f5_1_text, style=styles6['Normal'])
            form6_f5_1.append(form6_f5_1_text_approved)

            frame5 = Frame(160, 360, 170, 35)
            frame5.addFromList(form6_f5_1, can)

            # Text form6_f5_2
            form6_f5_2_text = pdf_print_dict_section5.get('s5_b5').get('s5_b5_effect_on_risk')
            form6_f5_2_text_approved = Paragraph(form6_f5_2_text, style=styles6['Normal'])
            form6_f5_2.append(form6_f5_2_text_approved)

            frame5 = Frame(340, 360, 70, 35)
            frame5.addFromList(form6_f5_2, can)

            # Text form6_f5_3
            form6_f5_3_text = pdf_print_dict_section5.get('s5_b5').get('s5_b5_residual_risk')
            form6_f5_3_text_approved = Paragraph(form6_f5_3_text, style=styles6['Normal'])
            form6_f5_3.append(form6_f5_3_text_approved)

            frame5 = Frame(410, 360, 67, 35)
            frame5.addFromList(form6_f5_3, can)

            # Text form6_f5_4
            form6_f5_4_text = pdf_print_dict_section5.get('s5_b5').get('s5_b5_measure_approved')
            form6_f5_4_text_approved = Paragraph(form6_f5_4_text, style=styles6['Normal'])
            form6_f5_4.append(form6_f5_4_text_approved)

            frame5 = Frame(479, 360, 64, 35)
            frame5.addFromList(form6_f5_4, can)

    can.save()


    new_pdf6 = PdfFileReader(packet)
    # read your existing PDF
    existing_pdf = PdfFileReader("media/Final_GDPR_report.pdf")
    output6 = PdfFileWriter()

    page = existing_pdf.getPage(5)
    page.mergePage(new_pdf6.getPage(0))
    output6.addPage(page)
    # finally, write "output" to a real file
    outputStream = open("pdf/page6.pdf", "wb")
    output6.write(outputStream)
    outputStream.close()

    # Page 7 ===================================================

    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=letter)
    can.setFontSize(10)

    # can.drawString(370, 577, "Digital Finance")
    # can.drawString(370, 562, "14 Jan 2022")
    # can.drawString(370, 549, "Kumar Das")

    flow_obj7_m_app_by = []
    flow_obj7_m_app_r_by = []
    flow_obj7_m_app_r_by_3 = []
    flow_obj7_m_app_r_by_4 = []
    flow_obj7_m_app_r_by_5 = []
    flow_obj7_m_app_r_by_6 = []
    flow_obj7_m_app_r_by_7 = []
    flow_obj7_m_app_r_by_8 = []
    flow_obj7_m_app_r_by_9 = []
    styles7 = getSampleStyleSheet()
    # p7_text = Paragraph(text7, style=styles4['Normal'])
    # flow_obj4.append(p7_text)


    # Frame 1 text
    text_m_approved = request.session.get('name_of_DPO')
    text_approved = Paragraph(text_m_approved, style=styles7['Normal'])
    flow_obj7_m_app_by.append(text_approved)

    frame4 = Frame(220, 640, 120, 25)
    frame4.addFromList(flow_obj7_m_app_by, can)
    # Frame 2 Text

    text_approved2 = request.session.get('title_of_DPO')
    text_approved_2 = Paragraph(text_approved2, style=styles7['Normal'])
    flow_obj7_m_app_r_by.append(text_approved_2)

    frame5 = Frame(220, 590, 120, 25)
    frame5.addFromList(flow_obj7_m_app_r_by, can)

    # Frame 3 Text

    text_approved3 = request.session.get('title_of_DPO')
    text_approved_3 = Paragraph(text_approved3, style=styles7['Normal'])
    flow_obj7_m_app_r_by_3.append(text_approved_3)

    frame5 = Frame(220, 540, 120, 25)
    frame5.addFromList(flow_obj7_m_app_r_by_3, can)

    # Frame 4 Text

    text_approved4 = request.session.get('summary_dpo_advice', 'DPO Advice')
    text_approved_4 = Paragraph(text_approved4, style=styles7['Normal'])
    flow_obj7_m_app_r_by_4.append(text_approved_4)

    frame5 = Frame(80, 440, 380, 65)
    frame5.addFromList(flow_obj7_m_app_r_by_4, can)

    # Frame 5 Text

    text_approved5 = request.session.get('dpo_advice_or_overruled', 'DPO Advice')
    text_approved_5 = Paragraph(text_approved5, style=styles7['Normal'])
    flow_obj7_m_app_r_by_5.append(text_approved_5)

    frame5 = Frame(220, 390, 120, 25)
    frame5.addFromList(flow_obj7_m_app_r_by_5, can)

    # Frame 6 Text Comment box 1

    text_approved6= request.session.get('comments1', 'No Comments')
    text_approved_6 = Paragraph(text_approved6, style=styles7['Normal'])
    flow_obj7_m_app_r_by_6.append(text_approved_6)

    frame5 = Frame(80, 295, 380, 65)
    frame5.addFromList(flow_obj7_m_app_r_by_6, can)

    # Frame 7 Text

    text_approved7 = request.session.get('response_reviewed_by', 'Response')
    text_approved_7 = Paragraph(text_approved7, style=styles7['Normal'])
    flow_obj7_m_app_r_by_7.append(text_approved_7)

    frame5 = Frame(220, 260, 120, 25)
    frame5.addFromList(flow_obj7_m_app_r_by_7, can)

    # Frame 8 Text Comment box 2

    text_approved8= request.session.get('comments2', 'No Comments')
    text_approved_8 = Paragraph(text_approved8, style=styles7['Normal'])
    flow_obj7_m_app_r_by_8.append(text_approved_8)

    frame5 = Frame(80, 170, 380, 55)
    frame5.addFromList(flow_obj7_m_app_r_by_8, can)

    # Frame 9 Text

    text_approved9 = request.session.get('title_of_DPO')
    text_approved_9 = Paragraph(text_approved9, style=styles7['Normal'])
    flow_obj7_m_app_r_by_9.append(text_approved_9)

    frame5 = Frame(220, 133, 120, 25)
    frame5.addFromList(flow_obj7_m_app_r_by_9, can)

    can.save()


    new_pdf7 = PdfFileReader(packet)
    # read your existing PDF
    existing_pdf = PdfFileReader("media/Final_GDPR_report.pdf")
    output7 = PdfFileWriter()

    page = existing_pdf.getPage(6)
    page.mergePage(new_pdf7.getPage(0))
    output7.addPage(page)
    # finally, write "output" to a real file
    outputStream = open("pdf/page7.pdf", "wb")
    output7.write(outputStream)
    outputStream.close()

    pdfs = ['pdf/page1.pdf', 'pdf/page2.pdf', 'pdf/page3.pdf', 'pdf/page4.pdf', 'pdf/page5.pdf', 'pdf/page6.pdf', 'pdf/page7.pdf']

    merger = PdfMerger()

    for pdf in pdfs:
        merger.append(pdf)

    merger.write("media/final_pdf/result.pdf")
    merger.close()

    return "PDF 2.0 Generated"

    # return HttpResponse(open('media/final_pdf/result.pdf', 'rb'), content_type='application/pdf')
    # return "Pdf has been generated"

